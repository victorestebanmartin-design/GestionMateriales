# Aplicación de materiales - versión corregida
from flask import Flask, render_template_string, request, redirect, url_for, flash, jsonify, abort, send_file, make_response, session
import sqlite3, os, csv, io, logging, re, threading
from datetime import date, datetime, timedelta
from contextlib import contextmanager
from typing import Optional, List
from dataclasses import dataclass
from werkzeug.utils import secure_filename
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False
    print("Advertencia: openpyxl no está instalado. Funcionalidad Excel deshabilitada.")

# ================== Config & logging ==================
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'clave-super-secreta')  # flash & cookies

# Bases de datos separadas (usar rutas absolutas relativas al proyecto)
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_MATERIALES = os.path.join(BASE_DIR, "database", "materiales.db")
DB_OPERARIOS = os.path.join(BASE_DIR, "database", "operarios.db")
AVISO_DIAS = 7

# Evento para reinicio limpio desde el admin (lo escucha run_app_window.py)
_restart_event = threading.Event()

# Roles y credenciales por defecto (cámbialas por variables de entorno)
ADMIN_PASSWORD      = os.environ.get("ADMIN_PASSWORD", "admin123")
ALMACEN_PIN         = os.environ.get("ALMACEN_PIN", "almac123")
OPERARIO_PIN        = os.environ.get("OPERARIO_PIN", "")  # vacío = sin contraseña

# ================== Modelo ==================
@dataclass
class Material:
    id: int
    codigo: str
    caducidad: str
    estado: Optional[str] = None
    operario_numero: Optional[str] = None  # número del operario
    ean: Optional[str] = None
    descripcion: Optional[str] = None
    fecha_asignacion: Optional[str] = None  # timestamp ISO de última asignación

# ================== DB helpers ==================
@contextmanager
def get_db_materiales():
    conn = sqlite3.connect(DB_MATERIALES)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"DB error: {e}")
        raise
    finally:
        conn.close()

@contextmanager
def get_db_operarios():
    conn = sqlite3.connect(DB_OPERARIOS)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"DB error: {e}")
        raise
    finally:
        conn.close()

# Función de compatibilidad (usa materiales por defecto)
@contextmanager
def get_db():
    with get_db_materiales() as conn:
        yield conn

def init_db():
    """Inicializar ambas bases de datos creando tablas y operarios por defecto si no existen."""
    # ── materiales.db ──────────────────────────────────────────────
    with get_db_materiales() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS materiales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT UNIQUE NOT NULL,
                caducidad TEXT,
                estado TEXT DEFAULT 'precintado',
                operario_numero TEXT,
                ean TEXT,
                descripcion TEXT,
                fecha_asignacion TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ean_descriptions (
                ean TEXT PRIMARY KEY,
                descripcion TEXT NOT NULL,
                fecha_actualizacion DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)

    # ── operarios.db ───────────────────────────────────────────────
    with get_db_operarios() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS operarios (
                numero TEXT PRIMARY KEY,
                nombre TEXT NOT NULL,
                rol TEXT DEFAULT 'operario',
                activo INTEGER DEFAULT 1
            )
        """)
        # Solo crear el admin 999999 si la tabla está vacía (instalación limpia)
        c = conn.execute("SELECT COUNT(*) FROM operarios")
        if c.fetchone()[0] == 0:
            conn.execute("""
                INSERT INTO operarios (numero, nombre, rol, activo)
                VALUES ('999999', 'Administrador', 'admin', 1)
            """)

def row_to_material(r)->Material:
    return Material(**dict(r))

# ================== Manejo de Roles y Autenticación ==================
def get_operario_by_numero(numero: str):
    """Obtiene un operario por su número"""
    with get_db_operarios() as conn:
        c = conn.cursor()
        c.execute("SELECT numero, nombre, rol, activo FROM operarios WHERE numero = ? AND activo = 1", (numero,))
        row = c.fetchone()
        if row:
            return {
                'numero': row[0],
                'nombre': row[1],
                'rol': row[2],
                'activo': row[3]
            }
        return None

def authenticate_user(numero: str, pin: str = ""):
    """Autentica un usuario solo por número de operario"""
    operario = get_operario_by_numero(numero)
    if not operario:
        return None
    
    # Solo verificar que el operario esté activo
    return operario

def current_role():
    """Obtiene el rol del usuario actual desde la cookie"""
    return request.cookies.get("role", "")

def current_user():
    """Obtiene los datos del usuario actual"""
    numero = request.cookies.get("user_numero", "")
    if numero:
        return get_operario_by_numero(numero)
    return None

def require_role(*allowed_roles):
    """Decorador para requerir ciertos roles"""
    def decorator(f):
        def wrapper(*args, **kwargs):
            user_role = current_role()
            if user_role not in allowed_roles:
                flash(f"No tienes permisos para esta acción. Rol requerido: {', '.join(allowed_roles)}", "error")
                return redirect(url_for("login"))
            return f(*args, **kwargs)
        wrapper.__name__ = f.__name__
        return wrapper
    return decorator

def can_user_perform_action(action: str) -> bool:
    """Verifica si el usuario actual puede realizar una acción específica"""
    user_role = current_role()
    
    permissions = {
        'admin': ['create_user', 'edit_user', 'delete_user', 'register_materials', 'assign_materials', 'return_materials', 'retire_materials', 'waste_materials', 'import_data', 'export_data', 'view_reports', 'admin_panel'],
        'almacenero': ['register_materials', 'assign_materials', 'return_materials', 'retire_materials', 'waste_materials', 'view_reports'],
        'operario': ['assign_materials']  # Solo pueden asignarse materiales
    }
    
    return action in permissions.get(user_role, [])

# ================== Validaciones & fechas ==================
def codigo_valido(codigo: str) -> bool:
    return bool(re.fullmatch(r"\d{7}", (codigo or "").strip()))

def ean_valido(ean: str) -> bool:
    if not ean: return True
    return bool(re.fullmatch(r"\d{13}", ean.strip()))

def normalize_date_human(s: str) -> Optional[str]:
    """Acepta ddmmaa, ddmmaaaa o con separadores. Devuelve ISO YYYY-MM-DD."""
    if not s: return None
    s = s.strip()
    digits = re.sub(r"[^0-9]", "", s)
    if len(digits)==6:
        d=int(digits[0:2]); m=int(digits[2:4]); y=2000+int(digits[4:6])
    elif len(digits)==8:
        d=int(digits[0:2]); m=int(digits[2:4]); y=int(digits[4:8])
    else:
        for fmt in ("%d-%m-%y","%d/%m/%y","%d.%m.%y","%d-%m-%Y","%d/%m/%Y","%d.%m.%Y","%Y-%m-%d"):
            try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except: pass
        return None
    try:
        return date(y,m,d).strftime("%Y-%m-%d")
    except: return None

def parse_date(iso: str) -> Optional[date]:
    try: return datetime.strptime(iso, "%Y-%m-%d").date()
    except: return None

# ================== Catálogo / Operarios ==================
def get_desc(ean: str)->tuple[Optional[str], bool]:
    """Devuelve (descripción, existe_en_db)"""
    if not ean: return None, False
    with get_db() as conn:
        c=conn.cursor()
        # Primero buscar en catálogo
        c.execute("SELECT descripcion FROM catalogo WHERE ean=?", (ean,))
        r=c.fetchone()
        if r: return r[0], True
        # Si no está en catálogo, buscar en materiales
        c.execute("SELECT DISTINCT descripcion FROM materiales WHERE ean=? AND descripcion IS NOT NULL LIMIT 1", (ean,))
        r=c.fetchone()
        return r[0] if r else None, bool(r)

def upsert_desc(ean: str, descripcion: str)->bool:
    if not (ean and descripcion and ean_valido(ean)): return False
    with get_db() as conn:
        c=conn.cursor()
        c.execute("""INSERT INTO catalogo(ean,descripcion) VALUES(?,?)
                     ON CONFLICT(ean) DO UPDATE SET descripcion=excluded.descripcion""", (ean.strip(), descripcion.strip()))
        return True

def get_operario_nombre(num: str)->Optional[str]:
    if not num: return None
    with get_db_operarios() as conn:
        c=conn.cursor()
        c.execute("SELECT nombre FROM operarios WHERE numero=?", (num,))
        r=c.fetchone()
        return r["nombre"] if r else None

def get_operario_display(num: str)->str:
    """Devuelve 'numero - nombre' o solo el número si no encuentra el nombre"""
    if not num: return "-"
    nombre = get_operario_nombre(num)
    if nombre:
        return f"{num} - {nombre}"
    return num

def upsert_operario(num: str, nombre: str, rol: str = "operario", activo: int = 1)->bool:
    """Función mejorada para importación CSV - soporta rol y estado activo"""
    num=(num or "").strip(); nombre=(nombre or "").strip(); rol=(rol or "operario").strip()
    if not num or not nombre: return False
    
    # Validar rol
    valid_roles = ['operario', 'almacenero', 'admin']
    if rol.lower() not in valid_roles:
        rol = 'operario'  # Por defecto
    
    # Validar activo (1 o 0)
    try:
        activo = int(activo) if activo in [0, 1, '0', '1'] else 1
    except:
        activo = 1
    
    with get_db_operarios() as conn:
        c=conn.cursor()
        # Insertar o actualizar con todos los campos
        c.execute("""INSERT INTO operarios(numero,nombre,rol,activo) VALUES(?,?,?,?)
                     ON CONFLICT(numero) DO UPDATE SET 
                     nombre=excluded.nombre, rol=excluded.rol, activo=excluded.activo""", 
                 (num, nombre, rol.lower(), activo))
        return True

# ================== CRUD Operarios ==================
def get_all_operarios():
    """Obtiene todos los operarios con información completa"""
    with get_db_operarios() as conn:
        c = conn.cursor()
        c.execute("""SELECT numero, nombre, rol, activo 
                    FROM operarios ORDER BY numero""")
        columns = [desc[0] for desc in c.description]
        return [dict(zip(columns, row)) for row in c.fetchall()]

def get_operario_completo(numero: str):
    """Obtiene un operario completo por número"""
    with get_db_operarios() as conn:
        c = conn.cursor()
        c.execute("""SELECT numero, nombre, rol, activo 
                    FROM operarios WHERE numero = ?""", (numero,))
        row = c.fetchone()
        if row:
            columns = [desc[0] for desc in c.description]
            return dict(zip(columns, row))
        return None

def crear_operario(numero: str, nombre: str, rol: str = "operario"):
    """Crea un nuevo operario"""
    numero = numero.strip()
    nombre = nombre.strip()
    rol = rol.strip().lower()
    
    if not numero or not nombre:
        return False, "Número y nombre son obligatorios"
    
    # Validar rol
    valid_roles = ['operario', 'almacenero', 'admin']
    if rol not in valid_roles:
        return False, f"Rol debe ser uno de: {', '.join(valid_roles)}"
    
    # Verificar que no existe
    if get_operario_completo(numero):
        return False, "Ya existe un operario con ese número"
    
    try:
        with get_db_operarios() as conn:
            c = conn.cursor()
            c.execute("""INSERT INTO operarios(numero, nombre, rol, activo) 
                        VALUES(?, ?, ?, 1)""", (numero, nombre, rol))
            return True, "Operario creado exitosamente"
    except Exception as e:
        return False, f"Error al crear operario: {str(e)}"

def actualizar_operario(numero: str, nombre: str, rol: str):
    """Actualiza un operario existente"""
    nombre = nombre.strip()
    rol = rol.strip().lower()
    
    if not nombre:
        return False, "El nombre es obligatorio"
    
    # Validar rol
    valid_roles = ['operario', 'almacenero', 'admin']
    if rol not in valid_roles:
        return False, f"Rol debe ser uno de: {', '.join(valid_roles)}"
    
    try:
        with get_db_operarios() as conn:
            c = conn.cursor()
            c.execute("""UPDATE operarios SET nombre = ?, rol = ? 
                        WHERE numero = ?""", (nombre, rol, numero))
            if c.rowcount == 0:
                return False, "Operario no encontrado"
            return True, "Operario actualizado exitosamente"
    except Exception as e:
        return False, f"Error al actualizar operario: {str(e)}"

def toggle_operario_activo(numero: str):
    """Activa/desactiva un operario"""
    operario = get_operario_completo(numero)
    if not operario:
        return False, "Operario no encontrado"
    
    nuevo_estado = 0 if operario['activo'] else 1
    estado_texto = "activado" if nuevo_estado else "desactivado"
    
    try:
        with get_db_operarios() as conn:
            c = conn.cursor()
            c.execute("UPDATE operarios SET activo = ? WHERE numero = ?", 
                     (nuevo_estado, numero))
            return True, f"Operario {estado_texto} exitosamente"
    except Exception as e:
        return False, f"Error al cambiar estado: {str(e)}"

def eliminar_operario(numero: str):
    """Elimina un operario (soft delete - lo desactiva)"""
    try:
        with get_db_operarios() as conn:
            c = conn.cursor()
            # Verificar si tiene materiales asignados
            with get_db_materiales() as conn_mat:
                c_mat = conn_mat.cursor()
                c_mat.execute("SELECT COUNT(*) FROM materiales WHERE operario_numero = ?", (numero,))
                materiales_asignados = c_mat.fetchone()[0]
            
            if materiales_asignados > 0:
                return False, f"No se puede eliminar: tiene {materiales_asignados} materiales asignados"
            
            # Soft delete - desactivar
            c.execute("UPDATE operarios SET activo = 0 WHERE numero = ?", (numero,))
            if c.rowcount == 0:
                return False, "Operario no encontrado"
            return True, "Operario eliminado (desactivado) exitosamente"
    except Exception as e:
        return False, f"Error al eliminar operario: {str(e)}"

def get_estadisticas_operario(numero: str):
    """Obtiene estadísticas de un operario"""
    try:
        with get_db_materiales() as conn:
            c = conn.cursor()
            # Materiales asignados actualmente
            c.execute("SELECT COUNT(*) FROM materiales WHERE operario_numero = ?", (numero,))
            materiales_asignados = c.fetchone()[0]
            
            # Materiales por estado
            c.execute("""SELECT estado, COUNT(*) FROM materiales 
                        WHERE operario_numero = ? GROUP BY estado""", (numero,))
            estados = dict(c.fetchall())
            
            return {
                'materiales_asignados': materiales_asignados,
                'por_estado': estados
            }
    except:
        return {'materiales_asignados': 0, 'por_estado': {}}

# ================== Materiales CRUD ==================
def validar_consistencia_ean_descripcion(ean: str, desc: str) -> tuple[bool, Optional[str]]:
    """Valida que un EAN y descripción sean consistentes con los datos existentes.
    
    Retorna:
        (True, None) si es válido
        (False, mensaje_error) si hay inconsistencia
    """
    if not ean or not desc:
        return True, None  # Sin EAN o sin descripción, no hay que validar
    
    ean = ean.strip()
    desc = desc.strip()
    
    with get_db() as conn:
        c = conn.cursor()
        # Buscar si ya existe este EAN con una descripción diferente
        c.execute("""
            SELECT DISTINCT descripcion 
            FROM materiales 
            WHERE ean = ? AND descripcion != ? AND descripcion IS NOT NULL
            LIMIT 1
        """, (ean, desc))
        
        existing_desc = c.fetchone()
        if existing_desc:
            return False, f"EAN {ean} ya existe con descripción '{existing_desc[0]}'. No se puede usar '{desc}'"
    
    return True, None

def get_material(codigo: str)->Optional[Material]:
    with get_db() as conn:
        c=conn.cursor()
        c.execute("SELECT id,codigo,caducidad,estado,operario_numero,ean,descripcion,fecha_asignacion FROM materiales WHERE codigo=? LIMIT 1",(codigo,))
        r=c.fetchone()
        return row_to_material(r) if r else None

def insert_material(codigo: str, cad_raw: str, ean: Optional[str], desc: Optional[str])->bool:
    codigo=(codigo or "").strip()
    ean=(ean or "").strip()
    desc=(desc or "").strip()
    if not codigo_valido(codigo): return False
    if not ean_valido(ean): return False
    cad=normalize_date_human(cad_raw)
    if not cad: return False
    
    # Validar que la fecha de caducidad no esté ya vencida
    fecha_cad = parse_date(cad)
    if fecha_cad and fecha_cad < date.today():
        return False  # No permitir registro de productos ya caducados
    # Chequeo de duplicado
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT id FROM materiales WHERE codigo = ?", (codigo,))
        if c.fetchone():
            return False  # duplicado
    if ean and not desc:
        d = get_desc(ean)
        if d: desc = d
    
    if not desc:
        return False  # La descripción es obligatoria
        
    # Autocompletar descripción desde catálogo si existe
    if ean:
        d = get_desc(ean)[0]  # Solo nos interesa la descripción, no si existe
        if d:
            desc = d  # Usar la descripción del catálogo
        
    with get_db() as conn:
        c=conn.cursor()
        c.execute("""INSERT INTO materiales (codigo,caducidad,estado,operario_numero,ean,descripcion,fecha_asignacion)
                     VALUES (?,?,'precintado',NULL,?,?,NULL)""",(codigo,cad, ean if ean else None, desc))
    if ean and desc: upsert_desc(ean, desc)
    return True

@app.route("/api/get_descripcion_by_ean")
def api_get_descripcion_by_ean():
    ean = request.args.get('ean', '').strip()
    if not ean_valido(ean):
        return jsonify({'error': 'EAN inválido'}), 400
    desc, existe = get_desc(ean)
    return jsonify({
        'descripcion': desc,
        'existe': existe
    })

def update_material(codigo: str, cad_raw: Optional[str], ean: Optional[str], desc: Optional[str])->bool:
    if not codigo_valido(codigo): return False
    sets=[]; params=[]
    if cad_raw is not None and cad_raw!="":
        cad=normalize_date_human(cad_raw)
        if not cad: return False
        sets.append("caducidad=?"); params.append(cad)
    if ean is not None:
        if ean!="" and not ean_valido(ean): return False
        sets.append("ean=?"); params.append(ean.strip() or None)
    if desc is not None:
        sets.append("descripcion=?"); params.append(desc.strip() or None)
    if not sets: return False
    
    # Validar consistencia EAN-Descripción si ambos están siendo actualizados o ya existen
    if ean is not None and desc is not None:
        ean_final = ean.strip() if ean else None
        desc_final = desc.strip() if desc else None
        if ean_final and desc_final:
            es_valido, error_msg = validar_consistencia_ean_descripcion(ean_final, desc_final)
            if not es_valido:
                print(f"Error de consistencia: {error_msg}")
                return False  # Rechazar actualización por inconsistencia
    
    params.append(codigo)
    with get_db() as conn:
        c=conn.cursor()
        c.execute(f"UPDATE materiales SET {', '.join(sets)} WHERE codigo=?", tuple(params))
        ok = c.rowcount>0
    if ean and desc: upsert_desc(ean, desc)
    return ok

def set_estado_disponible_si_precintado(codigo: str):
    with get_db() as conn:
        c=conn.cursor()
        c.execute("UPDATE materiales SET estado='disponible' WHERE codigo=? AND LOWER(IFNULL(estado,''))='precintado'", (codigo,))
        return c.rowcount>0

def update_operario(codigo: str, operario: str) -> bool:
    """Asigna operario y graba fecha/hora del servidor en asignado_at."""
    with get_db() as conn:
        c=conn.cursor()
        c.execute("UPDATE materiales SET operario_numero=?, fecha_asignacion=datetime('now','localtime') WHERE codigo=?", (operario, codigo))
        return c.rowcount > 0

def devolver_material(codigo: str)->bool:
    with get_db() as conn:
        c=conn.cursor()
        c.execute("UPDATE materiales SET operario_numero=NULL WHERE codigo=?", (codigo,))
        return c.rowcount>0

def gastar_material(codigo: str)->bool:
    with get_db() as conn:
        c=conn.cursor()
        c.execute("UPDATE materiales SET estado='gastado', operario_numero=NULL WHERE codigo=?", (codigo,))
        return c.rowcount>0

def retirar_material(codigo: str)->bool:
    with get_db() as conn:
        c=conn.cursor()
        c.execute("UPDATE materiales SET estado='retirado', operario_numero=NULL WHERE codigo=?", (codigo,))
        return c.rowcount>0

def list_materiales_paged(estado_filter: Optional[str], q: str, offset: int, limit: int, operario_filter: str = "")->List[Material]:
    with get_db() as conn:
        c=conn.cursor()
        c.execute("SELECT id,codigo,caducidad,estado,operario_numero,ean,descripcion,fecha_asignacion FROM materiales")
        rows=[row_to_material(r) for r in c.fetchall()]

    def estado_calc(m: Material)->str:
        return estado_base(m.caducidad, m.operario_numero, m.estado)

    def vence_prox_calc(m: Material)->bool:
        cad = parse_date(m.caducidad)
        if not cad: return False
        hoy = date.today()
        return cad <= hoy + timedelta(days=AVISO_DIAS) and cad >= hoy

    def caducado_calc(m: Material)->bool:
        cad = parse_date(m.caducidad)
        if not cad: return False
        hoy = date.today()
        return cad < hoy

    # Función auxiliar para verificar si un material es precintado
    def es_precintado(material):
        eg = (material.estado or "").lower()
        return eg == "precintado" and not (material.operario_numero and str(material.operario_numero).strip())

    qn=(q or "").upper().strip()
    qop=(operario_filter or "").upper().strip()
    filtered=[]
    for m in rows:
        est=estado_calc(m)
        
        # Permitir que materiales "en uso" aparezcan en "vence prox" o "caducado" si cumplen las condiciones
        if estado_filter and estado_filter != "todos":
            if estado_filter == "precintado" and es_precintado(m):
                pass  # Incluir materiales precintados (con P al final)
            elif estado_filter == "vence prox" and vence_prox_calc(m):
                pass  # Incluir si vence pronto, incluso si está en uso
            elif estado_filter == "caducado" and caducado_calc(m):
                pass  # Incluir si está caducado, incluso si está en uso
            elif estado_filter != est:
                continue
        texto=(f"{m.codigo} {m.ean or ''} {m.descripcion or ''}").upper()
        if qn and qn not in texto: continue
        if qop:
            op_display = get_operario_display(m.operario_numero).upper()
            if qop not in op_display: continue
        filtered.append((m, est))
    filtered.sort(key=lambda t: (sort_key_estado(t[1]), t[0].caducidad, t[0].codigo))
    sliced=filtered[offset: offset+limit]
    return [t[0] for t in sliced]

# ================== Estados & visual ==================
def estado_base(caducidad: str, operario: Optional[str], estado_guardado: Optional[str]) -> str:
    eg = (estado_guardado or "").lower()
    if eg == "gastado":
        return "gastado"
    if eg == "retirado":
        return "retirado"
    if eg == "escaneado":
        return "escaneado"
    if operario and str(operario).strip():
        return "en uso"
    cad = parse_date(caducidad)
    if not cad:
        return "error fecha"
    hoy = date.today()
    if cad < hoy:
        return "caducado"
    if cad <= hoy + timedelta(days=AVISO_DIAS):
        return "vence prox"
    return "disponible"

def estado_label(caducidad: str, operario: Optional[str], estado_guardado: Optional[str]) -> str:
    eg = (estado_guardado or "").lower()
    base = estado_base(caducidad, operario, estado_guardado)
    if eg == "gastado":
        return "gastado"
    if eg == "retirado":
        return "retirado"
    if eg == "escaneado":
        return "escaneado"
    if not (operario and str(operario).strip()) and eg == "precintado":
        if base in ("disponible", "vence prox", "caducado"):
            return f"P·{base}"
        return "precintado"
    return base

def sort_key_estado(estado: str)->int:
    order={"caducado":0,"en uso":1,"vence prox":2,"disponible":3,"precintado":4,"retirado":5,"gastado":6,"escaneado":7,"error fecha":8}
    return order.get(estado, 99)

def badge_html(label: str)->str:
    base = label.replace("P·", "")
    colores = {
        "disponible":"#d8f5d0",
        "vence prox":"#fff3bf",
        "caducado":"#ffd6d6",
        "en uso":"#d7e3ff",
        "retirado":"#ffeaa7",
        "gastado":"#e9ecef",
        "escaneado":"#d1ecf1",
        "error fecha":"#f3cff3",
        "precintado":"#bee9f3",
    }
    text = {
        "disponible":"#0f5132",
        "vence prox":"#664d03",
        "caducado":"#842029",
        "en uso":"#0b3d91",
        "retirado":"#856404",
        "gastado":"#495057",
        "escaneado":"#0c5460",
        "error fecha":"#842055",
        "precintado":"#055160",
    }
    bg=colores.get(base,"#e9ecef"); fg=text.get(base,"#495057")
    pchip = " <span style='font-weight:700'>P</span>" if label.startswith("P·") else ""
    return f"<span style='padding:4px 10px;border-radius:999px;background:{bg};color:{fg};font-weight:600;font-size:.85em;white-space:nowrap;display:inline-block'>{base}{pchip}</span>"

# Las funciones de roles están definidas arriba en la sección "Manejo de Roles y Autenticación"

# ================== Init ==================
init_db()

# ================== PWA: Manifest e iconos ==================
@app.route('/manifest.json')
def pwa_manifest():
    manifest = {
        "name": "Gestión de Materiales",
        "short_name": "GestMat",
        "description": "Sistema de gestión de materiales y herramientas",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#f8f9fa",
        "theme_color": "#1a73e8",
        "icons": [
            {
                "src": "/static/icons/icon-192.png",
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any maskable"
            },
            {
                "src": "/static/icons/icon-512.png",
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any maskable"
            }
        ]
    }
    from flask import json as flask_json
    response = app.response_class(
        response=flask_json.dumps(manifest),
        mimetype='application/manifest+json'
    )
    return response

@app.route('/sw.js')
def pwa_service_worker():
    sw = (
        "self.addEventListener('install',e=>self.skipWaiting());\n"
        "self.addEventListener('activate',e=>e.waitUntil(clients.claim()));\n"
        "self.addEventListener('fetch',e=>e.respondWith(fetch(e.request).catch(()=>caches.match(e.request))));\n"
    )
    return app.response_class(
        response=sw,
        mimetype='application/javascript',
        headers={'Service-Worker-Allowed': '/'}
    )

@app.route('/favicon.ico')
def favicon():
    icon_path = os.path.join(BASE_DIR, 'static', 'icons', 'icon-192.png')
    if os.path.exists(icon_path):
        return send_file(icon_path, mimetype='image/png')
    return '', 204

# ================== API auxiliar ==================
@app.get("/api/desc_por_ean")
def api_desc_por_ean():
    ean=(request.args.get("ean") or "").strip()
    return jsonify({"descripcion": get_desc(ean) or ""})

@app.get("/api/operario_nombre")
def api_operario_nombre():
    num=(request.args.get("numero") or "").strip()
    return jsonify({"nombre": get_operario_nombre(num) or ""})

@app.route("/api/operario_add", methods=["POST"])
def api_operario_add():
    num=(request.form.get("numero") or "").strip()
    nom=(request.form.get("nombre") or "").strip()
    return jsonify({"ok": upsert_operario(num,nom)})

# Chequeo inmediato de conflicto EAN para operario/código
@app.get("/api/operario_conflicto_ean")
def api_operario_conflicto_ean():
    codigo = (request.args.get("codigo") or "").strip()
    oper_num = (request.args.get("operario_num") or "").strip()
    if not codigo_valido(codigo) or not oper_num:
        return jsonify({"ok": False, "conflicto": False})
    m = get_material(codigo)
    if not m or not m.ean:
        return jsonify({"ok": True, "conflicto": False})
    with get_db() as conn:
        c = conn.cursor()
        patron = f"{oper_num} - %"  # "num - nombre"
        c.execute("""SELECT codigo, descripcion FROM materiales
                     WHERE ean=? AND operario LIKE ? AND LOWER(IFNULL(estado,'')) NOT IN ('gastado', 'retirado', 'escaneado') AND codigo<>?""",
                  (m.ean, patron, codigo))
        r = c.fetchone()
    if r:
        return jsonify({
            "ok": True,
            "conflicto": True,
            "ean": m.ean,
            "otro_codigo": r["codigo"],
            "otra_desc": r["descripcion"] or ""
        })
    return jsonify({"ok": True, "conflicto": False})

# Chequeo de código duplicado para registro
@app.get("/api/check_codigo")
def api_check_codigo():
    codigo = (request.args.get("codigo") or "").strip()
    if not codigo_valido(codigo):
        return jsonify({"existe": False, "valido": False})
    m = get_material(codigo)
    return jsonify({"existe": bool(m), "valido": True})

# Scroll infinito: devolver filas en lotes
@app.get("/api/materiales")
def api_materiales():
    estado=request.args.get("estado","todos")
    q=request.args.get("q","")
    operario=request.args.get("operario","")
    try:
        offset=int(request.args.get("offset","0"))
        limit=int(request.args.get("limit","50"))
    except:
        offset=0; limit=50
    datos=[]
    for m in list_materiales_paged(estado,q,offset,limit,operario):
        base = estado_base(m.caducidad, m.operario_numero, m.estado)
        label = estado_label(m.caducidad, m.operario_numero, m.estado)
        asignado_at_formatted = "-"
        if m.fecha_asignacion and m.operario_numero:
          # Aceptar varios formatos de fecha existentes en BD
          dt = None
          for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f"):
            try:
              dt = datetime.strptime(m.fecha_asignacion, fmt)
              break
            except ValueError:
              pass
          if dt is None:
            try:
              dt = datetime.fromisoformat(m.fecha_asignacion.replace("Z", "+00:00"))
            except ValueError:
              dt = None
          asignado_at_formatted = dt.strftime("%d/%m/%Y %H:%M:%S") if dt else str(m.fecha_asignacion)
        
        # Determinar estado crítico para materiales en uso
        estado_critico = None
        if base == "en uso":
            cad = parse_date(m.caducidad)
            if cad:
                hoy = date.today()
                if cad < hoy:
                    estado_critico = "caducado"
                elif cad <= hoy + timedelta(days=AVISO_DIAS):
                    estado_critico = "vence prox"
        
        datos.append({
            "id": m.id,
            "codigo": m.codigo,
            "ean": m.ean or "-",
            "descripcion": m.descripcion or "-",
            "caducidad": m.caducidad,
            "estado": base,
            "estado_label": label,
            "estado_html": badge_html(label),
            "asignado_at": asignado_at_formatted,
            "operario": get_operario_display(m.operario_numero),
            "operario_numero": m.operario_numero or "",
            "estado_critico": estado_critico,
        })
    return jsonify(datos)

# ================== API de Autenticación ==================
@app.route("/api/auth", methods=["POST"])
def api_auth():
    try:
        data = request.get_json()
        numero = data.get('numero', '').strip()
        
        if not numero:
            return jsonify({'success': False, 'message': 'Número de operario requerido'})
        
        # Autenticar usuario
        operario = authenticate_user(numero)
        if not operario:
            return jsonify({'success': False, 'message': 'Número de operario no válido o usuario inactivo'})
        
        return jsonify({
            'success': True,
            'user': {
                'numero': operario['numero'],
                'nombre': operario['nombre'],
                'rol': operario['rol']
            }
        })
    
    except Exception as e:
        logger.error(f"Error en autenticación: {e}")
        return jsonify({'success': False, 'message': 'Error interno del servidor'})

# Info material para asignación (alertas vence/caduca)
@app.get("/api/info_material")
def api_info_material():
    codigo=(request.args.get("codigo") or "").strip()
    if not codigo_valido(codigo): return jsonify({"existe": False})
    m=get_material(codigo)
    if not m: return jsonify({"existe": False})
    cad=parse_date(m.caducidad); hoy=date.today()
    caducado=False; vence_prox=False
    if cad:
        caducado = cad<hoy
        vence_prox = (not caducado) and (cad <= hoy+timedelta(days=AVISO_DIAS))
    return jsonify({
        "existe": True,
        "estado": estado_base(m.caducidad, m.operario_numero, m.estado),
        "estado_label": estado_label(m.caducidad, m.operario_numero, m.estado),
        "caducidad": m.caducidad,
        "descripcion": m.descripcion or "",
        "ean": m.ean or "",
        "caducado": caducado,
        "vence_prox": vence_prox,
        "asignado_at": m.fecha_asignacion or ""
    })

def get_productos_caducados_total():
    """Obtiene TODOS los productos caducados, independientemente de su estado (gastado/retirado/etc)"""
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT codigo, descripcion, caducidad, estado, operario_numero FROM materiales")
        rows = c.fetchall()
    
    def caducado_calc(caducidad: str) -> bool:
        cad = parse_date(caducidad)
        if not cad: return False
        hoy = date.today()
        return cad < hoy
    
    caducados = []
    for r in rows:
        if caducado_calc(r["caducidad"]):
            caducados.append({
                'codigo': r['codigo'],
                'descripcion': r['descripcion'], 
                'caducidad': r['caducidad'],
                'estado': r['estado'] or 'disponible',
                'operario': r['operario_numero'] or ''
            })
    
    return caducados

# Contadores por estado (para los botones) - MEJORADO con más datos
@app.get("/api/contadores")
def api_contadores():
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT codigo, caducidad, estado, operario_numero, descripcion FROM materiales WHERE (procesado_excel IS NULL OR procesado_excel = 0)")
        rows = c.fetchall()
    from collections import Counter
    ctr = Counter()
    
    # Listas para alertas específicas
    caducados_criticos = []
    vencen_hoy = []
    vencen_manana = []
    
    def vence_prox_calc(caducidad: str) -> bool:
        cad = parse_date(caducidad)
        if not cad: return False
        hoy = date.today()
        return cad <= hoy + timedelta(days=AVISO_DIAS) and cad >= hoy

    def caducado_calc(caducidad: str) -> bool:
        cad = parse_date(caducidad)
        if not cad: return False
        hoy = date.today()
        return cad < hoy

    hoy = date.today()
    manana = hoy + timedelta(days=1)
    
    for codigo, caducidad, estado, operario_numero, descripcion in rows:
        eg = (estado or "").lower()
        cad = parse_date(caducidad)
        
        # Contar estado base
        base = estado_base(caducidad, operario_numero, estado)
        ctr[base] += 1
        
        # Detectar alertas críticas (excluir escaneados)
        if cad and eg not in ["gastado", "retirado", "escaneado"]:
            if cad < hoy:  # Caducado
                caducados_criticos.append({
                    'codigo': codigo,
                    'descripcion': descripcion or 'Sin descripción',
                    'caducidad': caducidad,
                    'dias_caducado': (hoy - cad).days,
                    'operario': get_operario_display(operario_numero) if operario_numero else None
                })
            elif cad == hoy:  # Vence hoy
                vencen_hoy.append({
                    'codigo': codigo,
                    'descripcion': descripcion or 'Sin descripción',
                    'caducidad': caducidad,
                    'operario': get_operario_display(operario_numero) if operario_numero else None
                })
            elif cad == manana:  # Vence mañana
                vencen_manana.append({
                    'codigo': codigo,
                    'descripcion': descripcion or 'Sin descripción',
                    'caducidad': caducidad,
                    'operario': get_operario_display(operario_numero) if operario_numero else None
                })
        
        # Contar "vence prox": incluir materiales que vencen pronto Y no están gastados/retirados/escaneados
        if eg not in ["gastado", "retirado", "escaneado"] and vence_prox_calc(caducidad):
            if base != "vence prox":
                ctr["vence prox"] += 1
                
        # Contar "caducado": incluir materiales caducados Y no están gastados/retirados/escaneados
        if eg not in ["gastado", "retirado", "escaneado"] and caducado_calc(caducidad):
            if base != "caducado":
                ctr["caducado"] += 1
                
        # Contar precintado específicamente
        if eg == "precintado" and not operario_numero:
            ctr["precintado"] += 1

    # Calcular métricas adicionales
    total_materiales = sum(ctr.values())
    total_activos = ctr.get("disponible", 0) + ctr.get("en uso", 0) + ctr.get("vence prox", 0) + ctr.get("precintado", 0)
    
    return jsonify({
        # Contadores básicos
        "caducado":   ctr.get("caducado", 0),
        "en uso":     ctr.get("en uso", 0),
        "vence prox": ctr.get("vence prox", 0),
        "disponible": ctr.get("disponible", 0),
        "precintado": ctr.get("precintado", 0),
        "retirado":   ctr.get("retirado", 0),
        "gastado":    ctr.get("gastado", 0),
        "escaneado":  ctr.get("escaneado", 0),
        
        # Métricas adicionales
        "total_materiales": total_materiales,
        "total_activos": total_activos,
        "porcentaje_uso": round((ctr.get("en uso", 0) / total_activos * 100) if total_activos > 0 else 0, 1),
        
        # Alertas específicas
        "alertas": {
            "caducados_criticos": caducados_criticos[:5],  # Solo los primeros 5
            "vencen_hoy": vencen_hoy,
            "vencen_manana": vencen_manana,
            "total_caducados": len(caducados_criticos),
            "total_vencen_hoy": len(vencen_hoy),
            "total_vencen_manana": len(vencen_manana)
        }
    })

@app.get("/api/verificar_consistencia_ean")
def api_verificar_consistencia_ean():
    """Endpoint para verificar consistencia EAN-Descripción"""
    with get_db() as conn:
        c = conn.cursor()
        # Buscar EANs con múltiples descripciones
        c.execute("""
            SELECT ean, COUNT(DISTINCT descripcion) as desc_count,
                   GROUP_CONCAT(DISTINCT descripcion) as descripciones,
                   COUNT(*) as total_items
            FROM materiales 
            WHERE ean IS NOT NULL AND ean != ''
            GROUP BY ean
            HAVING COUNT(DISTINCT descripcion) > 1
            ORDER BY desc_count DESC, ean
        """)
        inconsistencias = c.fetchall()
        
        # Estadísticas generales
        c.execute("SELECT COUNT(*) FROM materiales WHERE ean IS NOT NULL AND ean != ''")
        total_con_ean = c.fetchone()[0]
        
        c.execute("SELECT COUNT(DISTINCT ean) FROM materiales WHERE ean IS NOT NULL AND ean != ''")
        eans_unicos = c.fetchone()[0]
        
    inconsistencias_detalle = []
    for ean, desc_count, descripciones, total in inconsistencias:
        inconsistencias_detalle.append({
            "ean": ean,
            "descripciones_count": desc_count,
            "descripciones": descripciones.split(',') if descripciones else [],
            "total_materiales": total
        })
    
    return jsonify({
        "consistente": len(inconsistencias) == 0,
        "inconsistencias_count": len(inconsistencias),
        "inconsistencias": inconsistencias_detalle,
        "estadisticas": {
            "total_materiales_con_ean": total_con_ean,
            "eans_unicos": eans_unicos
        }
    })

# ================== Autenticación simple ==================

# No hay login obligatorio, acceso libre al dashboard
@app.get("/logout")
def logout():
  resp = redirect(url_for("home"))
  resp.delete_cookie("role")
  resp.delete_cookie("user_numero")
  resp.delete_cookie("user_name")
  flash("Sesión cerrada correctamente.", "info")
  return resp

# ================== Admin ==================
@app.route("/admin", methods=["GET","POST"])
def admin():
    # Si no es admin, pedir autenticación
    if current_role()!="admin":
        if request.method=="POST":
            numero = request.form.get("numero", "").strip()
            operario = get_operario_by_numero(numero)
            if not operario or operario['rol'] != 'admin':
                flash("Solo el usuario admin puede acceder.", "error")
                return redirect(url_for("home"))
            # Login admin: set cookie y recargar
            resp = redirect(url_for("admin"))
            resp.set_cookie("role", "admin", max_age=60*60*8, httponly=True)
            resp.set_cookie("user_numero", operario['numero'], max_age=60*60*8, httponly=True)
            resp.set_cookie("user_name", operario['nombre'], max_age=60*60*8, httponly=True)
            flash("Bienvenido admin", "success")
            return resp
        # GET: mostrar formulario
        return render_template_string("""
        <html><head><title>Admin</title></head><body style='font-family:sans-serif;background:#f8f9fa;padding:40px'>
        <div style='max-width:400px;margin:0 auto;background:#fff;padding:30px;border-radius:12px;box-shadow:0 2px 10px rgba(0,0,0,.08)'>
        <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:20px'>
        <h2 style='margin:0'>🔐 Acceso administrador</h2>
        <a href='{{ url_for("home") }}' style='text-decoration:none;color:#007bff;font-weight:bold'>🏠 Inicio</a>
        </div>
        {% with messages = get_flashed_messages(with_categories=true) %}
          {% if messages %}
            {% for cat,msg in messages %}
              <div style='padding:10px;margin:10px 0;border-radius:6px;background:#f8d7da;color:#721c24'>{{ msg }}</div>
            {% endfor %}
          {% endif %}
        {% endwith %}
        <form method='post'>
        <label>Número de usuario admin</label>
        <input name='numero' required autofocus style='width:100%;padding:12px;margin:10px 0;border-radius:8px;border:1px solid #ccc'>
        <button type='submit' style='width:100%;padding:12px;background:#007bff;color:#fff;border:none;border-radius:8px;cursor:pointer'>Entrar</button>
        </form>
        <div style='margin-top:15px;padding:10px;background:#e7f3ff;border-radius:8px;font-size:12px;color:#004085'>
          <strong>Usuarios admin:</strong> 999999, US4281
        </div>
        </div></body></html>
        """)
    
    if request.method=="POST":
        accion=request.form.get("accion","")
        if accion=="import_operarios":
            f=request.files.get("archivo")
            if not f or f.filename=="": flash("Sube CSV o Excel operarios","error"); return redirect(url_for("admin"))
            
            n=0
            errors=0
            rows_data = []
            
            # Determinar tipo de archivo
            filename = f.filename.lower()
            
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                # Procesar archivo Excel
                try:
                    from openpyxl import load_workbook
                    import tempfile
                    import os
                    
                    # Guardar archivo temporal
                    temp_dir = tempfile.gettempdir()
                    temp_path = os.path.join(temp_dir, f"temp_operarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                    f.save(temp_path)
                    
                    # Leer Excel
                    wb = load_workbook(temp_path)
                    ws = wb.active
                    
                    for row in ws.iter_rows(values_only=True):
                        if row and len([cell for cell in row if cell is not None]) >= 2:
                            # Convertir None a string vacío y limpiar
                            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
                            rows_data.append(clean_row)
                    
                    # Limpiar archivo temporal
                    try:
                        os.unlink(temp_path)
                    except:
                        pass
                        
                except Exception as e:
                    flash(f"Error leyendo archivo Excel: {e}", "error")
                    return redirect(url_for("admin"))
            
            else:
                # Procesar archivo CSV
                try:
                    # Intentar diferentes codificaciones
                    try:
                        content = f.read().decode("utf-8")
                    except UnicodeDecodeError:
                        try:
                            f.seek(0)
                            content = f.read().decode("latin-1")
                        except:
                            f.seek(0)
                            content = f.read().decode("utf-8", errors="ignore")
                    
                    # Detectar separador (coma o punto y coma)
                    delimiter = ';' if ';' in content.split('\n')[0] else ','
                    reader=csv.reader(io.StringIO(content), delimiter=delimiter)
                    rows_data = list(reader)
                    
                except Exception as e:
                    flash(f"Error leyendo archivo CSV: {e}", "error")
                    return redirect(url_for("admin"))
            
            # Procesar filas (tanto de Excel como CSV)
            for row in rows_data:
                if not row or len(row)<2: continue
                # Soportar 2, 3 o 4 columnas: numero,nombre[,rol[,activo]]
                try:
                    numero_raw = str(row[0]).strip()
                    nombre_raw = str(row[1]).strip()
                    rol = str(row[2]).strip() if len(row)>2 else "operario"
                    activo = str(row[3]).strip() if len(row)>3 else "1"
                    
                    # Saltar filas de encabezado
                    if numero_raw.lower() in ['numero', 'id', 'código', 'codigo']:
                        continue
                    
                    # Añadir prefijo "US" si no lo tiene (para tarjetas de fichaje)
                    if numero_raw and not numero_raw.upper().startswith('US'):
                        # Solo añadir US si es un número o código válido
                        if numero_raw.isdigit() or numero_raw.isalnum():
                            numero = f"US{numero_raw}"
                            logger.info(f"Añadido prefijo US: '{numero_raw}' -> '{numero}'")
                        else:
                            numero = numero_raw  # Mantener formato original si no es numérico
                    else:
                        numero = numero_raw
                    
                    # Procesar formato "APELLIDOS, NOMBRE" -> "NOMBRE APELLIDOS"
                    if ',' in nombre_raw and len(nombre_raw.split(',')) == 2:
                        partes = nombre_raw.split(',')
                        apellidos = partes[0].strip()
                        nombre_parte = partes[1].strip()
                        nombre = f"{nombre_parte} {apellidos}".strip()
                        logger.info(f"Convertido nombre: '{nombre_raw}' -> '{nombre}'")
                    else:
                        nombre = nombre_raw
                    
                    # Convertir activo a entero
                    if str(activo).lower() in ['1', 'true', 'activo', 'si', 'sí']:
                        activo = 1
                    elif str(activo).lower() in ['0', 'false', 'inactivo', 'no']:
                        activo = 0
                    else:
                        activo = 1  # Por defecto activo
                    
                    if numero and nombre and upsert_operario(numero, nombre, rol, activo): 
                        n+=1
                    else:
                        errors+=1
                except Exception as e:
                    errors+=1
                    logger.error(f"Error importando operario {row}: {e}")
            
            msg = f"Operarios importados/actualizados: {n}"
            if errors > 0:
                msg += f", errores: {errors}"
            flash(msg, "success" if n > 0 else "error")
            return redirect(url_for("admin"))
        if accion=="import_materiales":
            f=request.files.get("archivo")
            if not f or f.filename=="": flash("Sube CSV materiales","error"); return redirect(url_for("admin"))
            content=f.read().decode("utf-8",errors="ignore")
            reader=csv.reader(io.StringIO(content))
            ok=upd=err=0
            for row in reader:
                if not row: continue
                codigo=(row[0] if len(row)>0 else "").strip()
                fecha=(row[1] if len(row)>1 else "").strip()
                ean=(row[2] if len(row)>2 else "").strip()
                desc=(row[3] if len(row)>3 else "").strip()
                if not codigo_valido(codigo) or not fecha or (ean and not ean_valido(ean)): err+=1; continue
                if insert_material(codigo, fecha, ean, desc): ok+=1
                elif update_material(codigo, fecha, ean, desc): upd+=1
                else: err+=1
            flash(f"Materiales importados: {ok}, actualizados: {upd}, errores: {err}", "success"); return redirect(url_for("admin"))
        if accion=="export_cleanup":
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            import tempfile
            import os
            
            with get_db() as conn:
                c = conn.cursor()
                # Obtener materiales gastados y retirados
                c.execute("SELECT codigo, descripcion FROM materiales WHERE LOWER(estado) IN ('gastado', 'retirado') ORDER BY codigo")
                materiales = c.fetchall()
                
                if not materiales:
                    flash("No hay materiales gastados o retirados para exportar", "error")
                    return redirect(url_for("admin"))
                
                # Crear archivo Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Materiales Gastados y Retirados"
                
                # Encabezados
                ws['A1'] = "Código"
                ws['B1'] = "Descripción"
                
                # Formato de encabezados
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                ws['A1'].font = header_font
                ws['A1'].fill = header_fill
                ws['B1'].font = header_font
                ws['B1'].fill = header_fill
                
                # Datos
                for idx, (codigo, descripcion) in enumerate(materiales, start=2):
                    ws[f'A{idx}'] = codigo
                    ws[f'B{idx}'] = descripcion or "-"
                
                # Ajustar ancho de columnas
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 50
                
                # Guardar archivo temporal
                temp_dir = tempfile.gettempdir()
                filename = f"materiales_gastados_retirados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filepath = os.path.join(temp_dir, filename)
                wb.save(filepath)
                
                # Eliminar los materiales de la base de datos
                c.execute("DELETE FROM materiales WHERE LOWER(estado) IN ('gastado', 'retirado')")
                deleted_count = c.rowcount
                
                # Enviar archivo como respuesta
                from flask import send_file
                try:
                    response = send_file(filepath, as_attachment=True, download_name=filename)
                    # Programar eliminación del archivo temporal después de enviarlo
                    @response.call_on_close
                    def cleanup_temp_file():
                        try:
                            os.unlink(filepath)
                        except:
                            pass
                    
                    flash(f"Exportados y eliminados {deleted_count} materiales", "success")
                    return response
                except Exception as e:
                    logger.error(f"Error enviando archivo: {e}")
                    flash(f"Error generando archivo Excel: {e}", "error")
                    return redirect(url_for("admin"))
        if accion=="delete_material":
            codigo=(request.form.get("codigo") or "").strip()
            if not codigo_valido(codigo): flash("Código inválido","error")
            else:
                with get_db() as conn:
                    c=conn.cursor()
                    c.execute("DELETE FROM materiales WHERE codigo=?", (codigo,))
                    n=c.rowcount
                flash(f"Materiales borrados: {n}", "success" if n else "error")
            return redirect(url_for("admin"))
        if accion=="op_upsert":
            numero=(request.form.get("numero") or "").strip()
            nombre=(request.form.get("nombre") or "").strip()
            rol=(request.form.get("rol") or "operario").strip()
            activo = 1 if request.form.get("activo") == "on" else 0
            
            if not numero or not nombre:
                flash("Número y nombre son obligatorios", "error")
                return redirect(url_for("admin"))
            
            try:
                with get_db_operarios() as conn:
                    c = conn.cursor()
                    c.execute("""INSERT OR REPLACE INTO operarios 
                                (numero, nombre, rol, activo) 
                                VALUES (?, ?, ?, ?)""", 
                             (numero, nombre, rol, activo))
                flash("Operario guardado exitosamente", "success")
            except Exception as e:
                logger.error(f"Error guardando operario: {e}")
                flash("Error al guardar operario", "error")
            return redirect(url_for("admin"))
        
        if accion=="op_delete":
            numero=(request.form.get("numero") or "").strip()
            if numero in ['admin', 'almacen']:
                flash("No se pueden eliminar usuarios del sistema", "error")
                return redirect(url_for("admin"))
            
            with get_db_operarios() as conn:
                c=conn.cursor()
                c.execute("DELETE FROM operarios WHERE numero=?", (numero,))
                n=c.rowcount
            flash(f"Operarios eliminados: {n}", "success" if n else "error"); return redirect(url_for("admin"))
        
        if accion=="op_toggle":
            numero=(request.form.get("numero") or "").strip()
            if numero in ['admin', 'almacen']:
                flash("No se pueden desactivar usuarios del sistema", "error")
                return redirect(url_for("admin"))
            
            with get_db_operarios() as conn:
                c=conn.cursor()
                c.execute("UPDATE operarios SET activo = 1 - activo WHERE numero=?", (numero,))
                n=c.rowcount
            flash(f"Estado cambiado para {n} operario(s)", "success" if n else "error")
            return redirect(url_for("admin"))
        if accion=="update_ean_description":
            ean = (request.form.get("ean") or "").strip()
            nueva_desc = (request.form.get("nueva_descripcion") or "").strip()
            
            if not ean or not nueva_desc:
                flash("EAN y descripción son obligatorios", "error")
                return redirect(url_for("admin"))
            
            try:
                with get_db() as conn:
                    c = conn.cursor()
                    # Actualizar todos los materiales con este EAN
                    c.execute("UPDATE materiales SET descripcion = ? WHERE ean = ?", (nueva_desc, ean))
                    materiales_actualizados = c.rowcount
                    
                    # Actualizar también la tabla de descripciones EAN
                    c.execute("INSERT OR REPLACE INTO ean_descriptions (ean, descripcion) VALUES (?, ?)", (ean, nueva_desc))
                
                flash(f"Actualizados {materiales_actualizados} materiales con EAN {ean}", "success")
            except Exception as e:
                logger.error(f"Error actualizando EAN: {e}")
                flash("Error al actualizar descripción", "error")
            
            return redirect(url_for("admin"))

    # Obtener EANs únicos con sus descripciones y conteos
    eans_data = []
    with get_db() as conn:
        c = conn.cursor()
        c.execute("""
            SELECT ean, descripcion, COUNT(*) as cantidad
            FROM materiales 
            WHERE ean IS NOT NULL AND ean != ''
            GROUP BY ean, descripcion
            ORDER BY ean, cantidad DESC
        """)
        
        # Agrupar por EAN
        eans_dict = {}
        for row in c.fetchall():
            ean = row['ean']
            if ean not in eans_dict:
                eans_dict[ean] = {
                    'ean': ean,
                    'descripcion_principal': row['descripcion'],
                    'total_materiales': 0,
                    'descripciones': []
                }
            
            eans_dict[ean]['total_materiales'] += row['cantidad']
            eans_dict[ean]['descripciones'].append({
                'descripcion': row['descripcion'],
                'cantidad': row['cantidad']
            })
        
        eans_data = list(eans_dict.values())
        eans_data.sort(key=lambda x: x['total_materiales'], reverse=True)

    with get_db_operarios() as conn:
        c=conn.cursor()
        c.execute("SELECT numero, nombre, rol, activo FROM operarios ORDER BY numero")
        columns = ['numero', 'nombre', 'rol', 'activo']
        ops = [dict(zip(columns, row)) for row in c.fetchall()]
    
    return render_template_string(tpl_admin(), operarios=ops, eans_data=eans_data)

# ================== Exportar/Importar Materiales ==================
@app.route('/admin/exportar_materiales')
def exportar_materiales():
    """Exportar base de datos de materiales a Excel"""
    if current_role() != "admin":
        flash("Acceso denegado", "error")
        return redirect(url_for("home"))
    
    if not EXCEL_DISPONIBLE:
        flash("❌ Funcionalidad Excel no disponible. Instale openpyxl.", "error")
        return redirect('/admin')
    
    try:
        conn = sqlite3.connect(DB_MATERIALES)
        cursor = conn.cursor()
        
        # Obtener todos los materiales con información completa
        cursor.execute("""
            SELECT 
                id, codigo, ean, descripcion, caducidad, estado, 
                operario_numero, fecha_asignacion, fecha_registro, fecha_registro
            FROM materiales 
            ORDER BY fecha_registro DESC
        """)
        
        materiales = cursor.fetchall()
        conn.close()
        
        # Crear archivo Excel en memoria
        wb = Workbook()
        ws = wb.active
        ws.title = "Materiales"
        
        # Definir encabezados con estilo
        encabezados = [
            'ID', 'Código', 'EAN', 'Descripción', 'Caducidad', 'Estado',
            'Operario', 'Asignado En', 'Creado En', 'Actualizado En'
        ]
        
        # Escribir encabezados con formato
        for col, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col, value=encabezado)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Escribir datos
        for fila, material in enumerate(materiales, 2):
            for col, valor in enumerate(material, 1):
                ws.cell(row=fila, column=col, value=valor)
        
        # Ajustar ancho de columnas
        columnas_ancho = [5, 12, 15, 30, 12, 12, 15, 18, 18, 18]
        for col, ancho in enumerate(columnas_ancho, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = ancho
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta HTTP
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=materiales_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        flash(f"✅ Exportación Excel completada: {len(materiales)} materiales descargados", "success")
        return response
        
    except Exception as e:
        flash(f"❌ Error en exportación Excel: {str(e)}", "error")
        return redirect('/admin')

@app.route('/admin/importar_materiales', methods=['POST'])
def importar_materiales():
    """Importar materiales desde archivo Excel o CSV"""
    if current_role() != "admin":
        flash("Acceso denegado", "error")
        return redirect(url_for("home"))
    
    if 'archivo' not in request.files:
        flash('No se seleccionó ningún archivo', 'error')
        return redirect('/admin')
    
    archivo = request.files['archivo']
    if archivo.filename == '':
        flash('No se seleccionó ningún archivo', 'error')
        return redirect('/admin')
    
    extension = archivo.filename.lower().split('.')[-1]
    if extension not in ['csv', 'xlsx', 'xls']:
        flash('Solo se permiten archivos CSV, XLS o XLSX', 'error')
        return redirect('/admin')
    
    try:
        materiales_data = []
        
        if extension == 'csv':
            # Leer CSV
            contenido = archivo.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(contenido))
            materiales_data = list(csv_reader)
            
        elif extension in ['xlsx', 'xls']:
            # Leer Excel
            if not EXCEL_DISPONIBLE:
                flash("❌ Funcionalidad Excel no disponible. Instale openpyxl.", "error")
                return redirect('/admin')
            
            # Guardar archivo temporalmente
            temp_path = os.path.join(os.path.dirname(__file__), f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{extension}")
            archivo.save(temp_path)
            
            try:
                wb = openpyxl.load_workbook(temp_path)
                ws = wb.active
                
                # Obtener encabezados (primera fila)
                encabezados = []
                for col in range(1, ws.max_column + 1):
                    valor = ws.cell(row=1, column=col).value
                    if valor:
                        encabezados.append(str(valor).strip())
                    else:
                        break
                
                # Leer datos
                for fila in range(2, ws.max_row + 1):
                    fila_data = {}
                    tiene_datos = False
                    
                    for col, encabezado in enumerate(encabezados, 1):
                        valor = ws.cell(row=fila, column=col).value
                        if valor is not None:
                            fila_data[encabezado] = str(valor).strip()
                            tiene_datos = True
                        else:
                            fila_data[encabezado] = ''
                    
                    if tiene_datos:
                        materiales_data.append(fila_data)
                
            finally:
                # Limpiar archivo temporal
                try:
                    os.unlink(temp_path)
                except:
                    pass
        
        # Procesar datos
        conn = sqlite3.connect(DB_MATERIALES)
        cursor = conn.cursor()
        
        materiales_importados = 0
        errores = []
        
        for fila_num, fila in enumerate(materiales_data, start=2):
            try:
                # Validar campos obligatorios
                codigo = fila.get('Código', '').strip()
                ean = fila.get('EAN', '').strip() 
                descripcion = fila.get('Descripción', '').strip()
                estado = fila.get('Estado', 'disponible').strip()
                
                if not codigo or not descripcion:
                    errores.append(f"Fila {fila_num}: Código y Descripción son obligatorios")
                    continue
                
                # Validar consistencia EAN-Descripción si hay EAN
                if ean:
                    valido, error_msg = validar_consistencia_ean_descripcion(ean, descripcion)
                    if not valido:
                        errores.append(f"Fila {fila_num}: {error_msg}")
                        continue
                
                # Preparar datos
                caducidad = fila.get('Caducidad', '').strip()
                operario = fila.get('Operario', '').strip() or None
                
                # Insertar material
                print(f"DEBUG: Insertando material {codigo} con operario_numero: {operario}")
                cursor.execute("""
                    INSERT INTO materiales 
                    (codigo, ean, descripcion, caducidad, estado, operario_numero, fecha_registro)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    codigo, ean or None, descripcion, caducidad or None, estado, 
                    operario, datetime.now().isoformat()
                ))
                
                materiales_importados += 1
                
            except Exception as e:
                errores.append(f"Fila {fila_num}: Error procesando - {str(e)}")
        
        conn.commit()
        conn.close()
        
        # Mostrar resultados
        if materiales_importados > 0:
            tipo_archivo = "Excel" if extension in ['xlsx', 'xls'] else "CSV"
            mensaje = f'✅ Importación {tipo_archivo} completada: {materiales_importados} materiales importados'
            if errores:
                mensaje += f' ({len(errores)} errores encontrados)'
            flash(mensaje, 'success')
            
            # Mostrar errores si los hay (máximo 5)
            if errores:
                for error in errores[:5]:
                    flash(f'⚠️ {error}', 'warning')
                if len(errores) > 5:
                    flash(f'... y {len(errores)-5} errores más', 'warning')
        else:
            flash('❌ No se importaron materiales', 'error')
            for error in errores[:3]:
                flash(f'🔥 {error}', 'error')
    
    except Exception as e:
        flash(f'❌ Error procesando archivo: {str(e)}', 'error')
    
    return redirect('/admin')

@app.route('/admin/borrar_materiales', methods=['POST'])
def borrar_todos_materiales():
    """Borrar todos los materiales de la base de datos"""
    if current_role() != "admin":
        flash("Acceso denegado", "error")
        return redirect(url_for("home"))
    
    # Verificar confirmación
    confirmacion = request.form.get('confirmacion', '').strip().upper()
    if confirmacion != 'BORRAR':
        flash("❌ Operación cancelada. Debes escribir 'BORRAR' para confirmar.", "error")
        return redirect('/admin')
    
    try:
        with get_db() as conn:
            c = conn.cursor()
            # Contar materiales antes de borrar
            c.execute("SELECT COUNT(*) FROM materiales")
            total_materiales = c.fetchone()[0]
            
            if total_materiales == 0:
                flash("ℹ️ La base de datos de materiales ya está vacía.", "info")
                return redirect('/admin')
            
            # Borrar todos los materiales
            c.execute("DELETE FROM materiales")
            conn.commit()
            
            # Borrar también las descripciones EAN huérfanas si existen
            c.execute("DELETE FROM ean_descriptions WHERE ean NOT IN (SELECT DISTINCT ean FROM materiales WHERE ean IS NOT NULL)")
            conn.commit()
            
            flash(f"✅ Base de datos limpiada exitosamente. Se eliminaron {total_materiales} materiales.", "success")
            logger.info(f"Admin borró todos los materiales: {total_materiales} registros eliminados")
            
    except Exception as e:
        flash(f"❌ Error al limpiar la base de datos: {str(e)}", "error")
        logger.error(f"Error borrando materiales: {str(e)}")
    
    return redirect('/admin')

# ================== Vistas por estado ==================
@app.get("/estado/<estado>")
def vista_estado(estado):
    if estado not in ["precintado","disponible","vence prox","caducado","en uso","retirado","gastado","escaneado"]:
        abort(404)
    return render_template_string(tpl_estado(), estado=estado)

# ================== Navegación entre aplicaciones ==================
@app.route("/switch/herramientas")
def switch_to_herramientas():
    """Cambia a la aplicación de herramientas"""
    session['app_origen'] = 'materiales'
    return redirect("http://localhost:5001")

# ================== Home ==================
@app.route("/", methods=["GET","POST"])
def home():
    # Acceso libre, sin login obligatorio
    role=current_role() or ""

    if request.method=="POST":
        accion=request.form.get("accion","")
        codigo=(request.form.get("codigo") or "").strip()

        # Registrar
        if accion=="registrar":
            if not require_role(["almacenero","admin"]): return redirect(url_for("home"))
            cad_raw=(request.form.get("caducidad") or "").strip()
            ean=(request.form.get("ean") or "").strip()
            desc=(request.form.get("descripcion") or "").strip()
            if not codigo_valido(codigo):
                flash("Código interno inválido (7 dígitos).","error")
                return redirect(url_for("home"))
            elif ean and not ean_valido(ean):
                flash("EAN inválido (debe tener exactamente 13 dígitos).","error")
                return redirect(url_for("home"))
            elif get_material(codigo):  # Verificación de duplicado
                flash(f"El código {codigo} ya existe. No se puede registrar.","error")
                return redirect(url_for("home"))
            else:
                # Validar fecha de caducidad antes de intentar insertar
                cad_normalizada = normalize_date_human(cad_raw)
                if cad_normalizada:
                    fecha_cad = parse_date(cad_normalizada)
                    if fecha_cad and fecha_cad < date.today():
                        flash("No se puede registrar: la fecha de caducidad ya ha vencido.","error")
                        return redirect(url_for("home"))
                
                if insert_material(codigo, cad_raw, ean, desc):
                    flash(f"Material {codigo} registrado (PRECINTADO).","success")
                    return redirect(url_for("home"))
                else:
                    flash("No se pudo registrar. Revisa datos (fecha inválida o descripción faltante).","error")
                    return redirect(url_for("home"))

        # Asignar directo con restricciones
        elif accion=="asignar_directo":
            if not require_role(["operario","almacenero","admin"]): return redirect(url_for("home"))
            oper_num=(request.form.get("operario_num") or "").strip()
            confirmado=(request.form.get("confirmado") or "")=="1"
            if not codigo_valido(codigo):
                flash("Código interno inválido (7 dígitos).","error")
            elif not oper_num:
                flash("Nº de operario obligatorio.","error")
            else:
                m=get_material(codigo)
                if not m:
                    flash("El código no existe. Regístralo primero.","error")
                else:
                    cad=parse_date(m.caducidad); hoy=date.today()
                    caducado = cad<hoy if cad else False
                    vence_prox = (not caducado) and (cad and cad<=hoy+timedelta(days=AVISO_DIAS))
                    if caducado:
                        flash("No se puede asignar: material CADUCADO.","error"); return redirect(url_for("home"))
                    if vence_prox and not confirmado:
                        flash(f"Atención: vence pronto ({m.caducidad}). Confirma para asignar.","warning"); return redirect(url_for("home"))

                    nombre=get_operario_nombre(oper_num)
                    if not nombre:
                        flash("Operario inexistente. Añádelo primero.","error"); return redirect(url_for("home"))

                    # Restricción: mismo operario + mismo EAN (no gastados/retirados/escaneados)
                    if m.ean:
                        with get_db() as conn:
                            c=conn.cursor()
                            c.execute("""SELECT COUNT(1) FROM materiales 
                                         WHERE ean=? AND operario_numero=? AND LOWER(IFNULL(estado,'')) NOT IN ('gastado', 'retirado', 'escaneado') AND codigo<>?""",
                                      (m.ean, oper_num, codigo))
                            cnt=c.fetchone()[0]
                        if cnt>0:
                            flash("No puedes asignarte este producto: ya tienes otro con el mismo EAN. Devuélvelo primero.", "error")
                            return redirect(url_for("home"))

                    if update_operario(codigo, f"{oper_num} - {nombre}"):
                        set_estado_disponible_si_precintado(codigo)
                        flash(f"Material {codigo} asignado a {oper_num} - {nombre}", "success")
                    else:
                        flash("No se pudo asignar el material.","error")
            return redirect(url_for("home"))

        # Devolver
        elif accion in ("devolver_rapido","devolver"):
            if not require_role(["almacenero","admin"]): return redirect(url_for("home"))
            if not codigo_valido(codigo):
                flash("Código interno inválido (7 dígitos).","error")
            elif devolver_material(codigo):
                flash(f"Material {codigo} devuelto.","success")
            else:
                flash("Error al devolver el material.","error")
            return redirect(url_for("home"))

        # Gastado
        elif accion in ("gastado_rapido","gastar"):
            if not require_role(["almacenero","admin"]): return redirect(url_for("home"))
            if not codigo_valido(codigo):
                flash("Código interno inválido (7 dígitos).","error")
            elif gastar_material(codigo):
                flash(f"Material {codigo} marcado como gastado.","success")
            else:
                flash("Error al marcar como gastado.","error")
            return redirect(url_for("home"))

        # Retirado
        elif accion in ("retirado_rapido","retirar"):
            if not require_role(["almacenero","admin"]): return redirect(url_for("home"))
            if not codigo_valido(codigo):
                flash("Código interno inválido (7 dígitos).","error")
            elif retirar_material(codigo):
                flash(f"Material {codigo} marcado como retirado.","success")
            else:
                flash("Error al marcar como retirado.","error")
            return redirect(url_for("home"))

    return render_template_string(tpl_home(), role=role)

# ================== Templates ==================
def tpl_login():
    return """
<!doctype html><html><head><meta charset="utf-8"><title>Login</title>
<style>
body{font-family:Segoe UI,Roboto,Arial,sans-serif;background:#f8f9fa;margin:0;padding:40px}
.card{max-width:480px;margin:0 auto;background:#fff;border-radius:12px;box-shadow:0 2px 10px rgba(0,0,0,.08);padding:20px}
h1{text-align:center;margin:0 0 12px}
label{display:block;margin-top:8px}
select,input,button{width:100%;padding:12px;border:1px solid #ced4da;border-radius:10px;font-size:16px}
button{background:#007bff;border-color:#007bff;color:#fff;margin-top:12px;cursor:pointer}
button:hover{background:#0069d9}
.alert{padding:10px;border-radius:8px;margin:10px 0}
.alert-error{background:#f8d7da;border:1px solid #f5c6cb;color:#721c24}
.small{color:#6c757d;text-align:center}
</style></head><body>
<div class="card">
  <h1>🔒 Acceso</h1>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}{% for cat,msg in messages %}
      {% if cat == 'error' %}
        <script>
          document.addEventListener('DOMContentLoaded', function() {
            mostrarDialogoError('{{ msg|e }}');
          });
        </script>
      {% else %}
        <div class="alert alert-{{ cat }}">{{ msg }}</div>
      {% endif %}
    {% endfor %}{% endif %}
  {% endwith %}
  
  <script>
    function mostrarDialogoError(mensaje) {
        const overlay = document.createElement('div');
        overlay.id = 'error-dialog-overlay';
        overlay.style.cssText = `
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0,0,0,0.5);
            z-index: 10000;
            display: flex;
            align-items: center;
            justify-content: center;
        `;
        
        const dialog = document.createElement('div');
        dialog.style.cssText = `
            background: white;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 4px 8px rgba(0,0,0,0.3);
            max-width: 400px;
            text-align: center;
            font-family: inherit;
        `;
        
        dialog.innerHTML = `
            <div style="color: #721c24; font-size: 18px; margin-bottom: 15px;">
                ❌ Error
            </div>
            <div style="color: #721c24; margin-bottom: 20px; line-height: 1.4;">
                ${mensaje}
            </div>
            <button id="error-dialog-btn" onclick="cerrarDialogoError()" style="
                background: #dc3545;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                cursor: pointer;
                font-size: 14px;
            ">Aceptar</button>
        `;
        
        overlay.appendChild(dialog);
        document.body.appendChild(overlay);
        
        // Enfocar el botón y añadir evento de teclado
        const button = document.getElementById('error-dialog-btn');
        button.focus();
        
        // Cerrar con Enter o Escape
        const handleKeydown = (e) => {
            if (e.key === 'Enter' || e.key === 'Escape') {
                cerrarDialogoError();
                document.removeEventListener('keydown', handleKeydown);
            }
        };
        
        document.addEventListener('keydown', handleKeydown);
    }

    function cerrarDialogoError() {
        const overlay = document.getElementById('error-dialog-overlay');
        if (overlay) {
            overlay.remove();
        }
    }
  </script>
  
  <form method="POST">
    <label>Número de Operario</label>
    <input type="text" name="numero" placeholder="Ej: 999999, US4281, US272..." required autofocus>
    <button type="submit">Entrar</button>
  </form>
  <div class="small" style="margin-top: 15px; padding: 10px; background: #e7f3ff; border-radius: 8px;">
    <strong>Acceso:</strong><br>
    • <strong>999999</strong> o <strong>US4281</strong> (Administrador)<br>
    • <strong>Almaceneros:</strong> US272, US25013<br>
    • <strong>Operarios:</strong> Usar número de tarjeta
</div>
</div>
</body></html>
"""

def tpl_admin():
    return """<!doctype html><html lang="es"><head>
<meta charset="utf-8"><title>Admin – Gestión de Materiales</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
*{box-sizing:border-box}
body{font-family:'Segoe UI',system-ui,sans-serif;margin:0;background:#f1f5f9;color:#1e293b;min-height:100vh}
.topbar{background:#0f172a;color:#fff;padding:0 24px;height:54px;display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:200;box-shadow:0 2px 12px rgba(0,0,0,.3)}
.topbar-logo{font-size:16px;font-weight:700;flex:1;color:#f8fafc;letter-spacing:-.2px}
.topbar-logo em{color:#60a5fa;font-style:normal}
.topbar a{color:#94a3b8;text-decoration:none;padding:6px 14px;border-radius:6px;font-size:13px;font-weight:500;transition:background .15s,color .15s}
.topbar a:hover{background:rgba(255,255,255,.12);color:#f8fafc}
.page{padding:20px 24px;max-width:1900px;margin:0 auto}
.alert{padding:12px 16px;border-radius:10px;margin-bottom:16px;font-size:13px}
.alert-success{background:#f0fdf4;border:1px solid #bbf7d0;color:#166534}
.alert-error{background:#fef2f2;border:1px solid #fecaca;color:#991b1b}
.card{background:#fff;border-radius:14px;box-shadow:0 1px 4px rgba(0,0,0,.07),0 2px 12px rgba(0,0,0,.04);padding:22px;margin-bottom:20px}
.card-head{display:flex;align-items:center;justify-content:space-between;padding-bottom:14px;margin-bottom:16px;border-bottom:1px solid #f1f5f9;gap:12px}
.card-title{font-size:15px;font-weight:700;color:#0f172a;display:flex;align-items:center;gap:8px;margin:0;flex-shrink:0}
.tiles{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:16px;margin-bottom:20px}
.tile{background:#fff;border-radius:14px;padding:18px 16px;box-shadow:0 1px 4px rgba(0,0,0,.07);display:flex;flex-direction:column;gap:10px;border-top:3px solid #e2e8f0}
.tile.amber{border-top-color:#f59e0b}.tile.cyan{border-top-color:#06b6d4}
.tile.emerald{border-top-color:#10b981}.tile.rose{border-top-color:#f43f5e}.tile.indigo{border-top-color:#6366f1}.tile.excel{border-top-color:#217346}
.tile-title{font-size:13px;font-weight:700;color:#1e293b}
.tile-desc{font-size:11px;color:#94a3b8;line-height:1.5;flex:1}
.row2{display:grid;grid-template-columns:3fr 2fr;gap:20px;margin-bottom:20px;align-items:start}
.btn{display:inline-flex;align-items:center;gap:6px;padding:8px 16px;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;border:none;text-decoration:none;line-height:1.2;transition:filter .12s;white-space:nowrap}
.btn:hover{filter:brightness(.9)}
.btn-primary{background:#3b82f6;color:#fff}.btn-success{background:#22c55e;color:#fff}
.btn-warning{background:#f59e0b;color:#fff}.btn-danger{background:#ef4444;color:#fff}
.btn-info{background:#06b6d4;color:#fff}.btn-secondary{background:#64748b;color:#fff}
.btn-ghost{background:#fff;border:1.5px solid #e2e8f0;color:#475569}
.btn-ghost:hover{background:#f8fafc}
.btn-sm{padding:6px 12px;font-size:12px}.btn-full{width:100%;justify-content:center}
.btn-row{display:flex;gap:8px;flex-wrap:wrap;align-items:center}
.fg{margin-bottom:12px}
.fg label{display:block;font-size:12px;font-weight:600;color:#374151;margin-bottom:4px}
.fg input,.fg select{width:100%;padding:9px 12px;border:1.5px solid #e2e8f0;border-radius:8px;font-size:13px;font-family:inherit;background:#fafafa;color:#1e293b;transition:border .15s}
.fg input:focus,.fg select:focus{outline:none;border-color:#3b82f6;background:#fff}
input[type=file]{padding:5px 8px;background:#fafafa;border:1.5px solid #e2e8f0;border-radius:8px;font-size:12px;width:100%}
.info-box{background:#f8fafc;border-left:3px solid #3b82f6;padding:10px 14px;border-radius:0 8px 8px 0;font-size:12px;color:#475569;line-height:1.7;margin-top:8px}
.info-box.danger{border-left-color:#ef4444;background:#fef2f2;color:#7f1d1d}
details>summary{cursor:pointer;font-size:12px;color:#3b82f6;font-weight:600;padding:4px 0;user-select:none}
details[open]>summary{color:#1d4ed8}
.code-block{background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;padding:12px;font-family:monospace;font-size:11px;color:#475569;line-height:1.8;margin-top:8px}
.table-wrap{border:1px solid #f1f5f9;border-radius:10px;overflow:hidden}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#f8fafc;color:#64748b;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.4px;padding:10px 12px;border-bottom:2px solid #e2e8f0;text-align:left;white-space:nowrap}
td{padding:10px 12px;border-bottom:1px solid #f8fafc;vertical-align:middle}
tr:last-child td{border-bottom:none}
tr:hover td{background:#fafcff}
code{background:#f1f5f9;color:#475569;padding:2px 6px;border-radius:4px;font-family:monospace;font-size:11px}
.badge{display:inline-block;padding:3px 10px;border-radius:99px;font-size:11px;font-weight:600}
.badge-ok{background:#dcfce7;color:#166534}.badge-warn{background:#fee2e2;color:#991b1b}
.badge-blue{background:#dbeafe;color:#1e40af}
.badge-red{background:#fee2e2;color:#991b1b}.badge-orange{background:#ffedd5;color:#9a3412}
.danger-zone{border:1.5px solid #fca5a5;border-radius:10px;background:#fef2f2;padding:18px;margin-top:16px}
.danger-zone-title{font-size:13px;font-weight:700;color:#991b1b;margin:0 0 8px}
.gh-card{background:linear-gradient(135deg,#0f172a,#1e3a5f);color:#f8fafc;border-radius:14px;padding:22px;margin-bottom:20px}
.gh-card p{font-size:12px;color:#94a3b8;margin:6px 0 14px}
.modal-ov{display:none;position:fixed;inset:0;background:rgba(15,23,42,.55);z-index:500}
.modal-ov .modal-box{position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);background:#fff;border-radius:16px;padding:28px;max-width:500px;width:92%;box-shadow:0 20px 60px rgba(0,0,0,.2)}
.modal-ov .modal-box h3{font-size:16px;font-weight:700;margin:0 0 20px;color:#0f172a}
hr.div{border:none;border-top:1px solid #f1f5f9;margin:16px 0}
</style></head><body>

<nav class="topbar">
  <span class="topbar-logo">⚙️ Panel de <em>Administración</em></span>
  <a href="{{ url_for('home') }}">← Inicio</a>
  <a href="{{ url_for('logout') }}">Cerrar sesión</a>
</nav>

<main class="page">
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}{% for cat,msg in messages %}
      {% if cat == 'error' %}
        <script>document.addEventListener('DOMContentLoaded',function(){mostrarDialogoError('{{ msg|e }}');});</script>
      {% else %}
        <div class="alert alert-{{ cat }}">{{ msg }}</div>
      {% endif %}
    {% endfor %}{% endif %}
  {% endwith %}

  <script>
    function mostrarDialogoError(mensaje) {
        const overlay = document.createElement('div');
        overlay.id = 'error-dialog-overlay';
        overlay.style.cssText = `
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0,0,0,0.5);
            z-index: 10000;
            display: flex;
            align-items: center;
            justify-content: center;
        `;
        
        const dialog = document.createElement('div');
        dialog.style.cssText = `
            background: white;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 4px 8px rgba(0,0,0,0.3);
            max-width: 400px;
            text-align: center;
            font-family: inherit;
        `;
        
        dialog.innerHTML = `
            <div style="color: #721c24; font-size: 18px; margin-bottom: 15px;">
                ❌ Error
            </div>
            <div style="color: #721c24; margin-bottom: 20px; line-height: 1.4;">
                ${mensaje}
            </div>
            <button id="error-dialog-btn" onclick="cerrarDialogoError()" style="
                background: #dc3545;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                cursor: pointer;
                font-size: 14px;
            ">Aceptar</button>
        `;
        
        overlay.appendChild(dialog);
        document.body.appendChild(overlay);
        
        const button = document.getElementById('error-dialog-btn');
        button.focus();
        
        const handleKeydown = (e) => {
            if (e.key === 'Enter' || e.key === 'Escape') {
                cerrarDialogoError();
                document.removeEventListener('keydown', handleKeydown);
            }
        };
        
        document.addEventListener('keydown', handleKeydown);
    }

    function cerrarDialogoError() {
        const overlay = document.getElementById('error-dialog-overlay');
        if (overlay) { overlay.remove(); }
    }
  </script>

  <!-- ════ QUICK TILES ════ -->
  <div class="tiles">
    <div class="tile" style="border-top-color:#0ea5e9">
      <div class="tile-title">📂 Carpeta Caducidades</div>
      <div class="tile-desc">Gestión de productos perecederos</div>
      <a href="file:///T:/Compartir/Produccion/GESTI%C3%93N%20CADUCIDADES/"
         class="btn btn-full btn-sm"
         style="background:#0ea5e9;color:#fff;justify-content:center"
         title="T:\\Compartir\\Produccion\\GESTI\u00d3N CADUCIDADES">📁 Abrir carpeta</a>
      <a href="file:///T:/Compartir/Produccion/GESTI%C3%93N%20CADUCIDADES/Gesti%C3%B3n%20de%20productos%20perecederos%20(19.09.2022).xlsx"
         class="btn btn-full btn-sm btn-ghost"
         style="justify-content:center;margin-top:4px;font-size:11px"
         title="Gesti\u00f3n de productos perecederos (19.09.2022)">📊 Abrir Excel</a>
      <hr style="border:none;border-top:1px solid #e2e8f0;margin:6px 0">
      <div style="font-size:10px;color:#64748b;margin-bottom:3px">O copia la ruta:</div>
      <div style="display:flex;gap:4px;align-items:center">
        <code style="font-size:10px;background:#f1f5f9;padding:3px 6px;border-radius:4px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">T:\\Compartir\\Produccion\\GESTI\u00d3N CADUCIDADES</code>
        <button id="btn-cp-carpeta"
                data-ruta="T:\\Compartir\\Produccion\\GESTIÓN CADUCIDADES"
                onclick="navigator.clipboard.writeText(this.dataset.ruta).then(()=>{var b=this;b.textContent='\u2705';setTimeout(()=>b.textContent='\U0001F4CB',1500)})"
                style="padding:3px 7px;border:1.5px solid #e2e8f0;border-radius:5px;background:#fff;cursor:pointer;font-size:12px;flex-shrink:0">📋</button>
      </div>
      <div style="display:flex;gap:4px;align-items:center;margin-top:4px">
        <code style="font-size:10px;background:#f1f5f9;padding:3px 6px;border-radius:4px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">Gestión de productos perecederos (19.09.2022).xlsx</code>
        <button id="btn-cp-excel"
                data-ruta="T:\\Compartir\\Produccion\\GESTIÓN CADUCIDADES\\Gestión de productos perecederos (19.09.2022).xlsx"
                onclick="navigator.clipboard.writeText(this.dataset.ruta).then(()=>{var b=this;b.textContent='\u2705';setTimeout(()=>b.textContent='\U0001F4CB',1500)})"
                style="padding:3px 7px;border:1.5px solid #e2e8f0;border-radius:5px;background:#fff;cursor:pointer;font-size:12px;flex-shrink:0">📋</button>
      </div>
    </div>
    <div class="tile emerald">
      <div class="tile-title">📈 Exportar y Limpiar</div>
      <div class="tile-desc">Exporta gastados/retirados a Excel y los elimina de la BD</div>
      <form method="POST" onsubmit="return confirm('¿Exportar a Excel y eliminar todos los gastados/retirados?')">
        <input type="hidden" name="accion" value="export_cleanup">
        <button type="submit" class="btn btn-success btn-full btn-sm">📥 Exportar + Limpiar</button>
      </form>
    </div>
    <div class="tile rose">
      <div class="tile-title">🗑️ Eliminar Material</div>
      <div class="tile-desc">Borra un material concreto por código de 7 dígitos</div>
      <form method="POST" onsubmit="return confirm('¿Eliminar el material?')" style="display:flex;gap:6px">
        <input type="hidden" name="accion" value="delete_material">
        <input type="text" name="codigo" placeholder="0000000" maxlength="7" required
               style="flex:1;padding:6px 8px;border:1.5px solid #e2e8f0;border-radius:6px;font-size:12px;width:0;min-width:0">
        <button type="submit" class="btn btn-danger btn-sm">Borrar</button>
      </form>
    </div>
    <div class="tile indigo">
      <div class="tile-title">✅ Dados de Baja</div>
      <div class="tile-desc" id="count-bajas">Cargando…</div>
      <button onclick="mostrarSeccionBajas()" class="btn btn-primary btn-full btn-sm">📋 Ver Historial</button>
    </div>
    <div class="tile excel">
      <div class="tile-title">📊 Procesar Bajas en Excel</div>
      <div class="tile-desc" id="count-pendientes-excel">Cargando…</div>
      <button id="btn-ejecutar-excel" onclick="ejecutarBajasExcel()" class="btn btn-success btn-full btn-sm">▶️ En este servidor</button>
      <pre id="excel-output" style="display:none;margin-top:6px;background:#f1f5f9;border-radius:6px;padding:8px;font-size:11px;max-height:100px;overflow-y:auto;white-space:pre-wrap;word-break:break-all;color:#1e293b"></pre>
      <hr style="border:none;border-top:1px solid #e2e8f0;margin:2px 0">
      <div style="display:flex;align-items:center;gap:5px">
        <span id="agente-badge" style="flex-shrink:0;display:inline-block;width:8px;height:8px;border-radius:50%;background:#94a3b8"></span>
        <span id="agente-estado-texto" style="font-size:10px;color:#64748b">Agente desconectado</span>
      </div>
      <button id="btn-enviar-agente" onclick="enviarAlAgente()" class="btn btn-info btn-full btn-sm">📡 Enviar al PC cliente</button>
      <button id="btn-cancelar-agente" onclick="cancelarSolicitudAgente()" class="btn btn-ghost btn-full btn-sm" style="display:none;font-size:11px">✖ Cancelar solicitud</button>
      <pre id="agente-output" style="display:none;margin-top:6px;background:#f1f5f9;border-radius:6px;padding:8px;font-size:11px;max-height:100px;overflow-y:auto;white-space:pre-wrap;word-break:break-all;color:#1e293b"></pre>
    </div>
    <div class="tile" style="border-top-color:#8b5cf6">
      <div class="tile-title">🛠️ Configurar Agente</div>
      <div class="tile-desc">Descarga los archivos e instrucciones para instalar el agente en el PC cliente</div>
      <button onclick="document.getElementById('seccion-agente-setup').scrollIntoView({behavior:'smooth'})" class="btn btn-full btn-sm" style="background:#8b5cf6;color:#fff">📖 Ver Instrucciones</button>
      <a href="/admin/descargar_agente_zip" class="btn btn-ghost btn-full btn-sm" style="font-size:11px">⬇️ Descargar ZIP</a>
    </div>
  </div>

  <!-- ════ 2-COL ROW ════ -->
  <div class="row2">

    <!-- ─── COL IZQUIERDA: Operarios ─── -->
    <div class="card">
      <div class="card-head">
        <h2 class="card-title">👷 Gestión de Operarios</h2>
        <div class="btn-row">
          <button onclick="mostrarModalCrear()" class="btn btn-success btn-sm">➕ Nuevo</button>
          <button onclick="cargarOperarios()" class="btn btn-ghost btn-sm">🔄</button>
          <button onclick="exportarOperarios()" class="btn btn-ghost btn-sm">📤 CSV</button>
        </div>
      </div>

      <details style="margin-bottom:16px">
        <summary>📂 Importación masiva desde CSV / Excel</summary>
        <div style="padding:14px 0 4px">
          <form method="POST" enctype="multipart/form-data" style="display:flex;gap:10px;align-items:flex-end;flex-wrap:wrap">
            <input type="hidden" name="accion" value="import_operarios">
            <div class="fg" style="flex:1;min-width:200px;margin:0">
              <label>Archivo CSV / Excel</label>
              <input type="file" name="archivo" accept=".csv,.xlsx,.xls" required>
            </div>
            <button type="submit" class="btn btn-info">📂 Importar</button>
          </form>
          <div class="info-box">
            Columnas: <code>numero</code> <code>nombre</code> <code>rol</code> <code>activo</code> ·
            Roles: operario, almacenero, admin · Activo: 1 / 0
          </div>
        </div>
      </details>

      <div id="tablaOperarios">
        <div style="text-align:center;padding:24px;color:#94a3b8;font-size:14px">🔄 Cargando operarios…</div>
      </div>

      <hr class="div">
      <div class="info-box" style="font-size:11px">
        <strong>Permisos:</strong>
        <strong>Admin</strong> – acceso total ·
        <strong>Almacenero</strong> – registrar, asignar, devolver, retirar, gastar ·
        <strong>Operario</strong> – solo asignar materiales
      </div>
    </div><!-- /col izquierda -->

    <!-- ─── COL DERECHA ─── -->
    <div>
      <!-- Base de datos materiales -->
      <div class="card">
        <div class="card-head">
          <h2 class="card-title">📊 Base de Datos de Materiales</h2>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:16px">
          <div>
            <div style="font-size:12px;font-weight:700;color:#166534;margin-bottom:6px">📤 Exportar</div>
            <p style="font-size:11px;color:#64748b;margin-bottom:8px;line-height:1.4">Descarga todos los materiales en Excel con formato profesional.</p>
            <a href="/admin/exportar_materiales" class="btn btn-success btn-full btn-sm">⬇️ Descargar Excel (.xlsx)</a>
          </div>
          <div>
            <div style="font-size:12px;font-weight:700;color:#1e40af;margin-bottom:6px">📥 Importar</div>
            <form method="POST" action="/admin/importar_materiales" enctype="multipart/form-data">
              <div class="fg" style="margin-bottom:6px">
                <label>Archivo (.xlsx / .csv)</label>
                <input type="file" name="archivo" accept=".xlsx,.xls,.csv" required>
              </div>
              <button type="submit" class="btn btn-primary btn-full btn-sm">⬆️ Subir e Importar</button>
            </form>
          </div>
        </div>

        <details>
          <summary>📋 Formato de archivos esperado</summary>
          <div class="code-block">
            <strong>Columnas (primera fila):</strong><br>
            Código · EAN · Descripción · Caducidad · Estado · Operario<br><br>
            Código y Descripción son obligatorios. Caducidad: YYYY-MM-DD.
          </div>
        </details>

        <hr class="div">

        <div class="danger-zone">
          <div class="danger-zone-title">🗑️ Limpiar toda la base de datos</div>
          <p style="font-size:12px;color:#7f1d1d;margin:0 0 12px;line-height:1.4">
            ⚠️ Elimina <strong>TODOS</strong> los materiales permanentemente. No se puede deshacer.
          </p>
          <form method="POST" action="/admin/borrar_materiales"
                onsubmit="return confirm('¿Eliminar TODOS los materiales?\\n\\nEsta acción NO SE PUEDE DESHACER.')">
            <div class="fg" style="margin-bottom:8px">
              <label style="color:#991b1b">Escribe <code>BORRAR</code> para confirmar</label>
              <input type="text" name="confirmacion" placeholder="BORRAR" required
                     style="font-family:monospace;border-color:#fca5a5">
            </div>
            <button type="submit" class="btn btn-danger btn-full btn-sm">🗑️ ELIMINAR TODOS LOS MATERIALES</button>
          </form>
        </div>
      </div><!-- /card db -->

      <!-- GitHub Update -->
      <div class="gh-card">
        <div class="card-title" style="color:#f8fafc">🔄 Actualización desde GitHub</div>
        <p>Descarga los últimos cambios y actualiza las dependencias automáticamente.</p>
        <div class="btn-row">
          <button id="btn-update" onclick="actualizarDesdeGitHub()" class="btn btn-success">🔄 Actualizar</button>
          <button id="btn-restart" onclick="reiniciarApp()" style="display:none" class="btn btn-warning">♻️ Reiniciar app</button>
        </div>
        <pre id="update-output"
             style="display:none;margin-top:14px;background:#020617;color:#a3e635;padding:14px;border-radius:8px;font-size:12px;white-space:pre-wrap;max-height:260px;overflow-y:auto;border:1px solid #1e3a5f"></pre>
      </div><!-- /gh-card -->
    </div><!-- /col derecha -->

  </div><!-- /row2 -->

  <!-- ════ EAN CATÁLOGO ════ -->
  <div class="card">
    <div class="card-head">
      <h2 class="card-title">🏷️ Catálogo EAN — Descripciones</h2>
      <span style="font-size:12px;color:#94a3b8">Consistencia de descripciones por código EAN</span>
    </div>

    {% if eans_data %}
    <div class="table-wrap">
      <table>
        <thead><tr>
          <th>EAN</th><th>Descripción principal</th>
          <th style="text-align:center">Materiales</th>
          <th>Variantes</th>
          <th style="text-align:center">Estado</th>
          <th style="text-align:center">Acción</th>
        </tr></thead>
        <tbody>
        {% for ean_info in eans_data %}
        <tr>
          <td><code>{{ ean_info.ean }}</code></td>
          <td><strong>{{ ean_info.descripcion_principal }}</strong></td>
          <td style="text-align:center"><span class="badge badge-blue">{{ ean_info.total_materiales }}</span></td>
          <td>{% for desc in ean_info.descripciones %}
            <div style="font-size:11px;color:#64748b">• {{ desc.descripcion }} ({{ desc.cantidad }})</div>
            {% endfor %}</td>
          <td style="text-align:center">
            {% if ean_info.descripciones|length > 1 %}
              <span class="badge badge-warn">⚠️ Inconsistente</span>
            {% else %}
              <span class="badge badge-ok">✅ OK</span>
            {% endif %}
          </td>
          <td style="text-align:center">
            <button onclick="editarEAN('{{ ean_info.ean }}','{{ ean_info.descripcion_principal }}')"
                    class="btn btn-ghost btn-sm">📝 Editar</button>
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>

    <!-- Modal EAN -->
    <div id="modalEAN" class="modal-ov">
      <div class="modal-box">
        <h3>📝 Editar Descripción EAN</h3>
        <form method="POST">
          <input type="hidden" name="accion" value="update_ean_description">
          <input type="hidden" name="ean" id="modal_ean">
          <div class="fg">
            <label>EAN</label>
            <code id="modal_ean_display" style="display:block;padding:4px 0;font-size:14px"></code>
          </div>
          <div class="fg">
            <label>Nueva descripción</label>
            <input type="text" name="nueva_descripcion" id="modal_descripcion" required>
          </div>
          <div class="btn-row" style="margin-top:8px">
            <button type="submit" class="btn btn-primary">💾 Actualizar todos</button>
            <button type="button" onclick="cerrarModalEAN()" class="btn btn-ghost">Cancelar</button>
          </div>
        </form>
      </div>
    </div>

    <script>
    function editarEAN(ean, desc) {
      document.getElementById('modal_ean').value = ean;
      document.getElementById('modal_ean_display').textContent = ean;
      document.getElementById('modal_descripcion').value = desc;
      document.getElementById('modalEAN').style.display = 'block';
    }
    function cerrarModalEAN() { document.getElementById('modalEAN').style.display = 'none'; }
    document.getElementById('modalEAN').addEventListener('click', function(e) {
      if (e.target === this) cerrarModalEAN();
    });
    </script>

    {% else %}
    <p style="color:#94a3b8;font-style:italic;font-size:13px">No hay EANs registrados en la base de datos.</p>
    {% endif %}
  </div><!-- /ean card -->

</main><!-- /page -->

<!-- ════ MODAL OPERARIO ════ -->
<div id="modalOperario" class="modal-ov">
  <div class="modal-box">
    <h3 id="modalTitulo">➕ Crear Nuevo Operario</h3>
    <form id="formOperario">
      <div class="fg">
        <label>Número de operario</label>
        <input type="text" id="operarioNumero" required placeholder="Ej: 001, US123…">
        <small style="color:#94a3b8;font-size:11px">Identificador único del operario</small>
      </div>
      <div class="fg">
        <label>Nombre completo</label>
        <input type="text" id="operarioNombre" required placeholder="Nombre y apellidos">
      </div>
      <div class="fg">
        <label>Rol</label>
        <select id="operarioRol" required>
          <option value="operario">👷 Operario</option>
          <option value="almacenero">📦 Almacenero</option>
          <option value="admin">⚙️ Administrador</option>
        </select>
      </div>
      <div id="estadisticasOperario" style="display:none;padding:12px;background:#f8fafc;border-radius:8px;margin-bottom:12px">
        <div style="font-size:12px;font-weight:600;margin-bottom:6px;color:#475569">📊 Estadísticas</div>
        <div id="statsContent"></div>
      </div>
      <div class="btn-row" style="justify-content:flex-end;margin-top:8px">
        <button type="button" onclick="cerrarModal()" class="btn btn-ghost">Cancelar</button>
        <button type="submit" id="btnGuardar" class="btn btn-primary">💾 Guardar</button>
      </div>
    </form>
  </div>
</div>

<script>
// ================== CRUD de Operarios ==================
let modoEdicion = false;
let operarioOriginal = '';

async function cargarOperarios() {
  try {
    const response = await fetch('/api/operarios');
    const data = await response.json();
    
    if (data.operarios) {
      mostrarTablaOperarios(data.operarios);
    } else {
      document.getElementById('tablaOperarios').innerHTML = 
        '<div style="text-align:center;padding:20px;color:#dc3545">❌ Error al cargar operarios</div>';
    }
  } catch (error) {
    console.error('Error:', error);
    document.getElementById('tablaOperarios').innerHTML = 
      '<div style="text-align:center;padding:20px;color:#dc3545">❌ Error de conexión</div>';
  }
}

function mostrarTablaOperarios(operarios) {
  let html = `
    <table style="width:100%;border-collapse:collapse;margin-top:10px">
      <tr style="background:#f8f9fa">
        <th style="padding:12px;border:1px solid #dee2e6">Nº</th>
        <th style="padding:12px;border:1px solid #dee2e6">Nombre</th>
        <th style="padding:12px;border:1px solid #dee2e6">Rol</th>
        <th style="padding:12px;border:1px solid #dee2e6">Estado</th>
        <th style="padding:12px;border:1px solid #dee2e6">Materiales</th>
        <th style="padding:12px;border:1px solid #dee2e6">Info</th>
        <th style="padding:12px;border:1px solid #dee2e6">Acciones</th>
      </tr>`;

  operarios.forEach(op => {
    const estadoColor = op.activo ? '#d4edda' : '#ffebee';
    const estadoTexto = op.activo ? '✅ Activo' : '❌ Inactivo';
    
    let rolColor = '#f3e5f5'; // operario
    if (op.rol === 'admin') rolColor = '#e3f2fd';
    if (op.rol === 'almacenero') rolColor = '#fff3e0';
    
    const materialesInfo = op.materiales_asignados || 0;
    
    html += `
      <tr style="background:${estadoColor}">
        <td style="padding:10px;border:1px solid #dee2e6"><strong>${op.numero}</strong></td>
        <td style="padding:10px;border:1px solid #dee2e6">${op.nombre}</td>
        <td style="padding:10px;border:1px solid #dee2e6">
          <span style="padding:4px 8px; border-radius:4px; font-size:12px; background:${rolColor}">
            ${op.rol.charAt(0).toUpperCase() + op.rol.slice(1)}
          </span>
        </td>
        <td style="padding:10px;border:1px solid #dee2e6">${estadoTexto}</td>
        <td style="padding:10px;border:1px solid #dee2e6;text-align:center">
          <span style="background:#e7f3ff;padding:2px 6px;border-radius:3px;font-size:11px">
            📦 ${materialesInfo}
          </span>
        </td>
        <td style="padding:10px;border:1px solid #dee2e6;font-size:11px;color:#666">
          <span style="background:#e9ecef;padding:2px 6px;border-radius:3px;font-size:10px">
            👤 ID: ${op.numero}
          </span>
        </td>
        <td style="padding:10px;border:1px solid #dee2e6">
          <div style="display:flex;gap:4px;flex-wrap:wrap">
            <button onclick="editarOperario('${op.numero}')" 
                    style="font-size:10px;padding:4px 8px;background:#ffc107;border:none;border-radius:3px;cursor:pointer"
                    title="Editar">
              ✏️
            </button>
            <button onclick="toggleOperario('${op.numero}')" 
                    style="font-size:10px;padding:4px 8px;background:${op.activo ? '#dc3545' : '#28a745'};color:white;border:none;border-radius:3px;cursor:pointer"
                    title="${op.activo ? 'Desactivar' : 'Activar'}">
              ${op.activo ? '🔒' : '🔓'}
            </button>
            <button onclick="eliminarOperario('${op.numero}', '${op.nombre}')" 
                    style="font-size:10px;padding:4px 8px;background:#6c757d;color:white;border:none;border-radius:3px;cursor:pointer"
                    title="Eliminar">
              🗑️
            </button>
          </div>
        </td>
      </tr>`;
  });

  html += '</table>';
  document.getElementById('tablaOperarios').innerHTML = html;
}

function mostrarModalCrear() {
  modoEdicion = false;
  operarioOriginal = '';
  document.getElementById('modalTitulo').textContent = '➕ Crear Nuevo Operario';
  document.getElementById('operarioNumero').value = '';
  document.getElementById('operarioNombre').value = '';
  document.getElementById('operarioRol').value = 'operario';
  document.getElementById('operarioNumero').disabled = false;
  document.getElementById('btnGuardar').textContent = '💾 Crear';
  document.getElementById('estadisticasOperario').style.display = 'none';
  document.getElementById('modalOperario').style.display = 'block';
}

async function editarOperario(numero) {
  try {
    const response = await fetch(`/api/operarios/${numero}`);
    const operario = await response.json();
    
    if (operario.error) {
      alert('Error: ' + operario.error);
      return;
    }
    
    modoEdicion = true;
    operarioOriginal = numero;
    document.getElementById('modalTitulo').textContent = '✏️ Editar Operario';
    document.getElementById('operarioNumero').value = operario.numero;
    document.getElementById('operarioNombre').value = operario.nombre;
    document.getElementById('operarioRol').value = operario.rol;
    document.getElementById('operarioNumero').disabled = true;
    document.getElementById('btnGuardar').textContent = '💾 Guardar Cambios';
    
    // Mostrar estadísticas
    if (operario.materiales_asignados !== undefined) {
      let statsHtml = `<div>📦 Materiales asignados: <strong>${operario.materiales_asignados}</strong></div>`;
      if (operario.por_estado) {
        Object.entries(operario.por_estado).forEach(([estado, cantidad]) => {
          statsHtml += `<div style="font-size:11px;margin-top:3px">• ${estado}: ${cantidad} materiales</div>`;
        });
      }
      document.getElementById('statsContent').innerHTML = statsHtml;
      document.getElementById('estadisticasOperario').style.display = 'block';
    }
    
    document.getElementById('modalOperario').style.display = 'block';
  } catch (error) {
    alert('Error al cargar datos del operario');
    console.error(error);
  }
}

async function toggleOperario(numero) {
  if (!confirm(`¿Cambiar el estado de activación del operario ${numero}?`)) return;
  
  try {
    const response = await fetch(`/api/operarios/${numero}/toggle`, {
      method: 'POST'
    });
    const result = await response.json();
    
    if (result.success) {
      alert(result.mensaje);
      cargarOperarios();
    } else {
      alert('Error: ' + result.mensaje);
    }
  } catch (error) {
    alert('Error de conexión');
    console.error(error);
  }
}

async function eliminarOperario(numero, nombre) {
  if (!confirm(`¿Está seguro de eliminar al operario "${nombre}" (${numero})?\n\nEsta acción lo desactivará permanentemente.`)) return;
  
  try {
    const response = await fetch(`/api/operarios/${numero}`, {
      method: 'DELETE'
    });
    const result = await response.json();
    
    if (result.success) {
      alert(result.mensaje);
      cargarOperarios();
    } else {
      alert('Error: ' + result.mensaje);
    }
  } catch (error) {
    alert('Error de conexión');
    console.error(error);
  }
}

function cerrarModal() {
  document.getElementById('modalOperario').style.display = 'none';
}

// Manejar envío del formulario
document.getElementById('formOperario').addEventListener('submit', async function(e) {
  e.preventDefault();
  
  const numero = document.getElementById('operarioNumero').value.trim();
  const nombre = document.getElementById('operarioNombre').value.trim();
  const rol = document.getElementById('operarioRol').value;
  
  if (!numero || !nombre) {
    alert('Número y nombre son obligatorios');
    return;
  }
  
  try {
    let url, method, data;
    
    if (modoEdicion) {
      url = `/api/operarios/${operarioOriginal}`;
      method = 'PUT';
      data = { nombre, rol };
    } else {
      url = '/api/operarios';
      method = 'POST';
      data = { numero, nombre, rol };
    }
    
    const response = await fetch(url, {
      method: method,
      headers: {
        'Content-Type': 'application/json'
      },
      body: JSON.stringify(data)
    });
    
    const result = await response.json();
    
    if (result.success) {
      alert(result.mensaje);
      cerrarModal();
      cargarOperarios();
    } else {
      alert('Error: ' + result.mensaje);
    }
  } catch (error) {
    alert('Error de conexión');
    console.error(error);
  }
});

// Cerrar modal al hacer clic fuera
document.getElementById('modalOperario').addEventListener('click', function(e) {
  if (e.target === this) {
    cerrarModal();
  }
});

async function exportarOperarios() {
  try {
    const response = await fetch('/api/operarios');
    const data = await response.json();
    
    if (data.operarios) {
      // Crear CSV
      let csvContent = "Número,Nombre,Rol,Estado,Materiales Asignados\\n";
      
      data.operarios.forEach(op => {
        const estado = op.activo ? 'Activo' : 'Inactivo';
        const materiales = op.materiales_asignados || 0;
        csvContent += `"${op.numero}","${op.nombre}","${op.rol}","${estado}","${materiales}"\\n`;
      });
      
      // Descargar archivo
      const blob = new Blob([csvContent], { type: 'text/csv;charset=utf-8;' });
      const link = document.createElement('a');
      const url = URL.createObjectURL(blob);
      link.setAttribute('href', url);
      link.setAttribute('download', `operarios_${new Date().toISOString().split('T')[0]}.csv`);
      link.style.visibility = 'hidden';
      document.body.appendChild(link);
      link.click();
      document.body.removeChild(link);
      
      alert('✅ Lista de operarios exportada exitosamente');
    } else {
      alert('❌ Error al exportar operarios');
    }
  } catch (error) {
    alert('❌ Error de conexión al exportar');
    console.error(error);
  }
}

// Cargar operarios al cargar la página
// Rellenar IP del servidor e inicializar badge de la sección setup
async function inicializarSeccionAgente() {
  // IP del servidor
  const ipTexto = document.getElementById('ip-servidor-texto');
  if (ipTexto) {
    ipTexto.textContent = window.location.host;
  }
}

document.addEventListener('DOMContentLoaded', function() {
  cargarOperarios();
  cargarContadorBajas();
  cargarPendientesExcel();
  cargarEstadoAgente();
  inicializarSeccionAgente();
  setInterval(cargarEstadoAgente, 8000);
  setInterval(() => {
    // Sincronizar badge de la sección setup con el estado real del agente
    const badgeTile = document.getElementById('agente-badge');
    const badgeSetup = document.getElementById('agente-setup-badge');
    const estadoSetup = document.getElementById('agente-setup-estado');
    if (badgeTile && badgeSetup) {
      badgeSetup.style.background = badgeTile.style.background;
      const online = badgeTile.style.background.includes('34c55') || badgeTile.style.background === 'rgb(34, 197, 94)';
      estadoSetup.textContent = online ? 'Agente conectado y escuchando' : 'Agente desconectado — ejecuta AGENTE_EXCEL.bat en el PC cliente';
      estadoSetup.style.color = online ? '#166534' : '#475569';
    }
  }, 2000);
});

// ── Actualización desde GitHub ─────────────────────────────────
async function actualizarDesdeGitHub() {
  const btn = document.getElementById('btn-update');
  const output = document.getElementById('update-output');
  const section = document.getElementById('update-section');
  
  btn.disabled = true;
  btn.textContent = '⏳ Actualizando…';
  output.style.display = 'block';
  output.textContent = 'Conectando con GitHub…';

  try {
    const resp = await fetch('/api/admin/update', { method: 'POST' });
    const data = await resp.json();

    output.textContent = data.output || data.mensaje || '(sin respuesta)';

    if (data.success) {
      if (data.hubo_cambios) {
        btn.textContent = '✅ Actualizado';
        document.getElementById('btn-restart').style.display = 'inline-block';
      } else {
        btn.textContent = '✅ Ya estás al día';
        btn.disabled = false;
      }
    } else {
      btn.textContent = '❌ Error — Reintentar';
      btn.disabled = false;
    }
  } catch(e) {
    output.textContent = 'Error de conexión: ' + e.message;
    btn.textContent = '❌ Error — Reintentar';
    btn.disabled = false;
  }
}

async function reiniciarApp() {
  const btn = document.getElementById('btn-restart');
  const output = document.getElementById('update-output');
  btn.disabled = true;
  btn.textContent = '⏳ Reiniciando…';
  
  try {
    await fetch('/api/admin/restart', { method: 'POST' });
  } catch(e) { /* Se espera que la conexión se corte */ }

  output.textContent += '\\n\\nServidor reiniciando… esperando que vuelva.';

  // Sondear hasta que el servidor responda, luego recargar
  let intentos = 0;
  const maxIntentos = 30; // hasta ~15 segundos
  const intervalo = setInterval(async () => {
    intentos++;
    try {
      const r = await fetch('/api/hora_servidor', { cache: 'no-store' });
      if (r.ok) {
        clearInterval(intervalo);
        output.textContent += '\\n✅ Servidor listo — recargando…';
        setTimeout(() => { window.location.reload(); }, 500);
      }
    } catch(e) {
      output.textContent = output.textContent.replace(/\\.+$/, '') + '.'.repeat(intentos % 4 + 1);
    }
    if (intentos >= maxIntentos) {
      clearInterval(intervalo);
      output.textContent += '\\n⚠️ Tardando más de lo esperado. Recarga la página manualmente.';
      btn.textContent = '🔄 Recargar';
      btn.disabled = false;
      btn.onclick = () => window.location.reload();
    }
  }, 500);
}

// ── Dados de Baja ──────────────────────────────────────────────
let _bajasSectionVisible = false;

async function cargarContadorBajas() {
  try {
    const r = await fetch('/api/bajas');
    const d = await r.json();
    const n = d.total || 0;
    document.getElementById('count-bajas').textContent =
      n === 0 ? 'Sin registros de bajas' :
      n === 1 ? '1 material dado de baja' :
      `${n} materiales dados de baja`;
  } catch {
    document.getElementById('count-bajas').textContent = 'Error al cargar';
  }
}

function mostrarSeccionBajas() {
  const sec = document.getElementById('seccion-bajas');
  if (!_bajasSectionVisible) {
    sec.style.display = 'block';
    _bajasSectionVisible = true;
    cargarTablaBajas();
    sec.scrollIntoView({ behavior: 'smooth', block: 'start' });
  } else {
    sec.scrollIntoView({ behavior: 'smooth', block: 'start' });
  }
}

async function cargarTablaBajas(filtro) {
  const tbody = document.getElementById('bajas-tbody');
  tbody.innerHTML = '<tr><td colspan="5" style="text-align:center;padding:18px;color:#64748b">Cargando…</td></tr>';
  try {
    const r = await fetch('/api/bajas');
    const d = await r.json();
    let rows = d.bajas || [];
    if (filtro) {
      const f = filtro.toLowerCase();
      rows = rows.filter(b =>
        (b.codigo || '').toLowerCase().includes(f) ||
        (b.descripcion || '').toLowerCase().includes(f) ||
        (b.operario_numero || '').toLowerCase().includes(f)
      );
    }
    if (rows.length === 0) {
      tbody.innerHTML = '<tr><td colspan="5" style="text-align:center;padding:18px;color:#64748b">No hay registros</td></tr>';
      return;
    }
    tbody.innerHTML = rows.map(b => `
      <tr>
        <td style="font-family:monospace;font-weight:600">${b.codigo || '—'}</td>
        <td>${b.descripcion || '—'}</td>
        <td><span class="badge badge-${b.estado_original === 'gastado' ? 'red' : 'orange'}">${b.estado_original || '—'}</span></td>
        <td>${b.operario_numero || '—'}</td>
        <td style="font-size:12px;white-space:nowrap">${b.fecha_baja || '—'}</td>
      </tr>`).join('');
  } catch {
    tbody.innerHTML = '<tr><td colspan="5" style="text-align:center;padding:18px;color:#ef4444">Error al cargar</td></tr>';
  }
}

// ── Procesar Bajas en Excel ───────────────────────────────────
async function cargarPendientesExcel() {
  try {
    const r = await fetch('/api/bajas_pendientes_excel');
    const d = await r.json();
    const n = (d.pendientes || []).length;
    document.getElementById('count-pendientes-excel').textContent =
      n === 0 ? 'Sin pendientes de procesar' :
      n === 1 ? '1 pendiente de procesar en Excel' :
      `${n} pendientes de procesar en Excel`;
  } catch {
    document.getElementById('count-pendientes-excel').textContent = 'Error al cargar';
  }
}

async function ejecutarBajasExcel() {
  const desc = document.getElementById('count-pendientes-excel').textContent;
  if (!confirm(`¿Ejecutar el proceso de bajas en Excel?\n\n${desc}\n\nAsegúrate de que el archivo Excel con la macro DAR_DE_BAJA esté abierto EN ESTE SERVIDOR.`)) return;
  const btn = document.getElementById('btn-ejecutar-excel');
  const output = document.getElementById('excel-output');
  btn.disabled = true;
  btn.textContent = '⏳ Procesando…';
  output.style.display = 'block';
  output.textContent = 'Iniciando proceso…';
  try {
    const r = await fetch('/api/admin/ejecutar_bajas_excel', { method: 'POST' });
    const d = await r.json();
    output.textContent = d.salida || '(sin salida)';
    if (d.success) {
      btn.textContent = '✅ Completado';
      setTimeout(() => { btn.disabled = false; btn.textContent = '▶️ En este servidor'; }, 4000);
      cargarPendientesExcel();
      cargarContadorBajas();
    } else {
      btn.textContent = '❌ Error — Reintentar';
      btn.disabled = false;
    }
  } catch(e) {
    output.textContent = 'Error de conexión: ' + e.message;
    btn.textContent = '❌ Error — Reintentar';
    btn.disabled = false;
  }
}

// ── Agente Cliente Excel ─────────────────────────────────
let _agentePollingInterval = null;

async function cargarEstadoAgente() {
  try {
    const r = await fetch('/api/admin/estado_solicitud_cliente');
    const d = await r.json();
    const badge   = document.getElementById('agente-badge');
    const texto   = document.getElementById('agente-estado-texto');
    const btnEnv  = document.getElementById('btn-enviar-agente');
    const btnCan  = document.getElementById('btn-cancelar-agente');
    const output  = document.getElementById('agente-output');
    if (d.agente_online) {
      badge.style.background = '#22c55e';
      texto.textContent = 'Agente conectado';
    } else {
      badge.style.background = '#94a3b8';
      texto.textContent = 'Agente desconectado';
    }
    const estado = d.estado || 'idle';
    if (estado === 'idle') {
      btnEnv.disabled = false; btnEnv.textContent = '📡 Enviar al PC cliente';
      btnCan.style.display = 'none'; output.style.display = 'none';
      _detenerPollingAgente();
    } else if (estado === 'pendiente') {
      btnEnv.disabled = true; btnEnv.textContent = '⏳ Esperando agente…';
      btnCan.style.display = 'inline-flex';
      output.style.display = 'block';
      output.textContent = 'Solicitud enviada. Esperando que el agente la recoja…';
      _iniciarPollingAgente();
    } else if (estado === 'procesando') {
      btnEnv.disabled = true; btnEnv.textContent = '⚙️ Procesando…';
      btnCan.style.display = 'inline-flex'; btnCan.textContent = '✖ Detener proceso';
      output.style.display = 'block';
      output.textContent = 'El agente está procesando las bajas en Excel…';
      _iniciarPollingAgente();
    } else if (estado === 'completado') {
      btnEnv.disabled = false; btnEnv.textContent = '✅ Completado — Volver a enviar';
      btnCan.style.display = 'none';
      output.style.display = 'block'; output.textContent = d.salida || '(sin salida)';
      cargarPendientesExcel(); cargarContadorBajas();
      _detenerPollingAgente();
    } else if (estado === 'error') {
      btnEnv.disabled = false; btnEnv.textContent = '❌ Error — Reintentar';
      btnCan.style.display = 'none';
      output.style.display = 'block'; output.textContent = d.salida || 'Error desconocido';
      _detenerPollingAgente();
    } else if (estado === 'cancelado') {
      btnEnv.disabled = false; btnEnv.textContent = '📡 Enviar al PC cliente';
      btnCan.style.display = 'none';
      output.style.display = 'block'; output.textContent = d.salida || 'Proceso detenido por el admin.';
      cargarPendientesExcel();
      _detenerPollingAgente();
    }
  } catch(e) { console.error('Estado agente error:', e); }
}

function _iniciarPollingAgente() {
  if (!_agentePollingInterval)
    _agentePollingInterval = setInterval(cargarEstadoAgente, 3000);
}
function _detenerPollingAgente() {
  if (_agentePollingInterval) { clearInterval(_agentePollingInterval); _agentePollingInterval = null; }
}

async function enviarAlAgente() {
  const desc = document.getElementById('count-pendientes-excel').textContent;
  if (!confirm(`¿Enviar solicitud al agente cliente?\n\n${desc}\n\nEl script baja_excel_agente.py debe estar corriendo en el PC que tiene Excel abierto.`)) return;
  try {
    const r = await fetch('/api/admin/solicitar_bajas_cliente', { method: 'POST' });
    const d = await r.json();
    if (!d.success) { alert('❌ ' + (d.mensaje || 'Error al enviar solicitud')); return; }
    cargarEstadoAgente();
    _iniciarPollingAgente();
  } catch(e) { alert('❌ Error de conexión: ' + e.message); }
}

async function cancelarSolicitudAgente() {
  if (!confirm('¿Cancelar la solicitud pendiente?')) return;
  try {
    await fetch('/api/admin/cancelar_solicitud_cliente', { method: 'POST' });
    cargarEstadoAgente();
  } catch(e) { alert('Error: ' + e.message); }
}
</script>

<!-- ════ SECCIÓN DADOS DE BAJA ════ -->
<div id="seccion-bajas" class="card" style="display:none;margin-top:24px">
  <div class="card-head" style="flex-wrap:wrap;gap:10px">
    <h2 class="card-title">✅ Historial de Bajas</h2>
    <div class="btn-row">
      <input type="text" id="bajas-filtro" placeholder="🔍 Filtrar…"
             oninput="cargarTablaBajas(this.value)"
             style="padding:6px 10px;border:1.5px solid #e2e8f0;border-radius:6px;font-size:13px;width:200px">
      <button onclick="cargarTablaBajas(document.getElementById('bajas-filtro').value)" class="btn btn-ghost btn-sm">🔄</button>
    </div>
  </div>
  <div style="overflow-x:auto">
    <table style="width:100%;border-collapse:collapse;font-size:13px">
      <thead>
        <tr style="background:#f8fafc;border-bottom:2px solid #e2e8f0">
          <th style="padding:10px 12px;text-align:left;color:#475569">Código</th>
          <th style="padding:10px 12px;text-align:left;color:#475569">Descripción</th>
          <th style="padding:10px 12px;text-align:left;color:#475569">Estado original</th>
          <th style="padding:10px 12px;text-align:left;color:#475569">Operario</th>
          <th style="padding:10px 12px;text-align:left;color:#475569">Fecha/hora baja</th>
        </tr>
      </thead>
      <tbody id="bajas-tbody">
        <tr><td colspan="5" style="text-align:center;padding:18px;color:#64748b">Haz clic en "Ver Historial" para cargar</td></tr>
      </tbody>
    </table>
  </div>
</div>

<!-- ════ SECCIÓN AGENTE BAJAS EXCEL ════ -->
<div id="seccion-agente-setup" class="card" style="margin-top:24px">
  <div class="card-head" style="flex-wrap:wrap;gap:10px">
    <h2 class="card-title">📡 Agente Bajas Excel — Configuración del PC cliente</h2>
    <a href="/admin/descargar_agente_zip" class="btn btn-success btn-sm">⬇️ Descargar todo (ZIP)</a>
  </div>

  <p style="color:#475569;font-size:13px;margin:0 0 16px">
    El agente es un pequeño programa que se ejecuta en el ordenador que tiene acceso al Excel compartido.
    Escucha órdenes del servidor y procesa las bajas automáticamente cuando el admin lo indica desde este panel.
  </p>

  <!-- Pasos -->
  <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:14px;margin-bottom:20px">

    <div style="background:#f0fdf4;border:1.5px solid #86efac;border-radius:10px;padding:14px">
      <div style="font-weight:700;color:#166534;margin-bottom:6px">① Copiar archivos al PC cliente</div>
      <p style="font-size:12px;color:#15803d;margin:0 0 10px">Descarga el ZIP y extráelo en el escritorio del PC que tiene Excel. Crea una carpeta llamada <code>AgenteExcel</code>.</p>
      <div style="display:flex;flex-direction:column;gap:5px">
        <a href="/admin/descargar_agente_zip" class="btn btn-success btn-sm" style="justify-content:center">⬇️ Descargar ZIP completo</a>
        <div style="font-size:11px;color:#16a34a;text-align:center">— o descarga por separado —</div>
        <a href="/admin/descargar_agente/INSTALAR_AGENTE.bat" class="btn btn-ghost btn-sm" style="justify-content:center;font-size:11px">📄 INSTALAR_AGENTE.bat</a>
        <a href="/admin/descargar_agente/AGENTE_EXCEL.bat" class="btn btn-ghost btn-sm" style="justify-content:center;font-size:11px">📄 AGENTE_EXCEL.bat</a>
        <a href="/admin/descargar_agente/baja_excel_agente.py" class="btn btn-ghost btn-sm" style="justify-content:center;font-size:11px">📄 baja_excel_agente.py</a>
        <a href="/admin/descargar_agente/baja_excel.py" class="btn btn-ghost btn-sm" style="justify-content:center;font-size:11px">📄 baja_excel.py</a>
      </div>
    </div>

    <div style="background:#eff6ff;border:1.5px solid #93c5fd;border-radius:10px;padding:14px">
      <div style="font-weight:700;color:#1e40af;margin-bottom:6px">② Instalar Python y dependencias</div>
      <p style="font-size:12px;color:#1d4ed8;margin:0 0 6px">En el PC cliente, si Python no está instalado:</p>
      <ol style="font-size:12px;color:#1d4ed8;margin:0 0 8px;padding-left:18px;line-height:1.8">
        <li>Descarga Python desde <strong>python.org/downloads</strong></li>
        <li>Durante la instalación marca <strong>"Add Python to PATH"</strong></li>
        <li>Doble clic en <strong>INSTALAR_AGENTE.bat</strong> — instala pywin32 y requests automáticamente</li>
      </ol>
      <div style="background:#dbeafe;border-radius:6px;padding:8px;font-size:11px;color:#1e40af">
        ℹ️ Solo hay que hacer esto una vez por PC
      </div>
    </div>

    <div style="background:#fdf4ff;border:1.5px solid #d8b4fe;border-radius:10px;padding:14px">
      <div style="font-weight:700;color:#6b21a8;margin-bottom:6px">③ Arrancar el agente</div>
      <ol style="font-size:12px;color:#7e22ce;margin:0 0 8px;padding-left:18px;line-height:1.8">
        <li>Abre el Excel compartido con las <strong>macros habilitadas</strong></li>
        <li>Doble clic en <strong>AGENTE_EXCEL.bat</strong></li>
        <li>La primera vez te pide la <strong>IP del servidor</strong> y la <strong>contraseña admin</strong></li>
        <li>Se queda en espera — ya no hace falta repetir este paso hasta cerrar la ventana</li>
      </ol>
      <div style="background:#f3e8ff;border-radius:6px;padding:8px;font-size:11px;color:#6b21a8">
        💡 IP del servidor: <strong id="ip-servidor-texto">cargando…</strong>
      </div>
    </div>

    <div style="background:#fff7ed;border:1.5px solid #fdba74;border-radius:10px;padding:14px">
      <div style="font-weight:700;color:#9a3412;margin-bottom:6px">④ Usar desde este panel</div>
      <ol style="font-size:12px;color:#c2410c;margin:0 0 8px;padding-left:18px;line-height:1.8">
        <li>En el tile <strong>"📊 Procesar Bajas en Excel"</strong> verás el punto verde cuando el agente esté conectado</li>
        <li>Pulsa <strong>"📡 Enviar al PC cliente"</strong></li>
        <li>El agente procesa las bajas en Excel automáticamente</li>
        <li>El historial se actualiza solo en <strong>"✅ Dados de Baja"</strong></li>
      </ol>
      <div style="background:#ffedd5;border-radius:6px;padding:8px;font-size:11px;color:#9a3412">
        ⚠️ El Excel debe estar abierto y visible en el PC cliente mientras se procesa
      </div>
    </div>

  </div>

  <!-- Estado actual del agente -->
  <div style="background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:10px;padding:14px;display:flex;align-items:center;gap:12px;flex-wrap:wrap">
    <span id="agente-setup-badge" style="display:inline-block;width:12px;height:12px;border-radius:50%;background:#94a3b8;flex-shrink:0"></span>
    <span id="agente-setup-estado" style="font-size:13px;font-weight:600;color:#475569">Comprobando estado del agente…</span>
    <span id="agente-setup-ultimo" style="font-size:11px;color:#94a3b8;margin-left:auto"></span>
  </div>
</div>
"""

def tpl_estado():
    return """
<!doctype html><html><head><meta charset="utf-8"><title>Estado: {{estado}}</title>
<style>
body{font-family:Segoe UI,Roboto,Arial,sans-serif;margin:0;padding:20px;background:#f8f9fa}
.container{max-width:1400px;margin:0 auto;background:#fff;border-radius:12px;box-shadow:0 2px 10px rgba(0,0,0,.08);padding:20px}
h1{text-transform:capitalize}
table{width:100%;border-collapse:collapse;margin-top:10px}
th,td{padding:10px;border:1px solid #e9ecef}
th{background:#f8f9fa}
th:nth-child(5),td:nth-child(5){width:85px;white-space:nowrap;font-size:12px}
th:nth-child(1),td:nth-child(1){width:50px;text-align:center}
th:nth-child(2),td:nth-child(2){width:90px}
th:nth-child(3),td:nth-child(3){width:110px}
th:nth-child(8),td:nth-child(8){width:85px;white-space:nowrap;font-size:12px}
.row-green{background:#f6fff2}
.row-amber{background:#fffbea}
.row-red{background:#fff1f0}
.row-critical-red{background-color:#ffd6d6 !important}
.row-critical-amber{background-color:#fff3bf !important}
.small{color:#6c757d}
.alerta-exportacion{position:fixed;top:80px;right:20px;background:#fff3cd;border:1px solid #ffeaa7;border-left:4px solid #f0ad4e;border-radius:6px;padding:12px;box-shadow:0 2px 10px rgba(0,0,0,0.1);z-index:1000;max-width:350px;font-family:inherit;animation:slideInRight 0.3s ease-out}
.alerta-contenido{display:flex;align-items:flex-start;gap:10px}
.alerta-icono{font-size:20px;flex-shrink:0;margin-top:2px}
.alerta-texto{flex:1;font-size:14px;line-height:1.3;color:#8a6d3b}
.alerta-texto strong{color:#6b5429}
.alerta-botones{display:flex;gap:6px;flex-shrink:0;margin-top:8px}
.btn-alerta-admin,.btn-alerta-cerrar{padding:5px 12px;border:none;border-radius:4px;cursor:pointer;font-size:12px;font-weight:500;transition:all 0.15s ease}
.btn-alerta-admin{background:#007bff;color:white;box-shadow:0 1px 3px rgba(0,123,255,0.3)}
.btn-alerta-admin:hover{background:#0056b3;box-shadow:0 2px 5px rgba(0,123,255,0.4)}
.btn-alerta-cerrar{background:#6c757d;color:white;box-shadow:0 1px 3px rgba(108,117,125,0.3)}
.btn-alerta-cerrar:hover{background:#545b62;box-shadow:0 2px 5px rgba(108,117,125,0.4)}
@keyframes slideInRight{from{transform:translateX(100%);opacity:0}to{transform:translateX(0);opacity:1}}
@media (max-width: 768px){.alerta-exportacion{top:10px;right:10px;left:10px;max-width:none;font-size:13px}.alerta-contenido{flex-direction:column;gap:8px}.alerta-texto{font-size:13px}.btn-alerta-admin,.btn-alerta-cerrar{padding:8px 12px;font-size:13px}}
.nav-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;border-bottom:2px solid #e9ecef;padding-bottom:15px}
.btn-home{background:#007bff;color:#fff;text-decoration:none;padding:10px 20px;border-radius:8px;font-weight:bold}
.btn-home:hover{background:#0056b3;color:#fff}
.filters{margin:15px 0;display:flex;gap:10px;align-items:center}
.filters input{padding:8px 12px;border:1px solid #ddd;border-radius:6px}
.filters button{background:#28a745;color:#fff;border:none;padding:8px 16px;border-radius:6px;cursor:pointer}
</style></head><body>
<div class="container">
  <div class="nav-header">
    <h1>📋 Estado: {{estado}}</h1>
    <a href="{{ url_for('home') }}" class="btn-home">🏠 Volver al Inicio</a>
  </div>
  
  <div class="filters">
    <input type="text" id="f_q" placeholder="Buscar por código, EAN o descripción..." style="min-width:300px">
    <button id="btnFiltrar">🔍 Filtrar</button>
    <button onclick="location.reload()" style="background:#6c757d">🔄 Limpiar</button>
  </div>
  
  <p class="small">Atajos: F2 Registrar · F3 Asignar · F4 Devolver · F5 Gastado</p>
  <table id="tbl">
    <thead>
      <tr><th>ID</th><th>Código</th><th>EAN</th><th>Descripción</th><th>Caducidad</th><th>Estado</th><th>Operario</th><th>Asignado</th></tr>
    </thead>
    <tbody id="body"></tbody>
  </table>
  <div id="sentinel" style="height:24px"></div>
</div>
<script>
const estado = "{{ estado }}";
let offset=0, loading=false, done=false;

async function loadMore(){
  if(loading||done) return; loading=true;
  const q = document.getElementById('f_q')?.value || '';
  const url = `/api/materiales?estado=${encodeURIComponent(estado)}&q=${encodeURIComponent(q)}&offset=${offset}&limit=50`;
  const res = await fetch(url);
  const data = await res.json();
  if(data.length===0){ done=true; return; }
  const tb=document.getElementById('body');
  for(const m of data){
    const tr=document.createElement('tr');
    if(m.estado==='disponible') tr.className='row-green';
    if(m.estado==='vence prox') tr.className='row-amber';
    if(m.estado==='caducado') tr.className='row-red';
    
    // Sombreado especial para materiales en uso con problemas de fecha
    if(m.estado==='en uso' && m.estado_critico==='caducado') {
        tr.className='row-critical-red';
    } else if(m.estado==='en uso' && m.estado_critico==='vence prox') {
        tr.className='row-critical-amber';
    }
    
    tr.innerHTML=`<td>${m.id}</td><td>${m.codigo}</td><td>${m.ean}</td>
                  <td>${m.descripcion}</td><td>${m.caducidad}</td>
                  <td>${m.estado_html}</td><td>${m.operario}</td><td>${m.asignado_at}</td>`;
    tb.appendChild(tr);
  }
  offset+=data.length; loading=false;
}
const io=new IntersectionObserver((e)=>{ if(e[0].isIntersecting) loadMore(); });
io.observe(document.getElementById('sentinel'));
loadMore();

// Funcionalidad del filtro
document.getElementById('btnFiltrar').onclick=()=>{ 
  document.getElementById('body').innerHTML=''; 
  offset=0; 
  done=false; 
  loadMore(); 
};

// Filtrar con Enter
document.getElementById('f_q').addEventListener('keypress', function(e){
  if(e.key === 'Enter') {
    document.getElementById('btnFiltrar').click();
  }
});

// Atajos también aquí
document.addEventListener('keydown', function(e){
  const tag = (e.target.tagName || '').toLowerCase();
  if (['input','textarea','select'].includes(tag)) return;
  if (e.key === 'F2'){ e.preventDefault(); window.opener && window.opener.document.getElementById('openReg')?.click(); }
  if (e.key === 'F3'){ e.preventDefault(); window.opener && window.opener.document.getElementById('openAsig')?.click(); }
  if (e.key === 'F4'){ e.preventDefault(); window.opener && window.opener.document.getElementById('openDev')?.click(); }
  if (e.key === 'F5'){ e.preventDefault(); window.opener && window.opener.document.getElementById('openGas')?.click(); }
});

// ---- Auto-logout tras 10 minutos sin actividad ----
(function(){
  const IDLE_MS = 10 * 60 * 1000;
  let t = null;
  function resetTimer(){
    if (t) clearTimeout(t);
    t = setTimeout(()=>{ window.location.href = '/logout'; }, IDLE_MS);
  }
  ['click','mousemove','keydown','touchstart','scroll'].forEach(ev=>{
    window.addEventListener(ev, resetTimer, {passive:true});
  });
  resetTimer();
})();
</script>
</body></html>
"""

def tpl_home():
    return """
<!doctype html><html><head><meta charset="utf-8"><title>Gestión de Materiales</title>
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<link rel="manifest" href="/manifest.json">
<meta name="theme-color" content="#1a73e8">
<link rel="icon" type="image/png" href="/static/icons/icon-192.png">
<link rel="apple-touch-icon" href="/static/icons/icon-192.png">
<style>
:root{
  --bg:#f8f9fa; --card:#fff; --shadow:0 2px 10px rgba(0,0,0,.08);
  --btn:#1a73e8; --btnh:#1558b0; --ok:#2e7d32; --warn:#c39200; --err:#c62828;
}
*{box-sizing:border-box}
body{font-family:Segoe UI,Roboto,Arial,sans-serif;margin:0;padding:16px;background:var(--bg)}
.container{max-width:1200px;margin:0 auto;background:var(--card);border-radius:14px;box-shadow:var(--shadow);padding:16px}
h1{margin:8px 0 12px;text-align:center}
.tag{display:inline-block;padding:6px 10px;border-radius:999px;background:#eef3ff;margin:6px 6px 0 0}
.rolebar{display:flex;justify-content:space-between;align-items:center;margin:6px 0 10px;gap:10px}
a.btnlink{padding:14px 16px;border-radius:14px;text-decoration:none;background:#eef3ff}

/* ===== WIDGETS DE ESTADO MEJORADOS ===== */
.statebar{
  display:grid;
  grid-template-columns:repeat(auto-fit,minmax(160px,1fr));
  gap:12px;margin:20px 0 16px;
  /* Forzar que todos los widgets normales estén en la misma línea */
}

/* Mejoras responsive para widgets */
@media (max-width: 768px) {
  .statebar{
    grid-template-columns:repeat(auto-fit,minmax(140px,1fr));
    gap:8px
  }
  .widget-count{font-size:24px}
  .critical-count{font-size:32px;margin:0 8px}
  .widget-trend{font-size:9px;padding:3px 6px}
  .critical-trend{font-size:10px;padding:4px 8px}
}

@media (max-width: 1200px) and (min-width: 769px) {
  .statebar{
    grid-template-columns:repeat(auto-fit,minmax(150px,1fr));
    gap:10px
  }
}

/* Para pantallas grandes, mantener widgets en línea */
@media (min-width: 1201px) {
  .statebar{
    grid-template-columns:repeat(6,1fr); /* Forzar 6 columnas máximo */
    gap:14px
  }
}

/* Widget crítico centrado - Estados dinámicos */
.critical-widget-container{
  display:flex;justify-content:center;
  margin:35px 0 25px; /* Mucha más separación para destacar como elemento principal */
  transition:all 0.5s ease;
  position:relative
}
/* Separador visual sutil antes del widget crítico */
.critical-widget-container::before{
  content:'';position:absolute;top:-20px;left:50%;
  transform:translateX(-50%);width:60px;height:2px;
  background:linear-gradient(90deg,transparent,#f39c12,transparent);
  border-radius:1px
}

/* Estado CALMADO (sin caducados) */
.widget-critical{
  position:relative;min-width:280px;max-width:350px;
  background:linear-gradient(135deg,#fff3cd,#ffeaa7,#f1c40f);
  border:2px solid #f39c12;box-shadow:0 4px 15px rgba(241,196,15,0.2);
  transform:scale(1.02);color:#856404;text-align:center;
  transition:all 0.8s cubic-bezier(0.4,0,0.2,1);
  padding:16px /* Padding uniforme con trend arriba */
}
/* Widget crítico ya no necesita padding especial */
.widget-critical::before{
  background:linear-gradient(90deg,rgba(243,156,18,0.5),transparent);
  height:2px;transition:all 0.5s ease
}
.widget-critical:hover{
  transform:scale(1.04) translateY(-2px);
  box-shadow:0 6px 20px rgba(241,196,15,0.3)
}

/* Estado CRÍTICO (con caducados) */
.widget-critical.has-expired{
  background:linear-gradient(135deg,#ff6b6b,#ee5a52,#e74c3c) !important;
  border:3px solid #fff !important;
  box-shadow:0 8px 32px rgba(231,76,60,0.4) !important;
  transform:scale(1.05) !important;
  animation:critical-pulse 2s infinite ease-in-out !important;
  color:#fff !important
}
.widget-critical.has-expired::before{
  background:linear-gradient(90deg,#fff,rgba(255,255,255,0.5),#fff) !important;
  height:3px !important
}
.widget-critical.has-expired:hover{
  transform:scale(1.08) translateY(-6px) !important;
  box-shadow:0 12px 40px rgba(231,76,60,0.6) !important
}

/* Iconos dinámicos */
.critical-icon{
  font-size:36px;width:60px;height:60px;
  background:rgba(133,100,4,0.1);color:#856404;
  border:2px solid rgba(133,100,4,0.2);
  transition:all 0.5s ease
}
.widget-critical.has-expired .critical-icon{
  background:rgba(255,255,255,0.2) !important;
  color:#fff !important;
  border:2px solid rgba(255,255,255,0.3) !important;
  animation:shake 1s infinite ease-in-out !important
}

/* Contadores dinámicos - Máximo protagonismo */
.critical-count{
  font-size:44px;font-weight:900;color:#856404;
  text-shadow:0 3px 6px rgba(133,100,4,0.25);
  transition:all 0.5s ease;flex:1;text-align:center;
  margin:0 12px;line-height:1.1 /* Más compacto verticalmente */
}
.widget-critical.has-expired .critical-count{
  color:#fff !important;
  text-shadow:0 3px 6px rgba(0,0,0,0.3) !important;
  animation:bounce-count 1.5s infinite ease-in-out !important
}

/* Labels dinámicos - Menos prominencia que el contador */
.critical-label{
  font-size:14px;font-weight:600; /* Menos peso que antes */
  color:#856404;opacity:0.9; /* Menos prominencia */
  text-transform:uppercase;letter-spacing:0.8px;
  text-shadow:0 1px 2px rgba(133,100,4,0.15);
  transition:all 0.5s ease;margin-top:8px
}
.widget-critical.has-expired .critical-label{
  color:#fff !important;opacity:0.95 !important;
  text-shadow:0 2px 4px rgba(0,0,0,0.25) !important;
  font-weight:700 !important
}

/* Trends dinámicos críticos - Header superior */
.critical-trend{
  position:static; /* Ya no absoluto */
  background:rgba(133,100,4,0.1);color:#856404;
  font-size:11px;padding:5px 10px;border-radius:8px;
  font-weight:500;text-transform:uppercase;letter-spacing:0.5px;
  transition:all 0.5s ease;text-align:center;line-height:1.1;
  opacity:0.85;margin-bottom:10px /* Header arriba del contador */
}
.widget-critical.has-expired .critical-trend{
  background:rgba(255,255,255,0.9) !important;
  color:#c0392b !important;
  font-weight:600 !important;
  opacity:0.95 !important;
  box-shadow:0 1px 3px rgba(0,0,0,0.1) !important
}

/* Glow dinámico */
.critical-widget-container.has-expired{
  animation:pulse-glow 2s infinite ease-in-out
}
.widget{
  position:relative;padding:14px;border-radius:14px;text-decoration:none;
  background:linear-gradient(135deg,var(--bg-from),var(--bg-to));
  box-shadow:0 4px 20px var(--shadow-color);
  border:2px solid var(--border-color);
  transition:all 0.3s cubic-bezier(0.4,0,0.2,1);
  overflow:hidden;min-height:90px;display:block /* Más altura para acomodar el trend arriba */
}
.widget::before{
  content:'';position:absolute;top:0;left:0;right:0;height:4px;
  background:linear-gradient(90deg,var(--accent-1),var(--accent-2));
  transform:scaleX(0);transition:transform 0.3s ease
}
.widget:hover{
  transform:translateY(-4px) scale(1.02);
  box-shadow:0 8px 30px var(--shadow-hover);
}
.widget:hover::before{transform:scaleX(1)}
.widget-header{
  display:flex;align-items:center;justify-content:space-between;
  margin-bottom:12px;min-height:50px;
  position:relative /* Ya no necesita padding-right */
}
.widget-icon{
  font-size:28px;width:50px;height:50px;border-radius:12px;
  display:flex;align-items:center;justify-content:center;
  background:var(--icon-bg);color:var(--icon-color);
  animation:pulse 2s infinite;flex-shrink:0 /* No se encoge */
}
.widget-count{
  font-size:32px;font-weight:800;color:var(--text-primary);
  text-shadow:0 2px 4px var(--text-shadow);
  animation:countUp 0.8s ease-out;
  flex:1;text-align:center;margin:0 10px /* Centrado con margen */
}
.widget-label{
  font-size:14px;font-weight:600;color:var(--text-secondary);
  text-transform:uppercase;letter-spacing:0.5px;margin-top:8px
}
.widget-trend{
  position:static; /* Ya no es absoluto */
  font-size:10px;padding:4px 8px;border-radius:6px;
  background:var(--trend-bg);color:var(--trend-color);
  font-weight:500;text-align:center;line-height:1.2;
  margin-bottom:8px;text-transform:uppercase;letter-spacing:0.5px;
  opacity:0.85 /* Comentario sutil arriba */
}

/* Estados específicos con gradientes y colores */
.widget-red{
  --bg-from:#ffe1e6;--bg-to:#ffd6d6;--border-color:#ff9aa2;
  --text-primary:#8b0000;--text-secondary:#a0000f;
  --icon-bg:rgba(255,87,87,0.2);--icon-color:#ff5757;
  --shadow-color:rgba(255,87,87,0.3);--shadow-hover:rgba(255,87,87,0.4);
  --accent-1:#ff6b6b;--accent-2:#ee5a52;--text-shadow:rgba(255,87,87,0.3);
  --trend-bg:rgba(255,255,255,0.8);--trend-color:#d63031
}
.widget-blue{
  --bg-from:#e3f2fd;--bg-to:#d7e3ff;--border-color:#90caf9;
  --text-primary:#0d47a1;--text-secondary:#1565c0;
  --icon-bg:rgba(33,150,243,0.2);--icon-color:#2196f3;
  --shadow-color:rgba(33,150,243,0.3);--shadow-hover:rgba(33,150,243,0.4);
  --accent-1:#42a5f5;--accent-2:#1e88e5;--text-shadow:rgba(33,150,243,0.3);
  --trend-bg:rgba(255,255,255,0.8);--trend-color:#1565c0
}
.widget-amber{
  --bg-from:#fff8e1;--bg-to:#fff3bf;--border-color:#ffcc02;
  --text-primary:#e65100;--text-secondary:#f57c00;
  --icon-bg:rgba(255,193,7,0.2);--icon-color:#ffc107;
  --shadow-color:rgba(255,193,7,0.3);--shadow-hover:rgba(255,193,7,0.4);
  --accent-1:#ffca28;--accent-2:#ffa000;--text-shadow:rgba(255,193,7,0.3);
  --trend-bg:rgba(255,255,255,0.8);--trend-color:#e65100
}
.widget-green{
  --bg-from:#e8f5e8;--bg-to:#d8f5d0;--border-color:#81c784;
  --text-primary:#1b5e20;--text-secondary:#2e7d32;
  --icon-bg:rgba(76,175,80,0.2);--icon-color:#4caf50;
  --shadow-color:rgba(76,175,80,0.3);--shadow-hover:rgba(76,175,80,0.4);
  --accent-1:#66bb6a;--accent-2:#43a047;--text-shadow:rgba(76,175,80,0.3);
  --trend-bg:rgba(255,255,255,0.8);--trend-color:#2e7d32
}
.widget-cyan{
  --bg-from:#e0f2f1;--bg-to:#bee9f3;--border-color:#4dd0e1;
  --text-primary:#00695c;--text-secondary:#00796b;
  --icon-bg:rgba(0,188,212,0.2);--icon-color:#00bcd4;
  --shadow-color:rgba(0,188,212,0.3);--shadow-hover:rgba(0,188,212,0.4);
  --accent-1:#26c6da;--accent-2:#00acc1;--text-shadow:rgba(0,188,212,0.3);
  --trend-bg:rgba(255,255,255,0.8);--trend-color:#00796b
}
.widget-orange{
  --bg-from:#fff3e0;--bg-to:#ffeaa7;--border-color:#ffb74d;
  --text-primary:#e65100;--text-secondary:#f57c00;
  --icon-bg:rgba(255,152,0,0.2);--icon-color:#ff9800;
  --shadow-color:rgba(255,152,0,0.3);--shadow-hover:rgba(255,152,0,0.4);
  --accent-1:#ffb74d;--accent-2:#fb8c00;--text-shadow:rgba(255,152,0,0.3);
  --trend-bg:rgba(255,255,255,0.8);--trend-color:#e65100
}
.widget-gray{
  --bg-from:#f5f5f5;--bg-to:#e9ecef;--border-color:#bdbdbd;
  --text-primary:#424242;--text-secondary:#616161;
  --icon-bg:rgba(158,158,158,0.2);--icon-color:#9e9e9e;
  --shadow-color:rgba(158,158,158,0.3);--shadow-hover:rgba(158,158,158,0.4);
  --accent-1:#bdbdbd;--accent-2:#757575;--text-shadow:rgba(158,158,158,0.3);
  --trend-bg:rgba(255,255,255,0.8);--trend-color:#616161
}

/* ===== ALERTAS FLOTANTES ===== */
.alert-container{
  position:fixed;top:20px;right:20px;z-index:2000;
  max-width:400px;pointer-events:none
}
.floating-alert{
  background:white;border-radius:16px;padding:20px;margin-bottom:16px;
  box-shadow:0 10px 40px rgba(0,0,0,0.15);
  border-left:6px solid var(--alert-color);
  animation:slideInRight 0.5s cubic-bezier(0.4,0,0.2,1);
  pointer-events:auto;position:relative;overflow:hidden
}
.floating-alert::before{
  content:'';position:absolute;top:0;left:0;right:0;height:2px;
  background:linear-gradient(90deg,var(--alert-color),transparent);
  animation:progress 5s linear forwards
}
.alert-header{
  display:flex;align-items:center;gap:12px;margin-bottom:12px
}
.alert-icon{
  font-size:24px;width:40px;height:40px;border-radius:50%;
  display:flex;align-items:center;justify-content:center;
  background:var(--alert-bg);color:var(--alert-color);
  animation:bounce 1s ease-in-out infinite alternate
}
.alert-title{font-size:16px;font-weight:700;color:#2c3e50}
.alert-body{color:#546e7a;font-size:14px;line-height:1.4;margin-bottom:16px}
.alert-actions{display:flex;gap:8px}
.alert-btn{
  padding:8px 16px;border:none;border-radius:8px;font-size:12px;
  font-weight:600;cursor:pointer;transition:all 0.2s ease
}
.alert-btn-primary{background:var(--alert-color);color:white}
.alert-btn-secondary{background:#ecf0f1;color:#546e7a}
.alert-btn:hover{transform:translateY(-1px);box-shadow:0 4px 12px rgba(0,0,0,0.15)}
.alert-close{
  position:absolute;top:16px;right:16px;background:none;border:none;
  font-size:20px;color:#bdc3c7;cursor:pointer;
  width:24px;height:24px;border-radius:50%;
  display:flex;align-items:center;justify-content:center
}
.alert-close:hover{background:#ecf0f1;color:#7f8c8d}

/* Tipos de alerta */
.alert-critical{--alert-color:#e74c3c;--alert-bg:rgba(231,76,60,0.1)}
.alert-warning{--alert-color:#f39c12;--alert-bg:rgba(243,156,18,0.1)}
.alert-info{--alert-color:#3498db;--alert-bg:rgba(52,152,219,0.1)}
.alert-success{--alert-color:#27ae60;--alert-bg:rgba(39,174,96,0.1)}

/* Animaciones */
@keyframes slideInRight{from{transform:translateX(100%);opacity:0}to{transform:translateX(0);opacity:1}}
@keyframes pulse{0%{transform:scale(1)}50%{transform:scale(1.05)}100%{transform:scale(1)}}
@keyframes bounce{0%{transform:translateY(0)}100%{transform:translateY(-4px)}}
@keyframes countUp{from{transform:scale(0.8);opacity:0}to{transform:scale(1);opacity:1}}
@keyframes progress{from{transform:scaleX(1)}to{transform:scaleX(0)}}

/* Animaciones críticas */
@keyframes critical-pulse{
  0%{box-shadow:0 8px 32px rgba(231,76,60,0.4)}
  50%{box-shadow:0 12px 40px rgba(231,76,60,0.7)}
  100%{box-shadow:0 8px 32px rgba(231,76,60,0.4)}
}
@keyframes pulse-glow{
  0%{filter:drop-shadow(0 0 10px rgba(231,76,60,0.3))}
  50%{filter:drop-shadow(0 0 20px rgba(231,76,60,0.6))}
  100%{filter:drop-shadow(0 0 10px rgba(231,76,60,0.3))}
}
@keyframes shake{
  0%,100%{transform:translateX(0)}
  25%{transform:translateX(-2px)}
  75%{transform:translateX(2px)}
}
@keyframes bounce-count{
  0%,100%{transform:scale(1)}
  50%{transform:scale(1.1)}
}
@keyframes blink{
  0%,100%{opacity:1}
  50%{opacity:0.7}
}

/* Responsive */
@media (max-width: 768px){
  .statebar{grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:12px}
  .widget{padding:16px}
  .widget-icon{font-size:24px;width:40px;height:40px}
  .widget-count{font-size:24px}
  .alert-container{top:10px;right:10px;left:10px;max-width:none}
}
.row-critical-red{background-color:#ffd6d6 !important}
.row-critical-amber{background-color:#fff3bf !important}

/* Botonera principal */
.toolbar{display:flex;flex-wrap:wrap;gap:16px;justify-content:center;margin:22px 0 6px}
.btn{padding:16px 20px;border:0;border-radius:14px;font-size:18px;color:#fff;background:var(--btn);cursor:pointer;min-width:220px;position:relative}
.btn:hover{background:var(--btnh)}
.btn-ok{background:#2e7d32} .btn-warn{background:#d99500} .btn-err{background:#c62828}
.shortcut{position:absolute;right:10px;top:8px;font-size:12px;opacity:.7;color:#eaf1ff}

/* Mensajes + tabla */
.alert{padding:12px;border-radius:10px;margin:10px 0}
.alert-success{background:#d4edda;border:1px solid #c3e6cb;color:#155724}
.alert-error{background:#f8d7da;border:1px solid #f5c6cb;color:#721c24}
.alert-warning{background:#fff3cd;border:1px solid #ffeeba;color:#856404}
table{width:100%;border-collapse:collapse;margin-top:12px;table-layout:fixed}
th,td{padding:7px 10px;border:1px solid #e9ecef;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;vertical-align:middle}
th{background:#f8f9fa}
/* ID: oculto */
th:nth-child(1),td:nth-child(1){display:none}
/* Código */
th:nth-child(2),td:nth-child(2){width:88px}
/* EAN */
th:nth-child(3),td:nth-child(3){width:135px;font-size:13px;color:#546e7a}
/* Descripción: ancho fijo */
th:nth-child(4),td:nth-child(4){width:220px}
/* Caducidad */
th:nth-child(5),td:nth-child(5){width:88px;font-size:12px;color:#546e7a;text-align:center}
/* Estado */
th:nth-child(6),td:nth-child(6){width:108px;text-align:center}
/* Operario: más ancho para nombres completos */
th:nth-child(7),td:nth-child(7){width:auto}
/* Asignado: más ancho para la fecha completa */
th:nth-child(8),td:nth-child(8){width:155px;font-size:12px;color:#78909c;text-align:center}
.row-green{background:#f6fff2}
.row-amber{background:#fffbea}
.row-red{background:#fff1f0}

/* Operario clickable */
.op-link{background:none;border:none;color:#1565c0;cursor:pointer;font-size:inherit;padding:2px 4px;border-radius:4px;text-decoration:underline;text-align:left;display:inline-block}
.op-link:hover{background:#e3f2fd;color:#0d47a1}

/* Descripción clickable */
.desc-link{background:none;border:none;color:#2e7d32;cursor:pointer;font-size:inherit;padding:2px 4px;border-radius:4px;text-align:left;display:inline-block;max-width:100%;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.desc-link:hover{background:#e8f5e9;color:#1b5e20;text-decoration:underline}

/* Filtro activo pills */
.filtro-pills{display:flex;flex-wrap:wrap;gap:8px;margin-top:6px}
#filtro-op-pill,#filtro-desc-pill{display:none;align-items:center;gap:8px;border-radius:20px;padding:6px 14px;font-size:13px;font-weight:600;width:fit-content}
#filtro-op-pill{background:#e3f2fd;border:1px solid #90caf9;color:#1565c0}
#filtro-desc-pill{background:#e8f5e9;border:1px solid #a5d6a7;color:#2e7d32}
#filtro-op-pill button,#filtro-desc-pill button{background:none;border:none;font-size:16px;cursor:pointer;line-height:1;padding:0 2px}
#filtro-op-pill button:hover,#filtro-desc-pill button:hover{color:#c62828}


/* Modales */
.modal-backdrop{position:fixed;inset:0;background:rgba(0,0,0,.45);display:none;align-items:center;justify-content:center;z-index:1000}
.modal{background:#fff;border-radius:16px;box-shadow:var(--shadow);width:min(700px,95%);padding:16px}
.modal header{display:flex;justify-content:space-between;align-items:center;font-size:20px;font-weight:700}
.modal .close{border:0;background:transparent;font-size:28px;cursor:pointer}
.modal .row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin:10px 0}
.modal label{min-width:160px;font-size:16px}
.modal input{flex:1;padding:12px;border:1px solid #ced4da;border-radius:10px;font-size:18px}
.modal footer{display:flex;gap:10px;justify-content:flex-end;margin-top:12px}
.warntext{color:#c62828;font-weight:700;margin-top:6px;display:none}
</style></head><body>
<div class="container">
  <div class="rolebar" style="display:flex;align-items:center;gap:10px">
    <div>
      <span class="tag">Rol: <b>{{ role or 'operario' }}</b></span>
      <span class="tag" id="reloj">🕒 <span id="hora-servidor">Cargando...</span></span>
      <span class="tag" style="background:#4CAF50;color:white">✨ VERSIÓN MEJORADA v2.0 ✨</span>
      <a class="btnlink" href="{{ url_for('logout') }}">Cerrar sesión</a>
      <a class="btnlink" href="{{ url_for('admin') }}">⚙️ Admin</a>
      <span class="btnlink" style="background:#ccc;color:#666;padding:6px 12px;border-radius:8px;cursor:not-allowed;opacity:0.6;" title="Próximamente disponible">🔧 Herramientas</span>
    </div>
  </div>

  <!-- Widgets de estado mejorados -->
  <div class="statebar" id="statebar">
    <a class="widget widget-blue" href="{{ url_for('vista_estado', estado='en uso') }}">
      <div class="widget-trend" id="trend-uso">Activo</div>
      <div class="widget-header">
        <div class="widget-icon">👷</div>
        <div class="widget-count" id="cnt-uso">0</div>
      </div>
      <div class="widget-label">En Uso</div>
    </a>
    
    <a class="widget widget-amber" href="{{ url_for('vista_estado', estado='vence prox') }}">
      <div class="widget-trend" id="trend-prox">Atención</div>
      <div class="widget-header">
        <div class="widget-icon">⏰</div>
        <div class="widget-count" id="cnt-prox">0</div>
      </div>
      <div class="widget-label">Vence Pronto</div>
    </a>
    
    <a class="widget widget-green" href="{{ url_for('vista_estado', estado='disponible') }}">
      <div class="widget-trend" id="trend-dispo">OK</div>
      <div class="widget-header">
        <div class="widget-icon">✅</div>
        <div class="widget-count" id="cnt-dispo">0</div>
      </div>
      <div class="widget-label">Disponibles</div>
    </a>
    
    <a class="widget widget-cyan" href="{{ url_for('vista_estado', estado='precintado') }}">
      <div class="widget-trend" id="trend-pre">Nuevo</div>
      <div class="widget-header">
        <div class="widget-icon">📦</div>
        <div class="widget-count" id="cnt-pre">0</div>
      </div>
      <div class="widget-label">Precintados</div>
    </a>
    
    <a class="widget widget-orange" href="{{ url_for('vista_estado', estado='retirado') }}">
      <div class="widget-trend" id="trend-ret">Archivado</div>
      <div class="widget-header">
        <div class="widget-icon">📤</div>
        <div class="widget-count" id="cnt-ret">0</div>
      </div>
      <div class="widget-label">Retirados</div>
    </a>
    
    <a class="widget widget-gray" href="{{ url_for('vista_estado', estado='gastado') }}">
      <div class="widget-trend" id="trend-gas">Finalizado</div>
      <div class="widget-header">
        <div class="widget-icon">🗑️</div>
        <div class="widget-count" id="cnt-gas">0</div>
      </div>
      <div class="widget-label">Gastados</div>
    </a>
  </div>

  <!-- Widget CRÍTICO de Caducados - Destacado y centrado -->
  <div class="critical-widget-container" id="critical-container">
    <a class="widget widget-critical" id="critical-widget" href="{{ url_for('vista_estado', estado='caducado') }}">
      <div class="widget-trend critical-trend" id="trend-cad">Todo OK</div>
      <div class="widget-header">
        <div class="widget-icon critical-icon" id="critical-icon">✅</div>
        <div class="widget-count critical-count" id="cnt-cad">0</div>
      </div>
      <div class="widget-label critical-label" id="critical-label">MATERIALES CADUCADOS</div>
    </a>
  </div>

  <!-- Contenedor de alertas flotantes -->
  <div class="alert-container" id="alertContainer"></div>

  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}{% for cat,msg in messages %}
      {% if cat == 'error' %}
        <script>
          document.addEventListener('DOMContentLoaded', function() {
            mostrarDialogoError('{{ msg|e }}');
          });
        </script>
      {% else %}
        <div class="alert alert-{{ cat }}">{{ msg }}</div>
      {% endif %}
    {% endfor %}{% endif %}
  {% endwith %}

  <!-- Botones táctiles grandes con atajos -->
  <div class="toolbar">
    <button class="btn btn-ok" id="openReg">➕ Registrar <span class="shortcut">(F2)</span></button>
    <button class="btn" id="openAsig">👷 Asignar <span class="shortcut">(F3)</span></button>
    <button class="btn btn-ok" id="openDev">↩️ Devolver <span class="shortcut">(F4)</span></button>
    <button class="btn btn-warn" id="openRet">📤 Retirado <span class="shortcut">(F6)</span></button>
    <button class="btn btn-err" id="openGas">🗑️ Gastado <span class="shortcut">(F5)</span></button>
  </div>

  <!-- Filtros -->
  <form id="filtros" style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-top:8px">
    <label>Estado:</label>
    <select name="estado" id="f_estado">
      {% for opt in ["todos","precintado","disponible","vence prox","caducado","en uso","gastado"] %}
        <option value="{{opt}}">{{opt}}</option>
      {% endfor %}
    </select>
    <label>Operario:</label>
    <input type="text" id="f_operario" placeholder="Nº o nombre de operario" style="width:180px">
    <label>Buscar:</label>
    <input type="text" id="f_q" placeholder="Código, EAN o descripción" style="width:200px">
    <button type="button" class="btn" id="btnFiltrar">🔍 Filtrar</button>
    <button type="button" class="btn" id="btnLimpiar" style="background:#6c757d">✖ Limpiar</button>
  </form>
  <div class="filtro-pills">
    <div id="filtro-op-pill">👷 Operario: <strong id="filtro-op-texto"></strong><button onclick="limpiarFiltroOperario()" title="Quitar filtro">×</button></div>
    <div id="filtro-desc-pill">📦 Producto: <strong id="filtro-desc-texto"></strong><button onclick="limpiarFiltroDesc()" title="Quitar filtro">×</button></div>
  </div>

  <!-- Tabla con scroll infinito -->
  <table>
    <thead>
      <tr><th>ID</th><th>Código</th><th>EAN</th><th>Descripción</th><th>Caducidad</th><th>Estado</th><th>Operario</th><th>Asignado</th></tr>
    </thead>
    <tbody id="body"></tbody>
  </table>
  <div id="sentinel" style="height:26px"></div>
</div>

<!-- Modales -->
<div class="modal-backdrop" id="mb-reg">
  <div class="modal">
    <header>Registrar material <button class="close" data-close="mb-reg">×</button></header>
    <form method="POST" id="formReg">
      <input type="hidden" name="accion" value="registrar">
      <div class="row"><label for="rg_ean">EAN</label><input id="rg_ean" name="ean" placeholder="13 dígitos (opcional)" autofocus></div>
      <div class="row"><label for="rg_desc">Descripción</label><input id="rg_desc" name="descripcion" placeholder="se autocompleta por EAN si existe"></div>
      <div class="row"><label for="rg_cod">Código interno</label><input id="rg_cod" name="codigo" placeholder="7 dígitos" required></div>
      <div class="row"><label for="rg_cad">Caducidad</label><input id="rg_cad" name="caducidad" placeholder="ddmmaa o ddmmaaaa" required></div>
      <footer><button type="button" class="btn" data-close="mb-reg">Cancelar</button><button type="submit" class="btn btn-ok">Registrar</button></footer>
    </form>
  </div>
</div>

<div class="modal-backdrop" id="mb-asig">
  <div class="modal">
    <header>Asignar material <button class="close" data-close="mb-asig">×</button></header>
    <form method="POST" id="formAsig">
      <input type="hidden" name="accion" value="asignar_directo">
      <input type="hidden" name="confirmado" id="asig_conf" value="0">
      <div class="row"><label for="as_cod">Código interno</label><input id="as_cod" name="codigo" placeholder="7 dígitos" required></div>
      <div class="row"><label for="as_num">Nº operario</label><input id="as_num" name="operario_num" required></div>
      <div class="row"><label for="as_nom">Nombre</label><input id="as_nom" placeholder="(auto)" readonly></div>
      <div id="conflicto_msg" class="warntext">—</div>
      <footer><button type="button" class="btn" data-close="mb-asig">Cancelar</button><button type="submit" id="btnAsignarSubmit" class="btn btn-ok">Asignar</button></footer>
    </form>
  </div>
</div>

<div class="modal-backdrop" id="mb-dev">
  <div class="modal">
    <header>Devolver material <button class="close" data-close="mb-dev">×</button></header>
    <form method="POST" id="formDev">
      <input type="hidden" name="accion" value="devolver_rapido">
      <div class="row"><label for="dv_cod">Código interno</label><input id="dv_cod" name="codigo" placeholder="7 dígitos" required></div>
      <footer><button type="button" class="btn" data-close="mb-dev">Cancelar</button><button type="submit" class="btn btn-ok">Devolver</button></footer>
    </form>
  </div>
</div>

<div class="modal-backdrop" id="mb-gas">
  <div class="modal">
    <header>Marcar gastado <button class="close" data-close="mb-gas">×</button></header>
    <form method="POST" id="formGas">
      <input type="hidden" name="accion" value="gastado_rapido">
      <div class="row"><label for="gs_cod">Código interno</label><input id="gs_cod" name="codigo" placeholder="7 dígitos" required></div>
      <footer><button type="button" class="btn" data-close="mb-gas">Cancelar</button><button type="submit" class="btn btn-err">Gastado</button></footer>
    </form>
  </div>
</div>

<div class="modal-backdrop" id="mb-ret">
  <div class="modal">
    <header>Marcar retirado <button class="close" data-close="mb-ret">×</button></header>
    <form method="POST" id="formRet">
      <input type="hidden" name="accion" value="retirado_rapido">
      <div class="row"><label for="rt_cod">Código interno</label><input id="rt_cod" name="codigo" placeholder="7 dígitos" required></div>
      <footer><button type="button" class="btn" data-close="mb-ret">Cancelar</button><button type="submit" class="btn btn-warn">Retirado</button></footer>
    </form>
  </div>
</div>

<!-- Modal de Autenticación -->
<div class="modal-backdrop" id="mb-auth">
  <div class="modal">
    <header id="auth-title">🔐 Identificación requerida <button class="close" data-close="mb-auth">×</button></header>
    <div style="margin: 15px 0; padding: 12px; background: #e3f2fd; border-radius: 8px; font-size: 14px;">
      <strong>Operación:</strong> <span id="auth-operation"></span><br>
      <strong>Permisos necesarios:</strong> <span id="auth-required-role"></span>
    </div>
    <form id="formAuth">
      <div class="row"><label for="auth_numero">Nº Operario</label><input id="auth_numero" placeholder="Escanea o ingresa tu número" required autofocus></div>
      <footer>
        <button type="button" class="btn" data-close="mb-auth">Cancelar</button>
        <button type="submit" class="btn btn-ok">Verificar</button>
      </footer>
    </form>
  </div>
</div>


<script>
// ====== helpers modal ======
// ====== Reloj hora del servidor ======
async function actualizarReloj(){
  try {
    const res = await fetch('/api/hora_servidor');
    const data = await res.json();
    document.getElementById("hora-servidor").textContent = data.full;
  } catch (e) {
    document.getElementById("hora-servidor").textContent = "Error al obtener hora";
  }
}

// ====== Widgets y Alertas Mejoradas ======
async function actualizarWidgets(){
  try {
    const res = await fetch('/api/contadores');
    const data = await res.json();
    
    // Actualizar contadores en widgets
    const caducados = data.caducado || 0;
    document.getElementById("cnt-cad").textContent = caducados;
    document.getElementById("cnt-uso").textContent = data["en uso"] || 0;
    document.getElementById("cnt-prox").textContent = data["vence prox"] || 0;
    document.getElementById("cnt-dispo").textContent = data.disponible || 0;
    document.getElementById("cnt-pre").textContent = data.precintado || 0;
    document.getElementById("cnt-ret").textContent = data.retirado || 0;
    document.getElementById("cnt-gas").textContent = data.gastado || 0;
    
    // Actualizar trends dinámicamente
    const porcentajeUso = data.porcentaje_uso || 0;
    document.getElementById("trend-uso").textContent = `${porcentajeUso}%`;
    
    // ===== WIDGET CRÍTICO DINÁMICO =====
    const criticalWidget = document.getElementById("critical-widget");
    const criticalContainer = document.getElementById("critical-container");
    const criticalIcon = document.getElementById("critical-icon");
    const criticalTrend = document.getElementById("trend-cad");
    
    if (caducados > 0) {
      // ESTADO CRÍTICO - Hay caducados
      criticalWidget.classList.add("has-expired");
      criticalContainer.classList.add("has-expired");
      criticalIcon.textContent = "🚨";
      criticalTrend.textContent = "¡URGENTE!";
    } else {
      // ESTADO CALMADO - No hay caducados
      criticalWidget.classList.remove("has-expired");
      criticalContainer.classList.remove("has-expired");
      criticalIcon.textContent = "✅";
      criticalTrend.textContent = "Todo OK";
    }
    
    // Mostrar alertas flotantes si hay caducidades críticas
    mostrarAlertas(data.alertas);
    
  } catch (e) {
    console.error("Error actualizando widgets:", e);
  }
}

function mostrarAlertas(alertas) {
  const container = document.getElementById("alertContainer");
  
  // Limpiar alertas anteriores (solo las automáticas)
  const alertasAntiguas = container.querySelectorAll('.auto-alert');
  alertasAntiguas.forEach(alert => alert.remove());
  
  // Alertas de caducados críticos
  if (alertas.caducados_criticos && alertas.caducados_criticos.length > 0) {
    const totalCaducados = alertas.total_caducados;
    const primeros = alertas.caducados_criticos.slice(0, 3);
    
    const alertaHtml = `
      <div class="floating-alert alert-critical auto-alert">
        <button class="alert-close" onclick="this.parentElement.remove()">×</button>
        <div class="alert-header">
          <div class="alert-icon">🚨</div>
          <div class="alert-title">¡${totalCaducados} Materiales Caducados!</div>
        </div>
        <div class="alert-body">
          ${primeros.map(item => `
            <div style="margin:4px 0;padding:4px 8px;background:rgba(231,76,60,0.1);border-radius:4px;font-size:12px">
              <strong>${item.codigo}</strong> - ${item.descripcion.substring(0,30)}${item.descripcion.length > 30 ? '...' : ''}
              <br><span style="color:#e74c3c;font-weight:600">${item.dias_caducado} días caducado</span>
              ${item.operario ? ` • Asignado a: ${item.operario}` : ''}
            </div>
          `).join('')}
          ${totalCaducados > 3 ? `<div style="margin-top:8px;color:#7f8c8d;font-size:12px">...y ${totalCaducados - 3} más</div>` : ''}
        </div>
        <div class="alert-actions">
          <button class="alert-btn alert-btn-primary" onclick="window.location.href='/estado/caducado'">
            Ver Todos
          </button>
          <button class="alert-btn alert-btn-secondary" onclick="this.closest('.floating-alert').remove()">
            Cerrar
          </button>
        </div>
      </div>
    `;
    container.insertAdjacentHTML('beforeend', alertaHtml);
  }
  
  // Alertas de vencimientos de hoy
  if (alertas.vencen_hoy && alertas.vencen_hoy.length > 0) {
    const vencenHoy = alertas.vencen_hoy.slice(0, 2);
    
    const alertaHtml = `
      <div class="floating-alert alert-warning auto-alert">
        <button class="alert-close" onclick="this.parentElement.remove()">×</button>
        <div class="alert-header">
          <div class="alert-icon">⏰</div>
          <div class="alert-title">¡${alertas.total_vencen_hoy} Vencen HOY!</div>
        </div>
        <div class="alert-body">
          ${vencenHoy.map(item => `
            <div style="margin:4px 0;padding:4px 8px;background:rgba(243,156,18,0.1);border-radius:4px;font-size:12px">
              <strong>${item.codigo}</strong> - ${item.descripcion.substring(0,30)}${item.descripcion.length > 30 ? '...' : ''}
              ${item.operario ? `<br>Asignado a: ${item.operario}` : ''}
            </div>
          `).join('')}
          ${alertas.total_vencen_hoy > 2 ? `<div style="margin-top:8px;color:#7f8c8d;font-size:12px">...y ${alertas.total_vencen_hoy - 2} más</div>` : ''}
        </div>
        <div class="alert-actions">
          <button class="alert-btn alert-btn-primary" onclick="window.location.href='/estado/vence%20prox'">
            Revisar
          </button>
          <button class="alert-btn alert-btn-secondary" onclick="this.closest('.floating-alert').remove()">
            OK
          </button>
        </div>
      </div>
    `;
    container.insertAdjacentHTML('beforeend', alertaHtml);
  }
  
  // Alertas de vencimientos de mañana
  if (alertas.vencen_manana && alertas.vencen_manana.length > 0) {
    const alertaHtml = `
      <div class="floating-alert alert-info auto-alert">
        <button class="alert-close" onclick="this.parentElement.remove()">×</button>
        <div class="alert-header">
          <div class="alert-icon">📅</div>
          <div class="alert-title">${alertas.total_vencen_manana} Vencen Mañana</div>
        </div>
        <div class="alert-body">
          Hay materiales programados para vencer mañana. Planifica su uso o devolución.
        </div>
        <div class="alert-actions">
          <button class="alert-btn alert-btn-primary" onclick="window.location.href='/estado/vence%20prox'">
            Ver Lista
          </button>
          <button class="alert-btn alert-btn-secondary" onclick="this.closest('.floating-alert').remove()">
            Recordar
          </button>
        </div>
      </div>
    `;
    container.insertAdjacentHTML('beforeend', alertaHtml);
  }
  
  // Auto-ocultar alertas después de 15 segundos
  setTimeout(() => {
    const alertasAuto = container.querySelectorAll('.auto-alert');
    alertasAuto.forEach(alert => {
      if (alert.parentElement) {
        alert.style.animation = 'slideInRight 0.3s reverse';
        setTimeout(() => alert.remove(), 300);
      }
    });
  }, 15000);
}

// Inicializar sistema
actualizarReloj();
actualizarWidgets();
setInterval(actualizarReloj, 10000);
setInterval(actualizarWidgets, 30000); // Actualizar widgets cada 30s
function openModal(id){ document.getElementById(id).style.display='flex'; }
function closeModal(id){ document.getElementById(id).style.display='none'; }
document.querySelectorAll('.close,[data-close]').forEach(el=>{
  el.addEventListener('click', ()=>{ closeModal(el.getAttribute('data-close') || el.closest('.modal-backdrop').id); });
});
// Configuración de permisos por operación
const operationPermissions = {
  'registrar': { roles: ['almacenero', 'admin'], modal: 'mb-reg', focus: 'rg_ean' },
  'asignar': { roles: ['operario', 'almacenero', 'admin'], modal: 'mb-asig', focus: 'as_cod' },
  'devolver': { roles: ['almacenero', 'admin'], modal: 'mb-dev', focus: 'dv_cod' },
  'retirado': { roles: ['almacenero', 'admin'], modal: 'mb-ret', focus: 'rt_cod' },
  'gastado': { roles: ['almacenero', 'admin'], modal: 'mb-gas', focus: 'gs_cod' }
};

// Variable global para la operación pendiente
let pendingOperation = null;
let currentUser = null;

// Función para verificar autenticación
function requireAuth(operation) {
  const config = operationPermissions[operation];
  if (!config) return false;
  
  // Mostrar modal de autenticación
  document.getElementById('auth-operation').textContent = operation.charAt(0).toUpperCase() + operation.slice(1);
  document.getElementById('auth-required-role').textContent = config.roles.join(', ');
  pendingOperation = operation;
  openModal('mb-auth');
  setTimeout(() => document.getElementById('auth_numero').focus(), 20);
  return false;
}

// Event listeners para botones principales
document.getElementById('openReg').onclick = () => requireAuth('registrar');
document.getElementById('openAsig').onclick = () => requireAuth('asignar');
document.getElementById('openDev').onclick = () => requireAuth('devolver');
document.getElementById('openRet').onclick = () => requireAuth('retirado');
document.getElementById('openGas').onclick = () => requireAuth('gastado');

// ====== Atajos de teclado globales ======
document.addEventListener('keydown', function(e){
  const tag = (e.target.tagName || '').toLowerCase();
  if (['input','textarea','select'].includes(tag)) return;
  if (e.key === 'F2'){ e.preventDefault(); document.getElementById('openReg').click(); }
  if (e.key === 'F3'){ e.preventDefault(); document.getElementById('openAsig').click(); }
  if (e.key === 'F4'){ e.preventDefault(); document.getElementById('openDev').click(); }
  if (e.key === 'F5'){ e.preventDefault(); document.getElementById('openGas').click(); }
  if (e.key === 'F6'){ e.preventDefault(); document.getElementById('openRet').click(); }
});

// ====== Manejo de autenticación ======
document.getElementById('formAuth').addEventListener('submit', async function(e) {
  e.preventDefault();
  
  const numero = document.getElementById('auth_numero').value.trim();
  
  if (!numero) {
    alert('Escanea o ingresa tu número de operario');
    return;
  }
  
  try {
    const response = await fetch('/api/auth', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ numero: numero })
    });
    
    const data = await response.json();
    
    if (data.success) {
      currentUser = data.user;
      const config = operationPermissions[pendingOperation];
      
      if (config && config.roles.includes(data.user.rol)) {
        closeModal('mb-auth');
        openModal(config.modal);
        setTimeout(() => {
          const focusElement = document.getElementById(config.focus);
          if (focusElement) focusElement.focus();
        }, 100);
      } else {
        alert(`No tienes permisos para esta operación. Se requiere: ${config.roles.join(', ')}`);
      }
    } else {
      alert('Número de operario no válido');
    }
  } catch (error) {
    alert('Error de conexión');
  }
  
  document.getElementById('auth_numero').value = '';
});

function mostrarAlertaCaducados(totalCaducados) {
    let alertaDiv = document.getElementById('alerta-caducados');
    if (!alertaDiv) {
        alertaDiv = document.createElement('div');
        alertaDiv.id = 'alerta-caducados';
        alertaDiv.className = 'alerta-exportacion';
        alertaDiv.innerHTML = `
            <div class="alerta-contenido">
                <span class="alerta-icono">⚠️</span>
                <div style="flex: 1;">
                    <div class="alerta-texto">
                        <strong>Productos Caducados Detectados</strong><br>
                        Hay <span id="total-caducados">${totalCaducados}</span> productos caducados (incluyendo retirados/gastados).<br>
                        <small>Se recomienda exportar los datos y actualizar el archivo de red.</small>
                    </div>
                    <div class="alerta-botones">
                        <button onclick="irAAdmin()" class="btn-alerta-admin">📊 Ir a Admin</button>
                        <button onclick="ocultarAlertaCaducados()" class="btn-alerta-cerrar">×</button>
                    </div>
                </div>
            </div>
        `;
        document.body.insertBefore(alertaDiv, document.body.firstChild);
    } else {
        document.getElementById('total-caducados').textContent = totalCaducados;
        alertaDiv.style.display = 'block';
    }
}

function ocultarAlertaCaducados() {
    const alertaDiv = document.getElementById('alerta-caducados');
    if (alertaDiv) {
        alertaDiv.style.display = 'none';
    }
}

function mostrarDialogoError(mensaje) {
    const overlay = document.createElement('div');
    overlay.id = 'error-dialog-overlay';
    overlay.style.cssText = `
        position: fixed;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        background: rgba(0,0,0,0.5);
        z-index: 10000;
        display: flex;
        align-items: center;
        justify-content: center;
    `;
    
    const dialog = document.createElement('div');
    dialog.style.cssText = `
        background: white;
        padding: 20px;
        border-radius: 10px;
        box-shadow: 0 4px 8px rgba(0,0,0,0.3);
        max-width: 400px;
        text-align: center;
        font-family: inherit;
    `;
    
    dialog.innerHTML = `
        <div style="color: #721c24; font-size: 18px; margin-bottom: 15px;">
            ❌ Error
        </div>
        <div style="color: #721c24; margin-bottom: 20px; line-height: 1.4;">
            ${mensaje}
        </div>
        <button id="error-dialog-btn" onclick="cerrarDialogoError()" style="
            background: #dc3545;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 5px;
            cursor: pointer;
            font-size: 14px;
        ">Aceptar</button>
    `;
    
    overlay.appendChild(dialog);
    document.body.appendChild(overlay);
    
    // Enfocar el botón y añadir evento de teclado
    const button = document.getElementById('error-dialog-btn');
    button.focus();
    
    // Cerrar con Enter o Escape
    const handleKeydown = (e) => {
        if (e.key === 'Enter' || e.key === 'Escape') {
            cerrarDialogoError();
            document.removeEventListener('keydown', handleKeydown);
        }
    };
    
    document.addEventListener('keydown', handleKeydown);
}

function cerrarDialogoError() {
    const overlay = document.getElementById('error-dialog-overlay');
    if (overlay) {
        overlay.remove();
    }
}

function irAAdmin() {
    window.location.href = '/admin';
}

async function loadCounters(){
  try{
    const r = await fetch('/api/contadores');
    const j = await r.json();
    document.getElementById('cnt-cad').textContent  = j["caducado"] ?? 0;
    document.getElementById('cnt-uso').textContent  = j["en uso"] ?? 0;
    document.getElementById('cnt-prox').textContent = j["vence prox"] ?? 0;
    document.getElementById('cnt-dispo').textContent= j["disponible"] ?? 0;
    document.getElementById('cnt-pre').textContent  = j["precintado"] ?? 0;
    document.getElementById('cnt-ret').textContent  = j["retirado"] ?? 0;
    document.getElementById('cnt-gas').textContent  = j["gastado"] ?? 0;
    
    // Mostrar alerta de productos caducados para exportación
    if (j.alerta_caducados && j.total_caducados > 0) {
        mostrarAlertaCaducados(j.total_caducados);
    } else {
        ocultarAlertaCaducados();
    }
  }catch(e){}
}
loadCounters();

// ====== Registro (con chequeo de duplicados) ======
function nextOnEnter(id,nextId,before){ const el=document.getElementById(id); if(!el) return;
  el.addEventListener('keydown', ev=>{ if(ev.key==='Enter'){ ev.preventDefault(); if(typeof before==='function') before(); const nx=document.getElementById(nextId); if(nx) nx.focus(); }});
}
async function checkEAN(){
  const ean=document.getElementById('rg_ean').value.trim();
  const desc=document.getElementById('rg_desc');
  if(!ean) return;
  if(!/^\\d{13}$/.test(ean)){ alert('EAN debe tener 13 dígitos'); return; }
  try{ 
    const r=await fetch('/api/get_descripcion_by_ean?ean='+encodeURIComponent(ean)); 
    const j=await r.json(); 
    if(j.descripcion){ 
      desc.value=j.descripcion;
      desc.setAttribute('readonly', 'true');
      document.getElementById('rg_cod').focus();
    } else {
      desc.removeAttribute('readonly');
      desc.focus();
    }
  }catch(e){}
}
async function checkCodigoDuplicado(){
  const cod = document.getElementById('rg_cod').value.trim();
  if(!/^\\d{7}$/.test(cod)) return;
  try{
    const r = await fetch('/api/check_codigo?codigo=' + encodeURIComponent(cod));
    const j = await r.json();
    if(j.existe){
      alert(`El código ${cod} ya existe. No se puede registrar.`);
      closeModal('mb-reg');
      window.location.href = '/';
    }
  }catch(e){}
}
document.getElementById('rg_ean').addEventListener('blur', checkEAN);
nextOnEnter('rg_ean','rg_desc', checkEAN);
nextOnEnter('rg_desc','rg_cod');
nextOnEnter('rg_cod','rg_cad', checkCodigoDuplicado);
nextOnEnter('rg_cad','', function() {
  document.getElementById('formReg').requestSubmit();
});
document.getElementById('rg_desc').addEventListener('keydown', e=>{ if(e.key==='Enter'){ e.preventDefault(); document.getElementById('formReg').requestSubmit(); }});
document.getElementById('rg_cod').addEventListener('blur', checkCodigoDuplicado);
document.getElementById('formReg').addEventListener('submit', async function(e){
  const cod=document.getElementById('rg_cod').value.trim();
  const ean=document.getElementById('rg_ean').value.trim();
  if(!/^\\d{7}$/.test(cod)){ e.preventDefault(); alert('Código interno = 7 dígitos'); return; }
  if(ean && !/^\\d{13}$/.test(ean)){ e.preventDefault(); alert('EAN debe tener 13 dígitos'); return; }
  try{
    const r = await fetch('/api/check_codigo?codigo=' + encodeURIComponent(cod));
    const j = await r.json();
    if(j.existe){
      e.preventDefault();
      alert(`El código ${cod} ya existe. No se puede registrar.`);
      closeModal('mb-reg');
      window.location.href = '/';
      return;
    }
  }catch(e){
    e.preventDefault();
    alert('Error al verificar el código.');
    return;
  }
  setTimeout(loadCounters, 500);
});

// ====== Asignación (conflicto inmediato) ======
const asNum = document.getElementById('as_num');
const asCod = document.getElementById('as_cod');
const asNom = document.getElementById('as_nom');
const asBtn = document.getElementById('btnAsignarSubmit');
const conflictoMsg = document.getElementById('conflicto_msg');

async function fillOperarioName(){
  const num=asNum.value.trim();
  asNom.value='';
  if(num){
    try{
      const r=await fetch('/api/operario_nombre?numero='+encodeURIComponent(num));
      const j=await r.json();
      if(j.nombre){ asNom.value=j.nombre; }
      else{
        if(confirm('Operario no existe. ¿Darlo de alta ahora?')){
          const nombre = prompt('Nombre del operario:','');
          if(nombre && nombre.trim()){
            const fd=new FormData(); fd.append('numero',num); fd.append('nombre',nombre.trim());
            const rr=await fetch('/api/operario_add',{method:'POST', body:fd});
            const jj=await rr.json();
            if(jj.ok){ asNom.value=nombre.trim(); alert('Operario guardado.'); }
            else alert('No se pudo guardar operario.');
          }
        }
      }
    }catch(e){}
  }
  await checkConflict();
}
async function checkConflict(){
  conflictoMsg.style.display='none';
  asBtn.disabled=false;
  const num=asNum.value.trim();
  const cod=asCod.value.trim();
  if(!/^\\d{7}$/.test(cod) || !num) return;
  try{
    const r=await fetch(`/api/operario_conflicto_ean?codigo=${encodeURIComponent(cod)}&operario_num=${encodeURIComponent(num)}`);
    const j=await r.json();
    if(j.ok && j.conflicto){
      conflictoMsg.textContent = `No puedes asignarte este producto: ya tienes otro con el mismo EAN (${j.ean}). Devuélvelo primero (código ${j.otro_codigo}${j.otra_desc? ' - '+j.otra_desc : ''}).`;
      conflictoMsg.style.display='block';
      asBtn.disabled = true;
    }
  }catch(e){}
}
asCod.addEventListener('keydown', ev=>{ if(ev.key==='Enter'){ ev.preventDefault(); asNum.focus(); } });
asNum.addEventListener('keydown', ev=>{ if(ev.key==='Enter'){ ev.preventDefault(); asNom.focus(); } });
asNum.addEventListener('blur', fillOperarioName);
asNum.addEventListener('input', ()=>{ conflictoMsg.style.display='none'; asBtn.disabled=false; });
asCod.addEventListener('blur', checkConflict);
asCod.addEventListener('input', ()=>{ conflictoMsg.style.display='none'; asBtn.disabled=false; });

document.getElementById('formAsig').addEventListener('submit', async function(e){
  const codigo=asCod.value.trim();
  const num=asNum.value.trim();
  if(!/^\\d{7}$/.test(codigo)){ e.preventDefault(); alert('Código interno = 7 dígitos'); return; }
  if(!num){ e.preventDefault(); alert('Nº de operario obligatorio'); return; }
  if(asBtn.disabled){ e.preventDefault(); alert('No puedes asignarte este producto hasta devolver el otro con el mismo EAN.'); return; }
  e.preventDefault();
  try{
    const r=await fetch('/api/info_material?codigo='+encodeURIComponent(codigo)); const j=await r.json();
    if(!j.existe){ alert('El código no existe. Regístralo primero.'); return; }
    if(j.estado==='gastado'){ alert('No se puede asignar: material gastado.'); return; }
    if(j.caducado){ alert('No se puede asignar: material CADUCADO.'); return; }
    if(j.vence_prox){
      if(!confirm('Atención: vence pronto ('+j.caducidad+'). ¿Asignar igualmente?')) return;
      document.getElementById('asig_conf').value='1';
    }
  }catch(e){ alert('No se pudo comprobar el estado.'); return; }
  this.submit();
  setTimeout(loadCounters, 500);
});

// ====== Scroll infinito ======
let offset=0, loading=false, done=false;
const bodyT=document.getElementById('body');
const estadoSel=document.getElementById('f_estado');
const qInp=document.getElementById('f_q');
const opInp=document.getElementById('f_operario');

async function loadMore(){
  if(loading||done) return; loading=true;
  const res=await fetch(`/api/materiales?estado=${encodeURIComponent(estadoSel.value)}&q=${encodeURIComponent(qInp.value)}&operario=${encodeURIComponent(opInp.value)}&offset=${offset}&limit=50`);
  const data=await res.json();
  if(data.length===0){ done=true; loading=false; return; }
  for(const m of data){
    const tr=document.createElement('tr');
    if(m.estado==='disponible') tr.className='row-green';
    if(m.estado==='vence prox') tr.className='row-amber';
    if(m.estado==='caducado') tr.className='row-red';
    
    // Sombreado especial para materiales en uso con problemas de fecha
    if(m.estado==='en uso' && m.estado_critico==='caducado') {
        tr.className='row-critical-red';
    } else if(m.estado==='en uso' && m.estado_critico==='vence prox') {
        tr.className='row-critical-amber';
    }
    
    // Celda de operario: clickable si tiene operario asignado
    let opCell;
    if(m.operario_numero){
      opCell=`<button class="op-link" data-num="${m.operario_numero}" data-display="${m.operario.replace(/"/g,'&quot;')}">${m.operario}</button>`;
    } else {
      opCell=m.operario;
    }
    // Celda de descripción: siempre clickable
    const descEsc=m.descripcion.replace(/"/g,'&quot;');
    const descCell=`<button class="desc-link" data-q="${descEsc}" title="Ver todos con este producto">${m.descripcion}</button>`;
    
    tr.innerHTML=`<td>${m.id}</td><td>${m.codigo}</td><td>${m.ean}</td><td>${descCell}</td><td>${m.caducidad}</td><td>${m.estado_html}</td><td>${opCell}</td><td>${m.asignado_at}</td>`;
    bodyT.appendChild(tr);
  }
  offset+=data.length; loading=false;
}

// Click en operario → filtrar tabla directamente
bodyT.addEventListener('click', function(e){
  const btnOp=e.target.closest('.op-link');
  if(btnOp){
    opInp.value=btnOp.dataset.num;
    mostrarPillOperario(btnOp.dataset.display);
    bodyT.innerHTML=''; offset=0; done=false; loadMore();
    return;
  }
  const btnDesc=e.target.closest('.desc-link');
  if(btnDesc){
    qInp.value=btnDesc.dataset.q;
    mostrarPillDesc(btnDesc.dataset.q);
    bodyT.innerHTML=''; offset=0; done=false; loadMore();
  }
});

function mostrarPillOperario(texto){
  document.getElementById('filtro-op-texto').textContent=texto;
  document.getElementById('filtro-op-pill').style.display='flex';
}
function limpiarFiltroOperario(){
  opInp.value='';
  document.getElementById('filtro-op-pill').style.display='none';
  bodyT.innerHTML=''; offset=0; done=false; loadMore();
}
function mostrarPillDesc(texto){
  document.getElementById('filtro-desc-texto').textContent=texto;
  document.getElementById('filtro-desc-pill').style.display='flex';
}
function limpiarFiltroDesc(){
  qInp.value='';
  document.getElementById('filtro-desc-pill').style.display='none';
  bodyT.innerHTML=''; offset=0; done=false; loadMore();
}

const io=new IntersectionObserver((e)=>{ if(e[0].isIntersecting) loadMore(); });
io.observe(document.getElementById('sentinel'));
loadMore();
document.getElementById('btnFiltrar').onclick=()=>{ if(!opInp.value) document.getElementById('filtro-op-pill').style.display='none'; if(!qInp.value) document.getElementById('filtro-desc-pill').style.display='none'; bodyT.innerHTML=''; offset=0; done=false; loadMore(); };
document.getElementById('btnLimpiar').onclick=()=>{ estadoSel.value='todos'; qInp.value=''; opInp.value=''; document.getElementById('filtro-op-pill').style.display='none'; document.getElementById('filtro-desc-pill').style.display='none'; bodyT.innerHTML=''; offset=0; done=false; loadMore(); };

document.addEventListener('click', e=>{
  if(e.target.classList.contains('modal-backdrop')) e.target.style.display='none';
});

// ---- Auto-logout tras 10 minutos sin actividad ----
(function(){
  const IDLE_MS = 10 * 60 * 1000;
  let t = null;
  function resetTimer(){
    if (t) clearTimeout(t);
    t = setTimeout(()=>{ window.location.href = '/logout'; }, IDLE_MS);
  }
  ['click','mousemove','keydown','touchstart','scroll'].forEach(ev=>{
    window.addEventListener(ev, resetTimer, {passive:true});
  });
  resetTimer();
})();

// ---- Service Worker (PWA install prompt) ----
if ('serviceWorker' in navigator) {
  window.addEventListener('load', () => {
    navigator.serviceWorker.register('/sw.js').catch(() => {});
  });
}
</script>
</body></html>
"""

@app.get("/api/operario/<numero>/materiales")
def api_operario_materiales(numero):
    """Retorna todos los materiales actualmente asignados a un operario."""
    with get_db() as conn:
        c = conn.cursor()
        c.execute(
            "SELECT id,codigo,caducidad,estado,operario_numero,ean,descripcion,fecha_asignacion "
            "FROM materiales WHERE operario_numero=?", (numero,)
        )
        rows = [row_to_material(r) for r in c.fetchall()]
    datos = []
    for m in rows:
        base = estado_base(m.caducidad, m.operario_numero, m.estado)
        if base in ("gastado", "retirado", "escaneado"):
            continue
        label = estado_label(m.caducidad, m.operario_numero, m.estado)
        datos.append({
            "id": m.id,
            "codigo": m.codigo,
            "ean": m.ean or "-",
            "descripcion": m.descripcion or "-",
            "caducidad": m.caducidad,
            "estado": base,
            "estado_html": badge_html(label),
        })
    nombre = get_operario_nombre(numero) or numero
    return jsonify({"materiales": datos, "numero": numero, "nombre": nombre})

@app.get("/api/hora_servidor")
def api_hora_servidor():
    ahora = datetime.now()
    dias = ["lunes","martes","miércoles","jueves","viernes","sábado","domingo"]
    meses = ["enero","febrero","marzo","abril","mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"]
    dia = dias[ahora.weekday()]
    mes = meses[ahora.month - 1]
    texto = f"{dia.capitalize()}, {ahora.day} de {mes} de {ahora.year} · {ahora.strftime('%H:%M:%S')}"
    return jsonify({
        "fecha": ahora.strftime("%Y-%m-%d"),
        "hora": ahora.strftime("%H:%M:%S"),
        "full": texto
    })


# ================== Bajas pendientes Excel ==================
def _ensure_procesado_excel_col():
    """Migración: añade la columna procesado_excel, crea la tabla bajas y limpia registros huérfanos."""
    with get_db_materiales() as conn:
        try:
            conn.execute("ALTER TABLE materiales ADD COLUMN procesado_excel INTEGER DEFAULT 0")
        except Exception:
            pass  # Ya existe
        conn.execute(
            """CREATE TABLE IF NOT EXISTS bajas (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo          TEXT,
                descripcion     TEXT,
                estado_original TEXT,
                operario_numero TEXT,
                fecha_baja      TEXT DEFAULT (datetime('now','localtime'))
            )"""
        )
        # Limpiar materiales ya procesados en Excel que no se borraron (registros huérfanos)
        conn.execute(
            "DELETE FROM materiales WHERE procesado_excel = 1 AND estado IN ('gastado', 'retirado')"
        )

@app.get("/api/bajas_pendientes_excel")
def api_bajas_pendientes_excel():
    """Devuelve materiales gastados/retirados no procesados en Excel. Solo admin."""
    if current_role() != "admin":
        return jsonify({"success": False}), 403
    _ensure_procesado_excel_col()
    with get_db_materiales() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """SELECT id, codigo, descripcion, estado, operario_numero, fecha_asignacion
               FROM materiales
               WHERE estado IN ('gastado','retirado')
                 AND (procesado_excel IS NULL OR procesado_excel = 0)
               ORDER BY fecha_asignacion ASC"""
        ).fetchall()
    return jsonify({"pendientes": [dict(r) for r in rows]})

@app.post("/api/marcar_procesado_excel/<int:mat_id>")
def api_marcar_procesado_excel(mat_id):
    """Marca un material como procesado en Excel. Solo admin."""
    if current_role() != "admin":
        return jsonify({"success": False}), 403
    _ensure_procesado_excel_col()
    with get_db_materiales() as conn:
        conn.execute("UPDATE materiales SET procesado_excel = 1 WHERE id = ?", (mat_id,))
    return jsonify({"success": True})

@app.post("/api/marcar_procesado_excel_bulk")
def api_marcar_procesado_excel_bulk():
    """Marca varios materiales como procesados en Excel. Solo admin."""
    if current_role() != "admin":
        return jsonify({"success": False}), 403
    ids = request.json.get("ids", [])
    if not ids:
        return jsonify({"success": False, "mensaje": "Sin IDs"}), 400
    _ensure_procesado_excel_col()
    with get_db_materiales() as conn:
        conn.executemany(
            "UPDATE materiales SET procesado_excel = 1 WHERE id = ?",
            [(i,) for i in ids]
        )
    return jsonify({"success": True, "procesados": len(ids)})

@app.get("/api/bajas")
def api_bajas():
    """Devuelve el historial de materiales dados de baja. Solo admin."""
    if current_role() != "admin":
        return jsonify({"success": False}), 403
    _ensure_procesado_excel_col()
    with get_db_materiales() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """SELECT id, codigo, descripcion, estado_original, operario_numero, fecha_baja
               FROM bajas
               ORDER BY fecha_baja DESC"""
        ).fetchall()
    return jsonify({"bajas": [dict(r) for r in rows], "total": len(rows)})

# ================== Agente Cliente Excel ==================
def _ensure_solicitud_cliente_table():
    with get_db_materiales() as conn:
        conn.execute(
            """CREATE TABLE IF NOT EXISTS solicitud_excel_cliente (
                id INTEGER PRIMARY KEY,
                estado TEXT DEFAULT 'idle',
                solicitada_en TEXT,
                completada_en TEXT,
                salida TEXT,
                ultimo_poll_agente TEXT
            )"""
        )
        conn.execute(
            "INSERT OR IGNORE INTO solicitud_excel_cliente (id, estado) VALUES (1, 'idle')"
        )

def _check_agent_token():
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return False
    token = auth[7:]
    # Acepta la contraseña admin
    if token == ADMIN_PASSWORD:
        return True
    # Acepta también el número de operario con rol admin
    op = get_operario_by_numero(token)
    return op is not None and op.get("rol") == "admin"

@app.get("/api/admin/estado_solicitud_cliente")
def api_estado_solicitud_cliente():
    """Estado actual de la solicitud al agente. Solo admin."""
    if current_role() != "admin":
        return jsonify({"success": False}), 403
    _ensure_solicitud_cliente_table()
    with get_db_materiales() as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM solicitud_excel_cliente WHERE id=1").fetchone()
    if not row:
        return jsonify({"estado": "idle", "agente_online": False})
    row = dict(row)
    agente_online = False
    if row.get("ultimo_poll_agente"):
        try:
            from datetime import datetime as _dt
            dt = _dt.fromisoformat(row["ultimo_poll_agente"])
            agente_online = (_dt.now() - dt).total_seconds() < 15
        except Exception:
            pass
    return jsonify({
        "estado": row["estado"],
        "salida": row["salida"],
        "solicitada_en": row["solicitada_en"],
        "completada_en": row["completada_en"],
        "agente_online": agente_online,
    })

@app.post("/api/admin/solicitar_bajas_cliente")
def api_solicitar_bajas_cliente():
    """Admin solicita al agente que procese las bajas. Solo admin."""
    if current_role() != "admin":
        return jsonify({"success": False}), 403
    _ensure_solicitud_cliente_table()
    with get_db_materiales() as conn:
        row = conn.execute("SELECT estado FROM solicitud_excel_cliente WHERE id=1").fetchone()
        if row and row[0] in ("pendiente", "procesando"):
            return jsonify({"success": False, "mensaje": "Ya hay una solicitud en curso."})
        conn.execute(
            """UPDATE solicitud_excel_cliente
               SET estado='pendiente', solicitada_en=datetime('now','localtime'),
                   completada_en=NULL, salida=NULL
               WHERE id=1"""
        )
    return jsonify({"success": True})

@app.post("/api/admin/cancelar_solicitud_cliente")
def api_cancelar_solicitud_cliente():
    """Admin cancela la solicitud (pendiente o procesando). Solo admin."""
    if current_role() != "admin":
        return jsonify({"success": False}), 403
    _ensure_solicitud_cliente_table()
    with get_db_materiales() as conn:
        conn.execute(
            "UPDATE solicitud_excel_cliente SET estado='cancelado', salida='Detenido por el admin' WHERE id=1"
        )
    return jsonify({"success": True})

@app.get("/api/agente/cancelado")
def api_agente_cancelado():
    """El agente comprueba si su trabajo fue cancelado. Auth: Bearer."""
    if not _check_agent_token():
        return jsonify({"error": "Token inválido"}), 401
    _ensure_solicitud_cliente_table()
    with get_db_materiales() as conn:
        row = conn.execute("SELECT estado FROM solicitud_excel_cliente WHERE id=1").fetchone()
    cancelado = row and row[0] == "cancelado"
    return jsonify({"cancelado": cancelado})

@app.get("/api/agente/poll")
def api_agente_poll():
    """El agente consulta si hay solicitud pendiente. Auth: Bearer <admin_password>."""
    if not _check_agent_token():
        return jsonify({"error": "Token inválido"}), 401
    _ensure_solicitud_cliente_table()
    with get_db_materiales() as conn:
        conn.execute(
            "UPDATE solicitud_excel_cliente SET ultimo_poll_agente=datetime('now','localtime') WHERE id=1"
        )
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT estado FROM solicitud_excel_cliente WHERE id=1").fetchone()
    estado = row["estado"] if row else "idle"
    return jsonify({"hay_solicitud": estado == "pendiente"})

@app.get("/api/agente/pendientes")
def api_agente_pendientes():
    """El agente descarga los materiales pendientes. Auth: Bearer."""
    if not _check_agent_token():
        return jsonify({"error": "Token inválido"}), 401
    _ensure_procesado_excel_col()
    with get_db_materiales() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """SELECT id, codigo, descripcion, estado, operario_numero
               FROM materiales
               WHERE estado IN ('gastado','retirado')
                 AND (procesado_excel IS NULL OR procesado_excel = 0)
               ORDER BY fecha_asignacion ASC"""
        ).fetchall()
    return jsonify({"pendientes": [dict(r) for r in rows]})

@app.post("/api/agente/iniciar")
def api_agente_iniciar():
    """El agente marca la solicitud como 'procesando'. Auth: Bearer."""
    if not _check_agent_token():
        return jsonify({"error": "Token inválido"}), 401
    _ensure_solicitud_cliente_table()
    with get_db_materiales() as conn:
        conn.execute("UPDATE solicitud_excel_cliente SET estado='procesando' WHERE id=1")
    return jsonify({"success": True})

@app.post("/api/agente/marcar_uno/<int:mat_id>")
def api_agente_marcar_uno(mat_id):
    """El agente reporta un material procesado: lo registra en bajas y lo elimina de materiales."""
    if not _check_agent_token():
        return jsonify({"error": "Token inválido"}), 401
    _ensure_procesado_excel_col()
    with get_db_materiales() as conn:
        row = conn.execute(
            "SELECT codigo, descripcion, estado, operario_numero FROM materiales WHERE id=?",
            (mat_id,)
        ).fetchone()
        if not row:
            return jsonify({"success": False, "mensaje": "Material no encontrado"}), 404
        conn.execute(
            """INSERT INTO bajas (codigo, descripcion, estado_original, operario_numero, fecha_baja)
               VALUES (?, ?, ?, ?, datetime('now','localtime'))""",
            (row[0], row[1], row[2], row[3])
        )
        conn.execute("DELETE FROM materiales WHERE id=?", (mat_id,))
    return jsonify({"success": True})

@app.post("/api/agente/completar")
def api_agente_completar():
    """El agente reporta la finalización. Los materiales ya fueron marcados por marcar_uno."""
    if not _check_agent_token():
        return jsonify({"error": "Token inválido"}), 401
    data = request.json or {}
    salida = data.get("salida", "")
    _ensure_solicitud_cliente_table()
    with get_db_materiales() as conn:
        conn.execute(
            """UPDATE solicitud_excel_cliente
               SET estado='completado', completada_en=datetime('now','localtime'), salida=?
               WHERE id=1""",
            (salida,)
        )
    return jsonify({"success": True})

@app.post("/api/agente/error")
def api_agente_error():
    """El agente reporta un error. Auth: Bearer."""
    if not _check_agent_token():
        return jsonify({"error": "Token inválido"}), 401
    mensaje = (request.json or {}).get("mensaje", "Error desconocido")
    _ensure_solicitud_cliente_table()
    with get_db_materiales() as conn:
        conn.execute(
            """UPDATE solicitud_excel_cliente
               SET estado='error', completada_en=datetime('now','localtime'), salida=?
               WHERE id=1""",
            (mensaje,)
        )
    return jsonify({"success": True})

# ================== Descarga de archivos del agente ==================
_ARCHIVOS_AGENTE = {
    "baja_excel_agente.py": "baja_excel_agente.py",
    "baja_excel.py":        "baja_excel.py",
    "AGENTE_EXCEL.bat":     "AGENTE_EXCEL.bat",
    "INSTALAR_AGENTE.bat":  "INSTALAR_AGENTE.bat",
}

@app.get("/admin/descargar_agente/<nombre>")
def descargar_archivo_agente(nombre):
    """Descarga uno de los archivos del agente. Solo admin."""
    if current_role() != "admin":
        abort(403)
    if nombre not in _ARCHIVOS_AGENTE:
        abort(404)
    ruta = os.path.join(BASE_DIR, _ARCHIVOS_AGENTE[nombre])
    if not os.path.isfile(ruta):
        abort(404)
    return send_file(ruta, as_attachment=True, download_name=nombre)

@app.get("/admin/descargar_agente_zip")
def descargar_agente_zip():
    """Descarga todos los archivos del agente en un ZIP. Solo admin."""
    if current_role() != "admin":
        abort(403)
    import zipfile, io
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for nombre, rel in _ARCHIVOS_AGENTE.items():
            ruta = os.path.join(BASE_DIR, rel)
            if os.path.isfile(ruta):
                zf.write(ruta, nombre)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="agente_bajas_excel.zip",
                     mimetype="application/zip")

@app.post("/api/admin/limpiar_procesados_excel")
def api_limpiar_procesados_excel():
    """Elimina de materiales los registros ya marcados como procesado_excel=1. Solo admin."""
    if current_role() != "admin":
        return jsonify({"success": False}), 403
    _ensure_procesado_excel_col()
    with get_db_materiales() as conn:
        cur = conn.execute(
            "DELETE FROM materiales WHERE procesado_excel = 1 AND estado IN ('gastado', 'retirado')"
        )
        eliminados = cur.rowcount
    return jsonify({"success": True, "eliminados": eliminados})

@app.post("/api/admin/ejecutar_bajas_excel")
def api_ejecutar_bajas_excel():
    """Ejecuta baja_excel.py en modo automático. Solo admin."""
    if current_role() != "admin":
        return jsonify({"success": False, "salida": "Acceso denegado"}), 403
    import subprocess, sys as _sys
    script = os.path.join(BASE_DIR, "baja_excel.py")
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    try:
        r = subprocess.run(
            [_sys.executable, script],
            cwd=BASE_DIR,
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=180, env=env
        )
        salida = r.stdout.strip() or "(sin salida)"
        if r.stderr.strip():
            salida += "\n[stderr]\n" + r.stderr.strip()
        return jsonify({"success": r.returncode == 0, "salida": salida})
    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "salida": "Tiempo de espera agotado (180 s)."}), 500
    except Exception as e:
        return jsonify({"success": False, "salida": str(e)}), 500

# ================== Actualización desde GitHub ==================
@app.post("/api/admin/update")
def api_admin_update():
    """Ejecuta git pull + pip install. Solo administradores."""
    if current_role() != "admin":
        return jsonify({"success": False, "mensaje": "Acceso denegado"}), 403

    import subprocess, sys as _sys
    output_lines = []

    # git fetch + reset (descarta cambios locales y sincroniza con remoto)
    try:
        # 1. Descargar cambios sin aplicar
        rf = subprocess.run(
            ["git", "fetch", "origin", "main"],
            cwd=BASE_DIR,
            capture_output=True, text=True, timeout=60
        )
        output_lines.append("── git fetch ──")
        output_lines.append(rf.stdout.strip() or "(sin salida)")
        if rf.stderr.strip():
            output_lines.append(rf.stderr.strip())

        # 2. Ver si hay diferencias entre local y remoto
        diff = subprocess.run(
            ["git", "diff", "HEAD", "FETCH_HEAD", "--name-only"],
            cwd=BASE_DIR, capture_output=True, text=True
        )
        hubo_cambios = bool(diff.stdout.strip())

        # 3. Forzar actualización descartando cualquier cambio local
        r = subprocess.run(
            ["git", "reset", "--hard", "FETCH_HEAD"],
            cwd=BASE_DIR,
            capture_output=True, text=True, timeout=30
        )
        output_lines.append("── git reset --hard FETCH_HEAD ──")
        output_lines.append(r.stdout.strip() or "(sin salida)")
        if r.stderr.strip():
            output_lines.append(r.stderr.strip())
    except FileNotFoundError:
        return jsonify({"success": False, "mensaje": "Git no está instalado o no está en el PATH."}), 500
    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "mensaje": "Tiempo de espera agotado al contactar GitHub."}), 500

    # ── Validación sintáctica de app.py ──────────────────────────
    app_py = os.path.join(BASE_DIR, "app.py")
    output_lines.append("")
    output_lines.append("── validación sintáctica ──")
    chk = subprocess.run(
        [_sys.executable, "-m", "py_compile", app_py],
        capture_output=True, text=True
    )
    if chk.returncode != 0:
        # Revertir solo app.py a la versión anterior
        subprocess.run(
            ["git", "checkout", "HEAD~1", "--", "app.py"],
            cwd=BASE_DIR, capture_output=True
        )
        error_msg = (chk.stderr.strip() or "Error de sintaxis desconocido")
        output_lines.append(f"❌ Error de sintaxis detectado — app.py REVERTIDO")
        output_lines.append(error_msg)
        return jsonify({
            "success": False,
            "hubo_cambios": hubo_cambios,
            "mensaje": f"Error de sintaxis en app.py. Revertido automáticamente.\n{error_msg}",
            "output": "\n".join(output_lines)
        }), 200
    output_lines.append("✅ app.py sin errores de sintaxis")

    # pip install si hubo cambios en requirements
    output_lines.append("")
    output_lines.append("── pip install ──")
    r2 = subprocess.run(
        [_sys.executable, "-m", "pip", "install", "-r",
         os.path.join(BASE_DIR, "requirements.txt"), "--quiet"],
        cwd=BASE_DIR,
        capture_output=True, text=True, timeout=120
    )
    output_lines.append(r2.stdout.strip() or "Sin cambios en dependencias.")
    if r2.stderr.strip():
        output_lines.append(r2.stderr.strip())

    return jsonify({
        "success": True,
        "hubo_cambios": hubo_cambios,
        "output": "\n".join(output_lines)
    })


@app.post("/api/admin/restart")
def api_admin_restart():
    """Reinicia el proceso de la aplicación. Solo administradores."""
    if current_role() != "admin":
        return jsonify({"success": False, "mensaje": "Acceso denegado"}), 403

    # Señalamos a run_app_window.py que cierre la ventana de forma limpia.
    # El hilo monitor en run_app_window.py llama window.destroy() tras 1.5s,
    # webview.start() retorna, y main() hace sys.exit(42).
    # start.bat detecta el código 42 y relanza el proceso sin conflicto de puerto.
    def _set_flag():
        import time
        time.sleep(1.5)
        _restart_event.set()

    threading.Thread(target=_set_flag, daemon=True).start()
    return jsonify({"success": True, "mensaje": "Reiniciando en 1.5 segundos…"})


# ================== Run ==================
if __name__ == "__main__":
    init_db()
    # Configuración para red local - permite acceso desde otros PCs
    # Para desarrollo local usar: host='127.0.0.1'
    # Para red local usar: host='0.0.0.0'
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    
    import sys as _sys_enc
    _out = open(_sys_enc.stdout.fileno(), mode='w', encoding='utf-8', buffering=1, closefd=False)
    _out.write(f"\n{'='*60}\n")
    _out.write(f"APLICACION DE GESTION DE MATERIALES INICIADA\n")
    _out.write(f"{'='*60}\n")
    _out.write(f"Servidor ejecutandose en:\n")
    _out.write(f"   Local: http://127.0.0.1:5000\n")
    _out.write(f"   Red:   http://{local_ip}:5000\n")
    _out.flush()
# ================== API CRUD Operarios ==================
@app.route("/api/operarios", methods=["GET", "POST"])
def api_operarios():
    """API para gestión de operarios"""
    if current_role() != "admin":
        return jsonify({"error": "Acceso denegado"}), 403
    
    if request.method == "GET":
        # Listar todos los operarios
        operarios = get_all_operarios()
        # Agregar estadísticas a cada operario
        for op in operarios:
            stats = get_estadisticas_operario(op['numero'])
            op.update(stats)
        return jsonify({"operarios": operarios})
    
    elif request.method == "POST":
        # Crear nuevo operario
        data = request.json or {}
        numero = data.get('numero', '').strip()
        nombre = data.get('nombre', '').strip()
        rol = data.get('rol', 'operario').strip()
        
        success, mensaje = crear_operario(numero, nombre, rol)
        return jsonify({"success": success, "mensaje": mensaje})

@app.route("/api/operarios/<numero>", methods=["GET", "PUT", "DELETE"])
def api_operario_individual(numero):
    """API para operario individual"""
    if current_role() != "admin":
        return jsonify({"error": "Acceso denegado"}), 403
    
    if request.method == "GET":
        # Obtener operario específico
        operario = get_operario_completo(numero)
        if not operario:
            return jsonify({"error": "Operario no encontrado"}), 404
        
        # Agregar estadísticas
        stats = get_estadisticas_operario(numero)
        operario.update(stats)
        return jsonify(operario)
    
    elif request.method == "PUT":
        # Actualizar operario
        data = request.json or {}
        nombre = data.get('nombre', '').strip()
        rol = data.get('rol', '').strip()
        
        success, mensaje = actualizar_operario(numero, nombre, rol)
        return jsonify({"success": success, "mensaje": mensaje})
    
    elif request.method == "DELETE":
        # Eliminar (desactivar) operario
        success, mensaje = eliminar_operario(numero)
        return jsonify({"success": success, "mensaje": mensaje})

@app.route("/api/operarios/<numero>/toggle", methods=["POST"])
def api_toggle_operario(numero):
    """API para activar/desactivar operario"""
    if current_role() != "admin":
        return jsonify({"error": "Acceso denegado"}), 403
    
    success, mensaje = toggle_operario_activo(numero)
    return jsonify({"success": success, "mensaje": mensaje})

if __name__ == "__main__":
    import sys as _sys2, socket as _sock2
    _o2 = open(_sys2.stdout.fileno(), mode='w', encoding='utf-8', buffering=1, closefd=False)
    _o2.write(f"{'='*60}\n")
    _o2.write(f"INICIANDO SERVIDOR DE GESTION DE MATERIALES\n")
    _o2.write(f"{'='*60}\n")
    _o2.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    _hn2 = _sock2.gethostname()
    _ip2 = _sock2.gethostbyname(_hn2)
    _o2.write(f"Servidor en: http://localhost:5000  |  http://{_ip2}:5000\n")
    _o2.write(f"{'='*60}\n\n")
    _o2.flush()

    app.run(host='0.0.0.0', port=5000, debug=False)
